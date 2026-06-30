"""Ingest raw layer data from the workbook (validation mode), CSV, or SQL."""
from __future__ import annotations
import datetime as dt
import re
import pandas as pd
from openpyxl import load_workbook

# Raw columns as they appear in Input Data B:AA (the SQL paste area)
RAW_COLS = {
    2: "program_id", 3: "layer_id", 5: "coverage", 6: "months_on_risk",
    7: "spacecraft_id", 8: "launch_date", 9: "inception", 10: "expiry",
    11: "orbit", 12: "on_risk_date", 13: "off_risk_date",
    14: "bus_manufacturer", 15: "prime_manufacturer",
    19: "program_type", 20: "underwriting_status", 21: "program_name",
    22: "entity", 23: "mapping_code", 24: "spacecraft_name",
    25: "bus_family", 26: "vehicle_family", 27: "layer_signed_exposure",
    32: "is_consortium",            # AF — consortium participation flag
}

# Workbook computed columns, captured only for reconciliation
WB_COMPUTED = {
    16: "wb_rpf", 17: "wb_on_risk_flag", 29: "wb_months_left",
    30: "wb_leo_debris_rpf", 31: "wb_debris_dr",
    33: "wb_equity_pct", 34: "wb_equity_usd", 35: "wb_layer_occ",
    36: "wb_exposure_usd", 37: "wb_per_sc",
    51: "wb_ext_qs", 54: "wb_net_ext_qs", 56: "wb_s3123_eligible",
    57: "wb_s3123_qs", 58: "wb_net_ext", 59: "wb_igr_qs_rate",
    60: "wb_igr_qs_ceded", 61: "wb_net_of_qs", 62: "wb_xol_ceded",
    64: "wb_net_of_xol",
    77: "wb_pf_fihl", 78: "wb_pf_ful", 79: "wb_pf_fiid", 81: "wb_pf_fibl_ceded",
    87: "wb_sw_fihl", 88: "wb_sw_ful", 89: "wb_sw_fiid", 91: "wb_sw_fibl_ceded",
    96: "wb_gd_fihl", 97: "wb_gd_ful", 98: "wb_gd_fiid", 100: "wb_gd_fibl_ceded",
    103: "wb_sd_fihl", 104: "wb_sd_ful", 105: "wb_sd_fiid", 107: "wb_sd_fibl_ceded",
    72: "wb_maxrisk_fihl", 73: "wb_maxrisk_net",
}


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s or s.upper() in ("NULL", "N/A", "NA", "-", "TBC"):
        return None
    try:
        return dt.date.fromisoformat(s[:10])
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d %b %Y", "%d-%b-%Y",
                "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None   # unparseable date string -> blank (surfaces in data quality)


def load_from_workbook(path: str, sheet: str = "Input Data",
                       first_row: int = 20, with_wb_values: bool = True) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    cols = dict(RAW_COLS)
    if with_wb_values:
        cols.update(WB_COMPUTED)
    max_col = max(cols.keys())

    records = []
    for row in ws.iter_rows(min_row=first_row, max_col=max_col, values_only=True):
        pid = row[1]  # col B (0-indexed 1)
        if pid is None:
            continue
        rec = {name: row[c - 1] for c, name in cols.items()}
        records.append(rec)
    wb.close()

    df = pd.DataFrame(records)
    for col in ["launch_date", "inception", "expiry", "on_risk_date", "off_risk_date"]:
        df[col] = df[col].map(_to_date)
    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


def load_from_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["launch_date", "inception", "expiry",
                                        "on_risk_date", "off_risk_date"])
    for col in ["launch_date", "inception", "expiry", "on_risk_date", "off_risk_date"]:
        df[col] = df[col].dt.date
    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


# ---------------- SQL ingest ----------------

_SQL_TARGETS = {
    "programid": "program_id", "layerid": "layer_id", "coverage": "coverage",
    "monthsonrisk": "months_on_risk", "spacecraftid": "spacecraft_id",
    "launchdate": "launch_date", "inception": "inception", "expiry": "expiry",
    "orbitcategory": "orbit", "orbit": "orbit",
    "onriskdate": "on_risk_date", "offriskdate": "off_risk_date",
    "busmanufacturer": "bus_manufacturer", "primemanufacturer": "prime_manufacturer",
    "programtype": "program_type", "underwritingstatus": "underwriting_status",
    "programname": "program_name", "entity": "entity", "mappingcode": "mapping_code",
    "spacecraftname": "spacecraft_name", "busfamily": "bus_family",
    "vehiclefamily": "vehicle_family",
    "layersignedexposure": "layer_signed_exposure",
    "layersignedexposureusd": "layer_signed_exposure",
    "isconsortium": "is_consortium",
    "action": "override_action",   # 'Add Layer' / 'Remove Layer' from manual table
    "placingbasis": "placing_basis",
    # View's own computed columns — captured as sql_* for engine-vs-view QA
    "qsigr": "sql_igr_qs_rate",
    "monthsleftonrisk": "sql_months_left",
    "externalqs": "sql_ext_qs",
    "netofexternalqs": "sql_net_ext_qs",
    "qsfiblceded": "sql_igr_qs_ceded",
    "netofqsigr": "sql_net_of_qs",
    "xolfiblceded": "sql_xol_ceded",
    "totalfiblceded": "sql_total_fibl_ceded",
    "netofxoligr": "sql_net_of_xol",
}
_SQL_REQUIRED = {"program_id", "layer_id", "spacecraft_id", "inception",
                 "off_risk_date", "orbit", "bus_manufacturer", "entity",
                 "mapping_code", "spacecraft_name", "layer_signed_exposure"}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _pick_driver():
    import pyodbc
    for want in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server",
                 "SQL Server Native Client 11.0", "SQL Server"):
        if want in pyodbc.drivers():
            return want
    raise RuntimeError("No SQL Server ODBC driver installed")


def load_from_sql(sql_cfg: dict, non_consortium: set | None = None) -> pd.DataFrame:
    """Load from the SQL view with automatic column-name mapping.

    Config block (config/<quarter>.yaml):
        ingest:
          source: sql
          non_consortium_spacecraft: [17163, 17950]
          sql:
            server: LON-SQLP-V005
            database: SpaceTrax_Data
            view: rds.vw_SpaceRDS_OnRisk
    """
    import pyodbc
    driver = sql_cfg.get("driver") or _pick_driver()
    conn = (f"DRIVER={{{driver}}};SERVER={sql_cfg['server']};"
            f"DATABASE={sql_cfg['database']};Trusted_Connection=yes;")
    if "18" in driver:
        conn += "TrustServerCertificate=yes;"
    cn = pyodbc.connect(conn, timeout=30)
    # Custom query takes precedence (supports {as_at} placeholder for
    # historical/as-at runs); query_file loads it from a .sql file.
    query = sql_cfg.get("query")
    if not query and sql_cfg.get("query_file"):
        from pathlib import Path as _P
        query = _P(sql_cfg["query_file"]).read_text()
    if query:
        query = query.format(as_at=sql_cfg.get("as_at", ""))
    else:
        query = f"SELECT * FROM {sql_cfg['view']}"
    raw = pd.read_sql(query, cn)
    cn.close()

    rename = {}
    for col in raw.columns:
        target = _SQL_TARGETS.get(_norm(col))
        if target and target not in rename.values():
            rename[col] = target
    df = raw.rename(columns=rename)

    missing = _SQL_REQUIRED - set(df.columns)
    if missing:
        raise ValueError(
            f"SQL view missing required fields: {sorted(missing)}. "
            f"View columns were: {list(raw.columns)}")

    for col in ["launch_date", "inception", "expiry", "on_risk_date", "off_risk_date"]:
        if col in df.columns:
            df[col] = df[col].map(_to_date)

    # Consortium flag: from the view if present, else config exceptions list
    if "is_consortium" in df.columns:
        df["is_consortium"] = df["is_consortium"].map(
            lambda v: str(v).strip().lower() in ("1", "true", "yes", "y"))
    else:
        nc = non_consortium or set()
        df["is_consortium"] = ~df["spacecraft_id"].isin(nc)

    # Optional descriptive fields the view doesn't carry
    for col in ["program_type", "underwriting_status"]:
        if col not in df.columns:
            df[col] = None

    df["layer_signed_exposure"] = pd.to_numeric(df["layer_signed_exposure"],
                                                errors="coerce")
    df = df[df["layer_signed_exposure"].notna()].reset_index(drop=True)

    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


def _apply_corrections(df, params):
    """Apply per-layer data corrections from config (e.g. wrong off-risk dates,
    restated exposures). Each correction:
        {program_id, layer_id, field, value, reason}
    Logged to load.last_corrections for the Data Corrections sheet. Date fields
    are parsed; per_sc/exposure corrections also sync layer_signed_exposure so
    the value flows through every downstream calc."""
    corrections = params.raw.get("data_corrections") or []
    applied = []
    df = df.copy()
    df["_pid"] = df["program_id"].astype(str)
    df["_lid"] = df["layer_id"].astype(str)
    DATE_FIELDS = ("off_risk_date", "on_risk_date", "inception", "expiry",
                   "launch_date")
    for c in corrections:
        mask = (df["_pid"] == str(c["program_id"])) & \
               (df["_lid"] == str(c["layer_id"]))
        n = int(mask.sum())
        if not n:
            applied.append({**c, "status": "NOT FOUND", "old": None})
            continue
        field = c["field"]
        old = df.loc[mask, field].iloc[0] if field in df.columns else None
        val = c["value"]
        if field in DATE_FIELDS:
            val = _to_date(val)
        df.loc[mask, field] = val
        # keep the raw signed-exposure column in sync with a per_sc correction
        if field in ("per_sc", "layer_signed_exposure"):
            for sync in ("per_sc", "layer_signed_exposure"):
                if sync in df.columns:
                    df.loc[mask, sync] = val
        applied.append({**c, "status": "APPLIED", "old": old, "rows": n})
    df = df.drop(columns=["_pid", "_lid"])
    load.last_corrections = applied
    return df


def load(params) -> pd.DataFrame:
    ing = params.ingest
    load.last_corrections = []
    if ing["source"] == "workbook":
        df = load_from_workbook(ing["workbook_path"], ing.get("sheet", "Input Data"),
                                ing.get("first_data_row", 20))
    elif ing["source"] == "csv":
        df = load_from_csv(ing["csv_path"])
    elif ing["source"] == "sql":
        nc = set(ing.get("non_consortium_spacecraft", []) or [])
        sql_cfg = dict(ing["sql"])
        sql_cfg.setdefault("as_at", str(getattr(params, "as_at", "")))
        df = load_from_sql(sql_cfg, non_consortium=nc)
    else:
        raise ValueError(f"Unknown source {ing['source']}")
    return _apply_corrections(df, params)
