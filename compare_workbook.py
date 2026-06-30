#!/usr/bin/env python3
"""Pipeline vs Workbook comparison — one formatted Excel, key figures only.

Reads the workbook's Summary (gross) and Netting Waterfall (netting steps)
tabs by LABEL matching, runs the pipeline per the config (SQL or workbook
source), and writes comparison.xlsx with side-by-side values, deltas and a
pass/fail flag.

Usage:
    python compare_workbook.py --config config/2026Q1.yaml --workbook "T:/path/to/Space_RDS_....xlsx"
    (--workbook falls back to ingest.workbook_path in the config)
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent))
from src.parameters import Params
from src import ingest, engine, scenarios, netting as net_mod

TOL = 1.0
SCENARIOS = ["Proton Flare", "Space Weather", "Generic Defect", "Space Debris", "Max Risk"]
WF_STEPS = ["Gross Loss", "less: External QS", "less: Other Ext RI",
            "less: QS Ceded to FIBL", "less: XoL Ceded to FIBL",
            "Net of XoL IGR (Retained)"]


# ---------- read workbook figures by label ----------

def _scen_key(label: str) -> str | None:
    s = str(label).lower()
    for scen in SCENARIOS:
        if scen.lower().rstrip("s") in s.replace("max risks", "max risk"):
            return scen
    return None


def read_workbook_figures(path: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)

    # Summary: entity blocks -> scenario -> gross (col D)
    ws = wb["Summary"]
    gross, current = {}, None
    for row in ws.iter_rows(min_row=1, max_row=60, max_col=4, values_only=True):
        b, d = row[1], row[3]
        if b in ("FIHL", "FUL", "FIID"):
            current = b
            continue
        if b in ("FIBL (Direct)",):
            current = None
        if current and b:
            scen = _scen_key(b)
            if scen and isinstance(d, (int, float)):
                gross[(current, scen)] = float(d)

    # Netting Waterfall: per entity x scenario step values
    ws = wb["Netting Waterfall"]
    steps, current = {}, None
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        b, c = row[1], row[2]
        if isinstance(b, str) and b.startswith("Netting Down"):
            # "Netting Down – FIID: Space Weather (...)"
            try:
                ent = b.split("–")[1].split(":")[0].strip()
                scen = _scen_key(b.split(":", 1)[1])
                current = (ent, scen) if scen else None
            except Exception:
                current = None
            continue
        if current and isinstance(b, str) and b.strip() in WF_STEPS \
                and isinstance(c, (int, float)):
            steps[(*current, b.strip())] = float(c)
    wb.close()
    return {"gross": gross, "steps": steps}


# ---------- pipeline figures ----------

def pipeline_figures(p) -> dict:
    df = ingest.load(p)
    df = engine.run_engine(df, p)
    per_layer, _, _ = scenarios.run_scenarios(df, p)
    nets = net_mod.entity_scenario_netting(per_layer, p)
    fihl = net_mod.fihl_gross(per_layer, p)

    gross, steps, detail = {}, {}, {}
    for _, r in fihl.iterrows():
        gross[("FIHL", r["scenario"])] = r["gross"]
        if r["detail"]:
            detail[("FIHL", r["scenario"])] = r["detail"]
    for _, r in nets.iterrows():
        key = (r["entity"], r["scenario"])
        gross[key] = r["gross"]
        if r.get("detail"):
            detail[key] = r["detail"]
        steps[(*key, "Gross Loss")] = r["gross"]
        steps[(*key, "less: External QS")] = -r["ext_qs"]
        steps[(*key, "less: Other Ext RI")] = -r["other_ext_ri"]
        steps[(*key, "less: QS Ceded to FIBL")] = -r["igr_qs_ceded"]
        steps[(*key, "less: XoL Ceded to FIBL")] = -r["xol_ceded"]
        steps[(*key, "Net of XoL IGR (Retained)")] = r["net"]
    return {"gross": gross, "steps": steps, "detail": detail}


# ---------- formatted output ----------

INK = "1A1815"; ACCENT = "B8341A"; CREAM = "FAF6EC"
GREEN = "2E7D32"; RED = "C62828"; RULE = "D8D2C4"

F_TITLE = Font(name="Georgia", size=16, bold=True, color=INK)
F_SUB = Font(name="Arial", size=10, italic=True, color="6A655F")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_ENT = Font(name="Georgia", size=12, bold=True, color=ACCENT)
F_CELL = Font(name="Arial", size=10, color=INK)
F_OK = Font(name="Arial", size=10, bold=True, color=GREEN)
F_BAD = Font(name="Arial", size=10, bold=True, color=RED)
FILL_HDR = PatternFill("solid", start_color=ACCENT)
FILL_ALT = PatternFill("solid", start_color=CREAM)
THIN = Border(bottom=Side(style="thin", color=RULE))
MONEY = '#,##0;(#,##0);"–"'


def write_comparison(out_path: str, wbk: dict, pipe: dict, quarter: str,
                     source: str):
    wb = Workbook()

    # ----- Sheet 1: headline grid -----
    ws = wb.active
    ws.title = "Headlines"
    ws.sheet_view.showGridLines = False
    ws["B2"] = f"Space RDS {quarter} · pipeline vs workbook"
    ws["B2"].font = F_TITLE
    ws["B3"] = f"Pipeline source: {source} · diffs beyond ±$1 flagged"
    ws["B3"].font = F_SUB

    headers = ["Scenario", "Workbook Gross", "Pipeline Gross", "Δ Gross",
               "Workbook Net", "Pipeline Net", "Δ Net", "Status"]
    widths = [24, 16, 16, 12, 16, 16, 12, 9]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 5
    for entity in ["FIHL", "FUL", "FIID"]:
        ws.cell(row=r, column=2, value=entity).font = F_ENT
        r += 1
        for j, h in enumerate(headers):
            c = ws.cell(row=r, column=2 + j, value=h)
            c.font = F_HDR; c.fill = FILL_HDR
            c.alignment = Alignment(horizontal="right" if j else "left")
        r += 1
        for k, scen in enumerate(SCENARIOS):
            wg = wbk["gross"].get((entity, scen))
            pg = pipe["gross"].get((entity, scen))
            wn = wbk["steps"].get((entity, scen, "Net of XoL IGR (Retained)"))
            pn = pipe["steps"].get((entity, scen, "Net of XoL IGR (Retained)"))
            def _d(a, b):
                if None in (a, b):
                    return None
                d = b - a
                return 0 if abs(d) < 0.005 else round(d, 2)
            row_cells = {
                2: scen, 3: wg, 4: pg, 5: _d(wg, pg),
                6: wn, 7: pn, 8: _d(wn, pn),
            }
            diffs = [v for c, v in row_cells.items() if c in (5, 8) and v is not None]
            ok = all(abs(d) <= TOL for d in diffs) if diffs else None
            for col, val in row_cells.items():
                c = ws.cell(row=r, column=col, value=val)
                c.font = F_CELL; c.border = THIN
                if k % 2 == 0:
                    c.fill = FILL_ALT
                if col > 2:
                    c.number_format = MONEY
                    c.alignment = Alignment(horizontal="right")
            s = ws.cell(row=r, column=9,
                        value=("✓" if ok else "✗") if ok is not None else "·")
            s.font = F_OK if ok else (F_BAD if ok is not None else F_CELL)
            s.alignment = Alignment(horizontal="center")
            s.border = THIN
            if k % 2 == 0:
                s.fill = FILL_ALT
            r += 1
        r += 1

    # ----- Sheet 2: netting steps for FUL / FIID -----
    ws2 = wb.create_sheet("Netting steps")
    ws2.sheet_view.showGridLines = False
    ws2["B2"] = "Netting waterfalls · step by step"
    ws2["B2"].font = F_TITLE
    ws2["B3"] = "Workbook = Netting Waterfall tab · Pipeline = layer-exact engine"
    ws2["B3"].font = F_SUB
    for i, w in enumerate([26, 16, 16, 12, 9], start=2):
        ws2.column_dimensions[chr(64 + i)].width = w

    r = 5
    for entity in ["FUL", "FIID"]:
        for scen in SCENARIOS:
            if not any((entity, scen, s) in wbk["steps"] for s in WF_STEPS) and \
               not any((entity, scen, s) in pipe["steps"] for s in WF_STEPS):
                continue
            det = pipe["detail"].get((entity, scen), "")
            ws2.cell(row=r, column=2,
                     value=f"{entity} · {scen}" + (f" — {det}" if det else "")
                     ).font = F_ENT
            r += 1
            for j, h in enumerate(["Step", "Workbook", "Pipeline", "Δ", "Status"]):
                c = ws2.cell(row=r, column=2 + j, value=h)
                c.font = F_HDR; c.fill = FILL_HDR
                c.alignment = Alignment(horizontal="right" if j else "left")
            r += 1
            for k, step in enumerate(WF_STEPS):
                wv = wbk["steps"].get((entity, scen, step))
                pv = pipe["steps"].get((entity, scen, step))
                d = None if None in (wv, pv) else (pv - wv)
                if d is not None and abs(d) < 0.005:
                    d = 0
                ok = None if d is None else abs(d) <= TOL
                vals = {2: step, 3: wv, 4: pv, 5: d}
                for col, val in vals.items():
                    c = ws2.cell(row=r, column=col, value=val)
                    c.font = F_CELL; c.border = THIN
                    if k % 2 == 0:
                        c.fill = FILL_ALT
                    if col > 2:
                        c.number_format = MONEY
                        c.alignment = Alignment(horizontal="right")
                s = ws2.cell(row=r, column=6,
                             value=("✓" if ok else "✗") if ok is not None else "·")
                s.font = F_OK if ok else (F_BAD if ok is not None else F_CELL)
                s.alignment = Alignment(horizontal="center"); s.border = THIN
                if k % 2 == 0:
                    s.fill = FILL_ALT
                r += 1
            r += 1

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--workbook", default=None,
                    help="path to the Excel workbook (defaults to ingest.workbook_path)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    p = Params.load(args.config)
    wb_path = args.workbook or p.ingest.get("workbook_path")
    if not wb_path:
        sys.exit("No workbook path: pass --workbook or set ingest.workbook_path")

    out = args.out or f"output/{p.quarter}/comparison.xlsx"
    Path(out).parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading workbook figures: {wb_path}")
    wbk = read_workbook_figures(wb_path)
    print(f"  {len(wbk['gross'])} gross figures · {len(wbk['steps'])} netting steps")

    print(f"Running pipeline ({p.ingest['source']} source)…")
    pipe = pipeline_figures(p)

    write_comparison(out, wbk, pipe, p.quarter, p.ingest["source"])
    print(f"Done → {out}")


if __name__ == "__main__":
    main()
