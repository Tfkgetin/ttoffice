"""Exports: CSVs, run manifest."""
from __future__ import annotations
import json
import datetime as dt
from pathlib import Path
import pandas as pd

from . import excel_report, persist, changes as changes_mod, engine as engine_mod, ingest as ingest_mod


def export(outdir: str, per_layer: pd.DataFrame, sw: pd.DataFrame,
           mr: pd.DataFrame, netting: pd.DataFrame, summary: pd.DataFrame,
           recon: pd.DataFrame, params, changes=None) -> Path:
    # `changes` is an optional caller-supplied Q-on-Q dict (run_space_rds builds
    # one from the frozen prior workbook). export computes its own `chg` below,
    # which is the source of truth; the caller's is used only as a fallback when
    # export can't build one (e.g. no prior run and no prior_workbook configured).
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)

    per_layer.drop(columns=[c for c in per_layer.columns if c.startswith("wb_")],
                   errors="ignore").to_csv(out / "per_layer.csv", index=False)
    sw.to_csv(out / "space_weather_by_manufacturer.csv", index=False)
    mr.to_csv(out / "max_risk_by_layer.csv", index=False)
    netting.to_csv(out / "netting_waterfalls.csv", index=False)
    summary.to_csv(out / "summary_grid.csv", index=False)
    recon.to_csv(out / "reconciliation.csv", index=False)

    # Q-on-Q: diff against the latest prior persisted run (before saving this one)
    prior_sel = params.raw.get("changes", {}).get("compare_to") \
        if isinstance(params.raw.get("changes"), dict) else None
    prior = persist.load_prior(str(params.as_at), explicit=prior_sel)
    if prior is None:
        print("  ⚠  NO PRIOR RUN FOUND in ./runs — Q-on-Q tabs will be empty:")
        print("     · 'Changes' tab will be skipped")
        print("     · 'WF · Exposure Bridge' / 'WF · Loss Movement' will show 'no prior'")
        existing = persist.list_runs()
        print(f"     runs/ currently holds: {existing or '(nothing)'}")
        print(f"     (cwd={Path.cwd()}; persist looks in {Path('runs').resolve()})")
        print("     To enable: put last quarter's run in runs/<as_at>/"
              "{layers.csv,grid.csv,meta.json}")

    # Build S3123 grid early so it can feed both the sheet and Q-on-Q
    s3grid = None
    if (params.raw.get("s3123_rds") or {}).get("enabled"):
        from . import s3123 as s3123_mod, s3123_sheet
        s3grid = s3123_mod.s3123_grid(per_layer, params)
        s3grid.to_csv(out / "s3123_grid.csv", index=False)

    chg = changes_mod.compute(per_layer, summary, prior, cur_s3=s3grid)

    # ------------------------------------------------------------------ #
    # FIX(recon 2026Q1): when the config names a frozen prior_workbook,
    # seed Q-on-Q from THAT (the filed prior) instead of the persisted run.
    # This is the manual's own "From_PQ" mechanism and, crucially, adds the
    # exec-basis columns (FIHL combined / FIBL receiver) the Change-Narrative
    # matrix needs — reproduced to the dollar against JJ's narrative tabs.
    # A persisted-run prior lacks those columns, so the Changes tab would
    # fall back to the Summary basis and not match JJ's Exec matrix.
    # ------------------------------------------------------------------ #
    prior_wb = params.raw.get("prior_workbook")
    if prior_wb:
        try:
            from . import prior_seed, netting as _net
            prior_layers = prior_seed.load_prior_workbook(prior_wb)
            # Restore prior S3123/equity add-ons if the auto book cached them as
            # formulas — keeps the FIHL combined prior on the same basis as current.
            prior_layers = prior_seed.restore_addons(prior_layers, per_layer, params)
            prior_grid = _net.summary_grid(prior_layers, params)
            label = params.raw.get("prior_as_at_label") or "prior (frozen workbook)"
            frozen_chg = prior_seed.build_changes(per_layer, summary,
                                                  prior_layers, prior_grid,
                                                  prior_as_at=label)
            # preserve the S3123 movement the persisted path already computed
            if chg and "s3123" in chg and "s3123" not in frozen_chg:
                frozen_chg["s3123"] = chg["s3123"]
            chg = frozen_chg
            print(f"      Q-on-Q seeded from frozen workbook: {prior_wb}")
        except Exception as e:  # noqa: BLE001
            print(f"      WARNING: frozen-workbook prior failed ({e}); "
                  f"using persisted-run prior")

    if chg is None and changes is not None:
        chg = changes  # caller-supplied (e.g. frozen-workbook) Q-on-Q fallback

    # Renewal-policy check: classify the quarter's dropped/expiring layers from
    # the underwriter's forward pointer (J.Pbi.Layers_t snapshot) — renewed &
    # bound / in progress / NTU-Declined / no pointer. Optional & guarded; falls
    # back silently to the spacecraft heuristic when the pointer feed is absent.
    if chg and "layers" in chg:
        from . import renewals
        import pandas as _pd
        watch = renewals.load_watch(params)
        # Merge UW-confirmed renewal_overrides (old→new links J.Pbi doesn't carry,
        # e.g. WorldView 341568 → 369418) into the watch — classification only.
        ow = renewals.overrides_watch(params.raw.get("renewal_overrides"))
        if ow is not None:
            watch = ow if watch is None else _pd.concat([watch, ow], ignore_index=True)
        if watch is not None:
            cur_pids = set(per_layer["program_id"]) if "program_id" in per_layer else set()
            for key in ("renewal_gaps", "dropped", "new"):
                fr = chg["layers"].get(key)
                if fr is not None and len(fr):
                    chg["layers"][key] = renewals.attach_watch(fr, watch, cur_pids)
        # Pointer-aware movement split — the single source of truth for what
        # renewed vs lapsed. renewal_gaps is reset to the layers that truly went
        # off the book (renewal not present as a new/renewed layer), so the
        # Changes tab, the console summary and Book Movement's Lapsed table agree.
        sp = renewals.split_movement(chg["layers"].get("new"), chg["layers"].get("dropped"))
        chg["layers"]["renewal_gaps"] = sp["lapsed"]
        gaps = sp["lapsed"]
        if gaps is not None and len(gaps) and "renewal_state" in getattr(gaps, "columns", []):
            summ = renewals.summarize(gaps["renewal_state"])
            print("      renewal check (lapsed): "
                  + ", ".join(f"{renewals.label(k)}={v}" for k, v in summ.items()))

    excluded = getattr(engine_mod.run_engine, "last_excluded", None)
    corrections = getattr(ingest_mod.load, "last_corrections", None)
    manual_includes = getattr(ingest_mod.load, "last_manual_includes", None)
    excel_report.write_results(
        str(out / f"Space_RDS_results_{params.as_at}.xlsx"),
        per_layer, sw, mr, summary, params,
        source=params.ingest.get("source", "?"), recon=recon,
        changes=chg, excluded=excluded,
        corrections=corrections, manual_includes=manual_includes)

    # S3123 (Lloyd's) RDS — separate syndicate tab + reconciliation
    if s3grid is not None:
        manual = (params.raw.get("s3123_rds") or {}).get("manual_compare")
        rec = s3123_mod.reconcile(s3grid, manual) if manual else None
        s3_qoq = chg.get("s3123") if chg else None
        import openpyxl as _opx
        _fp = str(out / f"Space_RDS_results_{params.as_at}.xlsx")
        _wb = _opx.load_workbook(_fp)
        s3123_sheet.write_s3123_sheet(
            _wb, s3grid, rec, as_at=str(params.as_at), qoq=s3_qoq)
        _wb.save(_fp)

    # Lloyd's RDS Summary (JJ's filed deliverable) — the workbook's LAST tab.
    # Rendered from the pipeline's COMPUTED S3123 grid (gross + net), so it does
    # not depend on the Lloyd's SQL views (which currently error in SQL). The
    # bus-type block is enriched from vw_SpaceRDS_SpaceWeather_Lloyds when that
    # view is readable; otherwise it shows a banner. Added after every other
    # sheet so it lands last.
    if s3grid is not None and (params.raw.get("s3123_rds") or {}).get("enabled"):
        from . import lloyds_sheet
        import openpyxl as _opx2
        views = ingest_mod.load_lloyds_rds_summary(params)   # may be None (view broken)
        sw_view = (views or {}).get("space_weather")
        lcfg = (params.raw.get("s3123_rds") or {}).get("lloyds_summary", {})
        # prior = JJ's filed previous-quarter RDS ({scenario: {gross, net}})
        s3_prior = dict(lcfg.get("previous_quarter") or {})
        if s3_prior:
            s3_prior["_label"] = lcfg.get("previous_quarter_label", "prior")
        appetite = lcfg.get("risk_appetite_usd")
        _fpl = str(out / f"Space_RDS_results_{params.as_at}.xlsx")
        _wbl = _opx2.load_workbook(_fpl)
        lloyds_sheet.write_lloyds_rds_summary(
            _wbl, s3grid, as_at=str(params.as_at), sw_view=sw_view,
            risk_appetite=appetite, prior=s3_prior,
            qoq_note=lcfg.get("qoq_note"), params=params)
        # The computed 'S3123 RDS' tab is superseded by the Lloyd's Summary —
        # keep it in the file (audit) but hide it from view.
        if "S3123 RDS" in _wbl.sheetnames:
            _wbl["S3123 RDS"].sheet_state = "hidden"
        _wbl.save(_fpl)
        print("      Lloyd's RDS Summary tab written (last tab); S3123 RDS hidden")

    # persist this run so future quarters can diff against it
    persist.save_run(params, per_layer, summary)
    if chg:
        print(f"      changes vs {chg['prior_as_at']}: "
              f"{chg['layers']['summary']['new_layers']} new, "
              f"{chg['layers']['summary']['dropped_layers']} dropped")
        if "s3123" in chg:
            s3d = chg["s3123"]
            sw_row = s3d[s3d["scenario"] == "Space Weather"]
            if len(sw_row):
                dg = sw_row.iloc[0].get("d_gross")
                dn = sw_row.iloc[0].get("d_net")
                print(f"      S3123 SW: gross {'+' if dg and dg>=0 else ''}"
                      f"{dg:,.0f}  net {'+' if dn and dn>=0 else ''}{dn:,.0f}"
                      if dg is not None else "      S3123: no prior to compare")

    manifest = {
        "quarter": params.quarter,
        "as_at": str(params.as_at),
        "run_at": dt.datetime.now().isoformat(timespec="seconds"),
        "rows": len(per_layer),
        "layers": int(per_layer["layer_key"].nunique()),
        "reconciliation_ok": bool((recon["status"] == "OK").all()),
        "outputs": sorted(f.name for f in out.iterdir() if f.is_file()
                          and f.suffix in (".csv", ".xlsx")),
    }
    (out / "manifest.json").write_text(json.dumps(manifest, indent=2))
    return out
