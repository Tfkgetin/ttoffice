"""Render the S3123 (Lloyd's) RDS as its own worksheet, separate from the IG
book, with an inline reconciliation to the manual Lloyd's submission."""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="1F2933"; NAVY="1B3A5C"; GREEN="2D6A3C"; RED="C0392B"
SOFT="6B7785"; RULE="DCE3EA"; CREAM="EEF3F8"; WHITE="FFFFFF"
AMBER="9A6410"; AMBERFILL="FBF3E4"
F="Calibri"
def _f(sz=10,b=False,c=INK): return Font(name=F,size=sz,bold=b,color=c)
def _fill(c): return PatternFill("solid",fgColor=c)
def _money(v):
    if v is None or v=="" or (isinstance(v,float) and v!=v): return "-"
    return "${:,.0f}".format(v)
thin=Side(style="thin",color=RULE)


def _quality_flags(grid):
    """FIX(recon 2026Q1, O2): the shipped 2026Q1 sheet printed $0 for Space
    Weather (population lacked lloyds_bus_type) and $0 for every net (stale /
    misconfigured grid) — both silently. Detect those states and return visible
    banner strings instead, so a broken feed reads as broken, not as zero risk."""
    flags = []
    try:
        g = grid.set_index("scenario")
        sw = g.loc["Space Weather"] if "Space Weather" in g.index else None
        if sw is not None and float(sw.get("gross") or 0) == 0 and \
                (sw.get("detail") in (None, "", "None") or sw.get("detail") != sw.get("detail")):
            flags.append("Space Weather is $0 with no bus-type detail — the "
                         "population is missing lloyds_bus_type (Lloyd's SW "
                         "view not joined). Value is UNAVAILABLE, not zero.")
        gross_pos = (grid["gross"].fillna(0) > 0)
        if gross_pos.any() and (grid.loc[gross_pos, "net"].fillna(0) == 0).all():
            flags.append("Every NET is $0 while gross is positive — check the "
                         "agg_xol config (an attachment of 0 wipes the net) "
                         "and that the CURRENT grid was passed to this sheet.")
    except Exception:
        pass
    return flags


def write_s3123_sheet(wb, grid, recon=None, notes=None, as_at="2026-04-01", qoq=None,
                      sheet_name="S3123 RDS", title="S3123 RDS — Syndicate 3123 (Lloyd's)",
                      subtitle=None, alt_groups=None):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws=wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGH",[3,22,30,16,16,16,16,42]): ws.column_dimensions[col].width=w
    r=2
    ws.cell(r,2,title).font=_f(16,True,NAVY)
    r+=1
    ws.cell(r,2,subtitle or f"Reported separately from the IG return · as at {as_at} · clean Lloyd's structure").font=_f(10,False,SOFT)
    r+=1
    # ---- data-quality banners (FIX recon 2026Q1, O2) ----
    for flag in _quality_flags(grid):
        cell=ws.cell(r,2,"⚠  "+flag); cell.font=_f(10,True,AMBER)
        cell.alignment=Alignment(wrap_text=True)
        for c in range(2,9): ws.cell(r,c).fill=_fill(AMBERFILL)
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
        ws.row_dimensions[r].height=28
        r+=1
    r+=1
    # ---- S3123 grid ----
    ws.cell(r,2,"SCENARIO").font=_f(9,True,WHITE); ws.cell(r,3,"WORST / BASIS").font=_f(9,True,WHITE)
    ws.cell(r,4,"GROSS").font=_f(9,True,WHITE); ws.cell(r,5,"NET").font=_f(9,True,WHITE)
    for c in range(2,6):
        ws.cell(r,c).fill=_fill(NAVY); ws.cell(r,c).alignment=Alignment(horizontal="right" if c>3 else "left")
    ws.cell(r,2).alignment=Alignment(horizontal="left"); ws.cell(r,3).alignment=Alignment(horizontal="left")
    r+=1
    for _,row in grid.iterrows():
        ws.cell(r,2,row["scenario"]).font=_f(10,True,INK)
        ws.cell(r,3,str(row["detail"]) if row["detail"] is not None else "").font=_f(10,False,SOFT)
        # FIX(recon 2026Q1, O2): a zero gross with no basis is an unavailable
        # value (e.g. SW without lloyds_bus_type) — print n/a, never $0.
        _unavail = (float(row.get("gross") or 0) == 0 and
                    (row.get("detail") in (None, "", "None")
                     or row.get("detail") != row.get("detail")))
        for c,key in [(4,"gross"),(5,"net")]:
            v = "n/a" if _unavail else _money(row[key])
            cell=ws.cell(r,c,v)
            cell.font=_f(10,False,AMBER if _unavail else INK)
            cell.alignment=Alignment(horizontal="right")
        for c in range(2,6): ws.cell(r,c).border=Border(bottom=thin)
        r+=1
    r+=1
    # ---- Space Debris altitude bands: the losing band, shown ----
    # The scenario reports only the band it picked, so a wrong pick (or the
    # all-LEO fallback) is invisible on the grid above. This makes the MAX
    # checkable and shows any LEO the bands do not claim.
    if alt_groups is not None and len(alt_groups):
        ws.cell(r,2,"SPACE DEBRIS — ALTITUDE BANDS").font=_f(11,True,NAVY); r+=1
        ws.cell(r,2,"Worst band is taken; the rest are shown so the pick can be "
                    "checked. Bounds come from config (altitude_groups).").font=_f(9,False,SOFT)
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8); r+=1
        for c,h in zip(range(2,7),["BAND","LAYERS","GROSS","NET","PICK"]):
            ws.cell(r,c,h).font=_f(9,True,WHITE); ws.cell(r,c).fill=_fill(SOFT)
            ws.cell(r,c).alignment=Alignment(horizontal="left" if c in (2,6) else "right")
        r+=1
        for _,g in alt_groups.iterrows():
            sel=bool(g.get("selected")); banded=bool(g.get("in_pick"))
            ws.cell(r,2,str(g["group"])).font=_f(10,sel,INK if banded else SOFT)
            lc=ws.cell(r,3,int(g["layers"])); lc.font=_f(10,False,INK)
            lc.alignment=Alignment(horizontal="right")
            for c,key in [(4,"gross"),(5,"net")]:
                v=_money(g[key]) if (banded or key=="gross") else ""
                cell=ws.cell(r,c,v); cell.font=_f(10,sel,INK if banded else SOFT)
                cell.alignment=Alignment(horizontal="right")
            note="◄ selected" if sel else ("" if banded else "not banded — excluded")
            ws.cell(r,6,note).font=_f(9,sel,NAVY if sel else SOFT)
            for c in range(2,7): ws.cell(r,c).border=Border(bottom=thin)
            r+=1
        r+=1
    # ---- reconciliation vs manual ----
    if recon is not None and len(recon):
        ws.cell(r,2,"RECONCILIATION vs MANUAL (JJ, 2026Q1)").font=_f(11,True,NAVY); r+=1
        heads=["SCENARIO","AUTO GROSS","MANUAL GROSS","Δ GROSS","AUTO NET","MANUAL NET","WHY THEY DIFFER"]
        for c,h in zip(range(2,9),heads):
            ws.cell(r,c,h).font=_f(9,True,WHITE); ws.cell(r,c).fill=_fill(SOFT)
            ws.cell(r,c).alignment=Alignment(horizontal="left" if c in (2,8) else "right")
        r+=1
        for _,row in recon.iterrows():
            ws.cell(r,2,row["scenario"]).font=_f(10,True,INK)
            for c,key in [(3,"auto_gross"),(4,"manual_gross"),(5,"d_gross"),(6,"auto_net"),(7,"manual_net")]:
                v=row.get(key); cell=ws.cell(r,c,_money(v)); cell.alignment=Alignment(horizontal="right")
                col=RED if (key=="d_gross" and v not in (None,"") and abs(v)>1000) else INK
                cell.font=_f(10,False,col)
            ws.cell(r,8,row.get("why","")).font=_f(9,False,INK)
            ws.cell(r,8).alignment=Alignment(wrap_text=True,vertical="top")
            for c in range(2,9): ws.cell(r,c).border=Border(bottom=thin)
            r+=1
    # ---- Q-on-Q movement ----
    if qoq is not None and len(qoq):
        r+=1
        ws.cell(r,2,"Q-on-Q MOVEMENT").font=_f(11,True,NAVY); r+=1
        for c,h in zip(range(2,8),["SCENARIO","PRIOR GROSS","CURRENT GROSS","Δ GROSS","CURRENT NET","Δ NET"]):
            ws.cell(r,c,h).font=_f(9,True,WHITE); ws.cell(r,c).fill=_fill("3E6C8C")
            ws.cell(r,c).alignment=Alignment(horizontal="left" if c==2 else "right")
        r+=1
        for _,row in qoq.iterrows():
            ws.cell(r,2,row["scenario"]).font=_f(10,True,INK)
            for c,key in [(3,"prior_gross"),(4,"cur_gross"),(5,"d_gross"),(6,"cur_net"),(7,"d_net")]:
                v=row.get(key); cell=ws.cell(r,c,_money(v)); cell.alignment=Alignment(horizontal="right")
                col=GREEN if (key in ("d_gross","d_net") and v is not None and not (isinstance(v,float) and v!=v) and v<0) else (RED if (key in ("d_gross","d_net") and v is not None and not (isinstance(v,float) and v!=v) and v>0) else INK)
                cell.font=_f(10,False,col)
            for c in range(2,8): ws.cell(r,c).border=Border(bottom=thin)
            r+=1
    r+=1
    if notes:
        ws.cell(r,2,"METHODOLOGY (clean Lloyd's vs IG)").font=_f(10,True,NAVY); r+=1
        for n in notes:
            ws.cell(r,2,"•  "+n).font=_f(9,False,INK); ws.cell(r,2).alignment=Alignment(wrap_text=True)
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8); r+=1
    return ws
