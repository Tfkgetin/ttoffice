"""Replace computed values on the Per Layer and Netting Waterfalls tabs with live
Excel formulas, so the reinsurance arithmetic recomputes in the workbook instead
of sitting as pasted numbers.

What becomes live, and what deliberately does not:

  Per Layer
    • treaty netting chain — igr_qs_ceded, net_of_qs, net_of_xol, total_fibl_ceded
      are exact arithmetic on same-row inputs, so they link in-row.
    • scenario losses — pf / gd / sd link to per_sc, the orbit, rpf and the
      scenario rates on the Parameters tab.
    • left as values — the raw inputs (ids, names, orbit, dates, per_sc, rates) and
      the treaty cessions that aren't cleanly invertible in-sheet (ext_qs, s3123_qs,
      equity_usd). per_sc and ext_qs are the inputs the chain runs off.

  Netting Waterfalls
    • the combined gross/net table and every cede-down step link to the canonical
      Summary cascade (Gross D, Ext QS E, IGR QS G, XoL I, Net J). The tab becomes a
      formula-linked presentation of Summary rather than a second copy of the numbers.

Run as a finishing pass after the workbook is built (excel_report.write_results):

    from . import linkify
    linkify.linkify(wb)

Idempotent and label-located, so it tolerates row/column shifts and re-runs cleanly.
The Python engine remains the source of truth; these formulas reproduce its
arithmetic transparently, and the reconciliation still diffs Python against source.
"""
from __future__ import annotations

SCEN_ORDER = ["Proton Flare", "Space Weather", "Generic Defect",
              "Space Debris", "Max Risk"]


# --------------------------------------------------------------------------- #
#  locators                                                                     #
# --------------------------------------------------------------------------- #
def _col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def _header_cols(ws, header_row=1):
    return {ws.cell(row=header_row, column=c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=header_row, column=c).value}


def _param_ref(params_ws, label):
    """Absolute ref to the value cell for a Parameters label, or None."""
    for r in range(1, params_ws.max_row + 1):
        if str(params_ws.cell(row=r, column=2).value).strip() == label:
            for c in range(3, 8):
                v = params_ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    return f"Parameters!${_col_letter(c)}${r}"
    return None


def _label_cell_ref(ws, label, startswith=False, col=3, sheet="Parameters"):
    """Absolute ref to col `col` on the row whose B-label matches `label`."""
    for r in range(1, ws.max_row + 1):
        b = str(ws.cell(row=r, column=2).value or "")
        if (b.startswith(label) if startswith else b.strip() == label):
            return f"{sheet}!${_col_letter(col)}${r}"
    return None


def _ensure_s3123_factors(pw):
    """Surface the S3123 consortium time factors on Parameters (idempotent).
    Returns (ref_from_2026, ref_before_2026)."""
    from copy import copy
    lab_ge = "S3123 consortium factor — inception \u2265 2026-01-01"
    lab_lt = "S3123 consortium factor — inception < 2026-01-01"
    r_ge = next((r for r in range(1, pw.max_row + 1)
                 if str(pw.cell(row=r, column=2).value or "") == lab_ge), None)
    r_lt = next((r for r in range(1, pw.max_row + 1)
                 if str(pw.cell(row=r, column=2).value or "") == lab_lt), None)
    if r_ge and r_lt:
        return f"Parameters!$C${r_ge}", f"Parameters!$C${r_lt}"

    last = pw.max_row
    while last > 1 and pw.cell(row=last, column=2).value in (None, ""):
        last -= 1
    base = last + 2

    def _style(dst_rc, src_rc):
        s = pw.cell(row=src_rc[0], column=src_rc[1])
        d = pw.cell(row=dst_rc[0], column=dst_rc[1])
        d.font = copy(s.font); d.fill = copy(s.fill)
        d.alignment = copy(s.alignment); d.border = copy(s.border)
        d.number_format = s.number_format

    # section header (style from an existing S3123 header row if present)
    hdr_src = next((r for r in range(1, pw.max_row + 1)
                    if str(pw.cell(row=r, column=2).value or "").startswith("Syndicate 3123")), 24)
    pw.cell(row=base, column=2, value="Syndicate 3123 consortium factors")
    _style((base, 2), (hdr_src, 2))
    rows = [(base + 1, lab_ge, 0.5), (base + 2, lab_lt, "=5/12")]
    for rr, lab, val in rows:
        pw.cell(row=rr, column=2, value=lab); _style((rr, 2), (25, 2))
        pw.cell(row=rr, column=3, value=val); _style((rr, 3), (25, 3))
        pw.cell(row=rr, column=4,
                value="inwards QS time factor (consortium schedule)")
        _style((rr, 4), (25, 4))
    return f"Parameters!$C${base + 1}", f"Parameters!$C${base + 2}"


def _ensure_equity_table(pw):
    """Surface the IG equity-share-by-inception-year table on Parameters
    (idempotent). Returns (year_range_ref, rate_range_ref) for SUMIFS."""
    from copy import copy
    header = "IG equity share by inception year"
    rates = [(2022, 0.0), (2024, 0.099), (2025, 0.075), (2026, 0.075)]

    hr = next((r for r in range(1, pw.max_row + 1)
               if str(pw.cell(row=r, column=2).value or "") == header), None)
    if hr:
        r0 = hr + 1
        r1 = r0
        while r1 <= pw.max_row and isinstance(pw.cell(row=r1, column=2).value, (int, float)):
            r1 += 1
        r1 -= 1
        return f"Parameters!$B${r0}:$B${r1}", f"Parameters!$C${r0}:$C${r1}"

    def _style(dst_rc, src_rc):
        s = pw.cell(row=src_rc[0], column=src_rc[1])
        d = pw.cell(row=dst_rc[0], column=dst_rc[1])
        d.font = copy(s.font); d.fill = copy(s.fill)
        d.alignment = copy(s.alignment); d.border = copy(s.border)
        d.number_format = s.number_format

    last = pw.max_row
    while last > 1 and pw.cell(row=last, column=2).value in (None, ""):
        last -= 1
    base = last + 2
    pw.cell(row=base, column=2, value=header)
    hdr_src = next((r for r in range(1, pw.max_row + 1)
                    if str(pw.cell(row=r, column=2).value or "").startswith("Syndicate 3123")), 24)
    _style((base, 2), (hdr_src, 2))
    r0 = base + 1
    for i, (yr, rate) in enumerate(rates):
        rr = r0 + i
        pw.cell(row=rr, column=2, value=yr); _style((rr, 2), (25, 2))
        pw.cell(row=rr, column=2).number_format = "0"
        pw.cell(row=rr, column=3, value=rate); _style((rr, 3), (25, 3))
        pw.cell(row=rr, column=4,
                value="FIHL equity participation by underwriting year")
        _style((rr, 4), (25, 4))
    r1 = r0 + len(rates) - 1
    return f"Parameters!$B${r0}:$B${r1}", f"Parameters!$C${r0}:$C${r1}"


def _summary_bases(summary_ws):
    """Map entity -> first-scenario row, by the four 'Proton Flare' rows in order
    (Summary lists FIHL, FUL, FIID, FIBL top to bottom)."""
    pf_rows = [r for r in range(1, summary_ws.max_row + 1)
               if summary_ws.cell(row=r, column=2).value == "Proton Flare"]
    order = ["FIHL", "FUL", "FIID", "FIBL"]
    return {order[i]: pf_rows[i] for i in range(min(len(order), len(pf_rows)))}


# --------------------------------------------------------------------------- #
#  Per Layer                                                                    #
# --------------------------------------------------------------------------- #
def linkify_per_layer(wb, sheet="Per Layer", params="Parameters"):
    if sheet not in wb.sheetnames or params not in wb.sheetnames:
        return
    ws = wb[sheet]
    H = _header_cols(ws, 1)
    need = ["per_sc", "ext_qs", "igr_qs_rate", "igr_qs_ceded", "net_of_qs",
            "xol_ceded", "net_of_xol", "total_fibl_ceded", "orbit", "rpf",
            "pf_fihl", "gd_fihl", "sd_fihl"]
    if any(k not in H for k in need):
        return
    L = {k: _col_letter(H[k]) for k in H}  # name -> column letter

    pw = wb[params]
    s3123_r = _param_ref(pw, "S3123 QS cession")
    pf_r = _param_ref(pw, "Proton Flare insured loss")
    gd_r = _param_ref(pw, "Generic Defect insured loss")
    sd_geo = _param_ref(pw, "Space Debris damage ratio — GEO-GSO")
    sd_leo = _param_ref(pw, "Space Debris damage ratio — LEO")
    sd_meo = _param_ref(pw, "Space Debris damage ratio — MEO")

    last = ws.max_row
    while last > 1 and ws.cell(row=last, column=H["per_sc"]).value in (None, ""):
        last -= 1

    per, ext, rate = L["per_sc"], L["ext_qs"], L["igr_qs_rate"]
    igr, nq, xol = L["igr_qs_ceded"], L["net_of_qs"], L["xol_ceded"]
    nx, fibl, orb, rpf = L["net_of_xol"], L["total_fibl_ceded"], L["orbit"], L["rpf"]
    offr, scid, incp, s3 = (L["off_risk_date"], L["spacecraft_id"],
                            L["inception"], L["s3123_qs"])

    # refs for RPF (as-at date + bands) and s3123 (cession, excluded, factors)
    asat_ref = _label_cell_ref(pw, "As-at date")
    band0 = _label_cell_ref(pw, "RPF: months \u2265 0")
    band_rng = None
    if band0:
        br = int(band0.split("$")[-1]); band_rng = f"Parameters!$C${br}:$C${br + 4}"
    cession_ref = _label_cell_ref(pw, "S3123 QS cession")
    excl_ref = _label_cell_ref(pw, "S3123 excluded spacecraft", startswith=True)
    f_ge, f_lt = _ensure_s3123_factors(pw)
    eq_year_rng, eq_rate_rng = _ensure_equity_table(pw)

    for r in range(2, last + 1):
        def put(name, formula):
            ws.cell(row=r, column=H[name]).value = formula
        # external QS cession (per_sc × S3123 cession rate)
        if s3123_r:
            put("ext_qs", f"={per}{r}*{s3123_r}")
        # RPF — banded on DATEDIF(as-at, off-risk, "M") against the Parameters table
        if asat_ref and band_rng:
            put("rpf", f"=IFERROR(LOOKUP(DATEDIF(DATEVALUE({asat_ref}),"
                       f'DATEVALUE({offr}{r}),"M"),{{0,6,12,18,999}},{band_rng}),0)')
        # S3123 inwards QS — per_sc × cession × consortium factor, zero if excluded
        if cession_ref and excl_ref and f_ge and f_lt:
            excluded = (f'ISNUMBER(SEARCH(","&{scid}{r}&",",'
                        f'","&SUBSTITUTE({excl_ref}," ","")&","))')
            factor = f"IF(DATEVALUE({incp}{r})>=DATE(2026,1,1),{f_ge},{f_lt})"
            put("s3123_qs", f"=IF({excluded},0,{per}{r}*{cession_ref}*{factor})")
        # IG equity share — per_sc × equity rate (by UW year) × consortium factor
        if "equity_usd" in H and eq_year_rng and eq_rate_rng and f_ge and f_lt:
            yr = f"VALUE(LEFT({incp}{r},4))"
            erate = f"SUMIFS({eq_rate_rng},{eq_year_rng},{yr})"
            ef = f"IF(DATEVALUE({incp}{r})>=DATE(2026,1,1),{f_ge},{f_lt})"
            put("equity_usd", f"={per}{r}*{erate}*{ef}")
        # treaty netting chain (exact in-row arithmetic)
        put("igr_qs_ceded", f"=({per}{r}-{ext}{r})*{rate}{r}")
        put("net_of_qs",    f"={per}{r}-{ext}{r}-{igr}{r}")
        put("net_of_xol",   f"={nq}{r}-{xol}{r}")
        put("total_fibl_ceded", f"={igr}{r}")
        # scenario losses (rates from Parameters, orbit-scoped)
        geo = f'ISNUMBER(SEARCH("GEO",{orb}{r}))'
        leo = f'ISNUMBER(SEARCH("LEO",{orb}{r}))'
        meo = f'ISNUMBER(SEARCH("MEO",{orb}{r}))'
        if pf_r:
            put("pf_fihl", f"=IF({geo},{per}{r}*{pf_r},0)")
        if gd_r:
            put("gd_fihl", f"=IF({geo},{per}{r}*{gd_r}*{rpf}{r},0)")
        if sd_geo and sd_leo and sd_meo:
            put("sd_fihl",
                f"={per}{r}*IF({geo},{sd_geo},IF({leo},{sd_leo},IF({meo},{sd_meo},0)))")


# --------------------------------------------------------------------------- #
#  Netting Waterfalls                                                           #
# --------------------------------------------------------------------------- #
def linkify_netting(wb, sheet="Netting Waterfalls", summary="Summary"):
    if sheet not in wb.sheetnames or summary not in wb.sheetnames:
        return
    ws = wb[sheet]
    bases = _summary_bases(wb[summary])
    if "FIHL" not in bases:
        return

    def srow(entity, scen):
        b = bases.get(entity)
        return None if b is None else b + SCEN_ORDER.index(scen)

    # --- combined gross / net tables ----------------------------------------
    # Two stacked tables built by _waterfalls, each headed
    #   'Scenario | FIHL | FUL | FIBL | FIID'  (cols 2..6, FIHL in col 3)
    # with the basis ('gross' / 'net') in the section title above each header.
    # Link every entity cell to the canonical Summary cascade so the tab is a
    # formula-linked view of Summary. NB FIHL mirrors Summary's COMBINED basis
    # (incl S3123 QS + IG equity) — the book-wide FIHL headline convention.
    ent_col = {"FIHL": 3, "FUL": 4, "FIBL": 5, "FIID": 6}
    for r in range(1, ws.max_row + 1):
        if not (ws.cell(row=r, column=2).value == "Scenario"
                and ws.cell(row=r, column=3).value == "FIHL"):
            continue
        basis = None
        for rr in range(r - 1, max(0, r - 4), -1):
            t = str(ws.cell(row=rr, column=2).value or "").lower()
            if "gross" in t:
                basis = "Gross"; break
            if "net" in t:
                basis = "Net"; break
        scol = "J" if basis == "Net" else "D"       # Summary Net / Gross
        rr = r + 1
        while rr <= ws.max_row and ws.cell(row=rr, column=2).value in SCEN_ORDER:
            scen = ws.cell(row=rr, column=2).value
            for ent, c in ent_col.items():
                sr = srow(ent, scen)
                if sr:
                    ws.cell(row=rr, column=c).value = f"=Summary!{scol}{sr}"
            rr += 1

    # --- cede-down detail blocks: link each step to Summary -----------------
    # Four blocks per scenario at label/amount column pairs:
    #   (2,3) FIHL · (5,6) FUL · (8,9) FIBL · (11,12) FIID
    # step offsets from the 'Gross Loss' row: 0 Gross · 1 less Ext QS ·
    # 2 less Other Ext RI · 3 less QS to FIBL · 4 less XoL to FIBL · 5 Net.
    # The Net row is already a live =SUM(...) written by the builder — leave it.
    for r in range(1, ws.max_row + 1):
        for label_col, amt_col in ((2, 3), (5, 6), (8, 9), (11, 12)):
            if ws.cell(row=r, column=label_col).value != "Gross Loss":
                continue
            head = str(ws.cell(row=r - 2, column=label_col).value or "")
            entity = next((e for e in ("FIHL", "FUL", "FIBL", "FIID")
                           if head.startswith(e)), None)
            scen = next((s for s in SCEN_ORDER if s in head), None)
            sr = srow(entity, scen) if entity and scen else None
            if not sr:
                continue
            ws.cell(row=r, column=amt_col).value = f"=Summary!D{sr}"            # Gross
            ws.cell(row=r + 1, column=amt_col).value = f"=-Summary!E{sr}"       # less Ext QS
            ws.cell(row=r + 2, column=amt_col).value = 0                        # Other Ext RI
            ws.cell(row=r + 3, column=amt_col).value = f"=-Summary!G{sr}"       # less IGR QS
            ws.cell(row=r + 4, column=amt_col).value = f"=-Summary!I{sr}"       # less XoL

    # --- chart scaffold (Step | Source | Base | Decrease | Total) ------------
    # link Source to Summary so the floating-waterfall mini-chart is live
    for r in range(1, ws.max_row + 1):
        sc_cell = next((c for c in range(2, ws.max_column + 1)
                        if ws.cell(row=r, column=c).value == "Source ($)"), None)
        if sc_cell is None:
            continue
        ncol = sc_cell - 1               # Step label column
        ocol = sc_cell                   # Source
        pcol, qcol, rcol = sc_cell + 1, sc_cell + 2, sc_cell + 3   # Base/Decrease/Total
        title = str(ws.cell(row=r - 1, column=ncol).value or "")
        entity = next((e for e in ("FUL", "FIID") if title.startswith(e)), None)
        scen = next((s for s in SCEN_ORDER if s in title), None)
        sr = srow(entity, scen) if entity and scen else None
        if not sr:
            continue
        O = _col_letter(ocol)
        # Source: Gross +D, Ext QS -E, IGR QS -G, IGR XoL -I, Net +J
        src = [f"=Summary!D{sr}", f"=-Summary!E{sr}", f"=-Summary!G{sr}",
               f"=-Summary!I{sr}", f"=Summary!J{sr}"]
        for i, formula in enumerate(src):
            ws.cell(row=r + 1 + i, column=ocol).value = formula
        rows = [r + 1 + i for i in range(5)]
        # Base (cumulative after each decrease), Decrease (=-Source), Total (gross/net bars)
        ws.cell(row=rows[0], column=pcol).value = 0
        ws.cell(row=rows[0], column=qcol).value = 0
        ws.cell(row=rows[0], column=rcol).value = f"={O}{rows[0]}"
        for i in (1, 2, 3):
            ws.cell(row=rows[i], column=pcol).value = f"={_col_letter(pcol)}{rows[i-1]}+{O}{rows[i]}" if i > 1 else f"={O}{rows[0]}+{O}{rows[1]}"
            ws.cell(row=rows[i], column=qcol).value = f"=-{O}{rows[i]}"
            ws.cell(row=rows[i], column=rcol).value = 0
        ws.cell(row=rows[4], column=pcol).value = 0
        ws.cell(row=rows[4], column=qcol).value = 0
        ws.cell(row=rows[4], column=rcol).value = f"={O}{rows[4]}"


def _find_row(ws, col, value, start=1, end=None):
    end = end or ws.max_row
    for r in range(start, end + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    return None


# --------------------------------------------------------------------------- #
#  Executive Summary  →  Summary / Per Layer                                    #
# --------------------------------------------------------------------------- #
def linkify_exec_summary(wb, sheet="Executive Summary", summary="Summary",
                         per_layer="Per Layer"):
    if sheet not in wb.sheetnames or summary not in wb.sheetnames:
        return
    ws = wb[sheet]
    bases = _summary_bases(wb[summary])
    if "FIHL" not in bases:
        return
    pl = f"'{per_layer}'"

    def srow(entity, scen):
        b = bases.get(entity)
        return None if b is None else b + SCEN_ORDER.index(scen)

    # headline cards (row with 'WORST CASE' eyebrow, big numbers two rows below)
    card = next((r for r in range(1, ws.max_row + 1)
                 if isinstance(ws.cell(row=r, column=2).value, str)
                 and "WORST CASE" in ws.cell(row=r, column=2).value), None)
    if card:
        vr = card + 2
        fihl = srow("FIHL", "Space Weather")
        # card 2 eyebrow pairs with card 1: FIHL gross -> FIHL net
        if isinstance(ws.cell(row=card, column=4).value, str):
            ws.cell(row=card, column=4).value = "WORST CASE · FIHL NET"
        if fihl:
            ws.cell(row=vr, column=2).value = f'="$"&TEXT(Summary!D{fihl}/1000000,"#,##0")&"m"'
            ws.cell(row=vr, column=4).value = f'="$"&TEXT(Summary!J{fihl}/1000000,"#,##0")&"m"'
        ws.cell(row=vr, column=6).value = f'="$"&TEXT(SUM({pl}!L:L)/1000000,"#,##0")&"m"'

    # headline table: header 'Entity' / 'Worst-case scenario'
    h = next((r for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=2).value == "Entity"
              and ws.cell(row=r, column=3).value in ("Worst-case scenario", "Binding scenario")), None)
    if h:
        for r in range(h + 1, h + 6):
            ent = ws.cell(row=r, column=2).value
            scen = ws.cell(row=r, column=3).value
            sr = srow(ent, scen) if ent in bases and scen in SCEN_ORDER else None
            if not sr:
                continue
            ws.cell(row=r, column=4).value = f"=Summary!D{sr}"       # Gross
            ws.cell(row=r, column=5).value = f"=Summary!J{sr}"       # Net retained
            ws.cell(row=r, column=6).value = f"=IFERROR(E{r}/D{r},0)"  # Net/Gross

    # ranked FUL table: header 'Scenario' + 'Gross' + 'Retained %'
    h = next((r for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=2).value == "Scenario"
              and ws.cell(row=r, column=4).value == "Gross"), None)
    if h:
        for r in range(h + 1, h + 6):
            scen = ws.cell(row=r, column=2).value
            sr = srow("FUL", scen) if scen in SCEN_ORDER else None
            if not sr:
                continue
            ws.cell(row=r, column=4).value = f"=Summary!D{sr}"       # Gross
            ws.cell(row=r, column=6).value = f"=Summary!J{sr}"       # Net retained
            ws.cell(row=r, column=5).value = f"=D{r}-F{r}"           # Ceded
            ws.cell(row=r, column=7).value = f"=IFERROR(F{r}/D{r},0)"  # Retained %

    # run-status: layers / exposure / GEO share
    r = _find_row(ws, 2, "On-risk layers")
    if r:
        ws.cell(row=r, column=3).value = f"=COUNTA({pl}!C2:C100000)"
    r = next((rr for rr in range(1, ws.max_row + 1)
              if str(ws.cell(row=rr, column=2).value or "").startswith("Total signed exposure")), None)
    if r:
        ws.cell(row=r, column=3).value = f"=SUM({pl}!L:L)"
    r = next((rr for rr in range(1, ws.max_row + 1)
              if str(ws.cell(row=rr, column=2).value or "").startswith("GEO-GSO share")), None)
    if r:
        ws.cell(row=r, column=3).value = (f'=SUMIFS({pl}!L:L,{pl}!G:G,"*GEO*")/SUM({pl}!L:L)')
        ws.cell(row=r, column=3).number_format = "0%"


# --------------------------------------------------------------------------- #
#  Portfolio  →  bridge totals, UW-year table, GEO share                        #
# --------------------------------------------------------------------------- #
def linkify_portfolio(wb, sheet="Portfolio", per_layer="Per Layer", changes="Changes"):
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    pl = f"'{per_layer}'"
    have_changes = changes in wb.sheetnames

    # exposure bridge: Opening / +New / -Run-off / Reval / Closing in col D
    op = next((r for r in range(1, ws.max_row + 1)
               if str(ws.cell(row=r, column=2).value or "").startswith("Opening in-force")), None)
    if op:
        new_r, run_r, rev_r, close_r = op + 1, op + 2, op + 3, op + 4
        if have_changes:
            c_new = _find_row(wb[changes], 2, "New exposure")
            c_drop = _find_row(wb[changes], 2, "Dropped exposure")
            if c_new:
                ws.cell(row=new_r, column=4).value = f"=Changes!C{c_new}"
            if c_drop:
                ws.cell(row=run_r, column=4).value = f"=-Changes!C{c_drop}"
        ws.cell(row=close_r, column=4).value = f"=SUM(D{op}:D{rev_r})"
        for rr in (new_r, run_r, rev_r):
            ws.cell(row=rr, column=5).value = f"=IFERROR(D{rr}/$D${op},0)"

    # UW-year table: rows under 'UW Year' header — SUMPRODUCT on inception year
    def _row_startswith(prefix):
        return next((r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=2).value or "").startswith(prefix)), None)

    h = _find_row(ws, 2, "UW Year")
    if h:
        # find total cell ($C$4 style) = the 'Total on-risk exposure' value
        tot = _row_startswith("Total on-risk exposure")
        tref = f"$C${tot}" if tot else None
        inc = f"{pl}!$I$2:$I$100000"
        expo = f"{pl}!$L$2:$L$100000"
        r = h + 1
        while r <= ws.max_row and ws.cell(row=r, column=2).value not in (None, ""):
            yr = ws.cell(row=r, column=2).value
            ys = str(int(yr)) if isinstance(yr, (int, float)) else str(yr).strip()
            if len(ys) == 4 and ys.isdigit():
                cond = f'(LEFT({inc},4)="{ys}")'
                ws.cell(row=r, column=3).value = f"=SUMPRODUCT({cond}*1)"
                ws.cell(row=r, column=4).value = f"=SUMPRODUCT({cond}*{expo})"
                if tref:
                    ws.cell(row=r, column=5).value = f"=IFERROR(D{r}/{tref},0)"
            r += 1

    # concentration: GEO-GSO share of exposure
    g = next((r for r in range(1, ws.max_row + 1)
              if str(ws.cell(row=r, column=2).value or "").startswith("GEO-GSO share")), None)
    tot = _row_startswith("Total on-risk exposure")
    if g and tot:
        ws.cell(row=g, column=3).value = (
            f'=SUMIFS({pl}!L:L,{pl}!G:G,"*GEO*")/$C${tot}')
        ws.cell(row=g, column=3).number_format = "0.0%"


# --------------------------------------------------------------------------- #
#  Changes  →  Δ and Δ% columns (current − prior)                               #
# --------------------------------------------------------------------------- #
def linkify_changes(wb, sheet="Changes", summary="Summary",
                    per_layer="Per Layer", params="Parameters"):
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    # locate each block header where col D starts with 'Current' and col F with 'Δ'
    for r in range(1, ws.max_row + 1):
        d = str(ws.cell(row=r, column=4).value or "")
        f = str(ws.cell(row=r, column=6).value or "")
        if d.startswith("Current") and f.startswith("Δ"):
            rr = r + 1
            while rr <= ws.max_row and isinstance(ws.cell(row=rr, column=4).value, (int, float)):
                ws.cell(row=rr, column=6).value = f"=D{rr}-E{rr}"
                ws.cell(row=rr, column=7).value = f"=IFERROR((D{rr}-E{rr})/E{rr},0)"
                rr += 1

    # De-hardcode the 'Current' column of the by-scenario blocks by linking it to
    # the canonical Summary cascade (the current run IS Summary). 'Prior' stays a
    # value — the prior quarter is not a workbook tab. Only cells that equal
    # Summary are linked, so no number moves:
    #   • Current Net  → Summary!J for every entity (net ties across the book)
    #   • Current Gross → Summary!D for FUL / FIID only. FIHL & FIBL carry the
    #     raw (ex-add-on / pre-receiver) gross here, which deliberately differs
    #     from Summary's combined display, so those stay as values.
    if summary not in wb.sheetnames:
        return
    bases = _summary_bases(wb[summary])
    if "FIHL" not in bases:
        return

    def srow(ent, scen):
        b = bases.get(ent)
        return None if (b is None or scen not in SCEN_ORDER) else b + SCEN_ORDER.index(scen)

    for r in range(1, ws.max_row + 1):
        if not (str(ws.cell(row=r, column=2).value or "") == "Entity"
                and str(ws.cell(row=r, column=3).value or "") == "Scenario"):
            continue
        d = str(ws.cell(row=r, column=4).value or "")
        if not d.startswith("Current"):
            continue
        is_net = "Net" in d
        scol = "J" if is_net else "D"
        rr = r + 1
        while rr <= ws.max_row:
            ent = ws.cell(row=rr, column=2).value
            scen = ws.cell(row=rr, column=3).value
            sr = srow(ent, scen) if ent in bases else None
            if sr is None:
                break
            if is_net or ent in ("FUL", "FIID"):
                ws.cell(row=rr, column=4).value = f"=Summary!{scol}{sr}"
            rr += 1

    # ---- Aggregated-exposure 'Current' block → live SUMIFS on Per Layer -------
    # The 'Prior' block stays a value (external quarter, no tab). The 'Current'
    # block (by orbit × entity) is a live SUMIFS on Per Layer, with the derived
    # columns built from the orbit sums (Total = GEO+MEO+LEO, GD-eligible =
    # GEO+MEO, PF-eligible = GEO, SD orbit-weighted = GEO·dr+MEO·dr+LEO·dr from
    # the Parameters damage ratios). Arithmetic verified against the baked values.
    if per_layer not in wb.sheetnames:
        return
    PL = f"'{per_layer}'"
    ENT, ORB, PSC = "C", "G", "L"
    dr = {}
    if params in wb.sheetnames:
        pw = wb[params]
        dr["GEO"] = _label_cell_ref(pw, "Space Debris damage ratio — GEO-GSO")
        dr["MEO"] = _label_cell_ref(pw, "Space Debris damage ratio — MEO")
        dr["LEO"] = _label_cell_ref(pw, "Space Debris damage ratio — LEO")

    def _orb(ent, pat):
        return (f'SUMIFS({PL}!{PSC}:{PSC},{PL}!{ENT}:{ENT},"{ent}",'
                f'{PL}!{ORB}:{ORB},"{pat}")')

    for r in range(1, ws.max_row + 1):
        if not (str(ws.cell(row=r, column=2).value or "") == "Current"
                and str(ws.cell(row=r, column=3).value or "") == "GEO-GSO"):
            continue
        rows = {}
        rr = r + 1
        while rr <= ws.max_row:
            lbl = str(ws.cell(row=rr, column=2).value or "")
            if not lbl or lbl.startswith("Δ"):
                break
            rows[lbl.split()[0]] = rr
            rr += 1
        for key, rw in rows.items():
            if key in ("FUL", "FIID", "FIBL"):
                ws.cell(row=rw, column=3).value = f"={_orb(key, '*GEO*')}"
                ws.cell(row=rw, column=4).value = f"={_orb(key, '*MEO*')}"
                ws.cell(row=rw, column=5).value = f"={_orb(key, '*LEO*')}"
            elif key == "FIHL":                       # FIHL (FUL+FIID)
                a, b = rows.get("FUL"), rows.get("FIID")
                if a and b:
                    for col in (3, 4, 5):
                        L = _col_letter(col)
                        ws.cell(row=rw, column=col).value = f"={L}{a}+{L}{b}"
            elif key == "Whole":                      # Whole portfolio
                parts = [rows.get(e) for e in ("FUL", "FIID", "FIBL") if rows.get(e)]
                for col in (3, 4, 5):
                    L = _col_letter(col)
                    ws.cell(row=rw, column=col).value = "=" + "+".join(
                        f"{L}{x}" for x in parts)
            # derived columns (built from the orbit sums just set)
            ws.cell(row=rw, column=6).value = f"=C{rw}+D{rw}+E{rw}"   # Total
            ws.cell(row=rw, column=8).value = f"=C{rw}+D{rw}"         # GD eligible (GEO+MEO)
            ws.cell(row=rw, column=9).value = f"=C{rw}"               # PF eligible (GEO)
            if dr.get("GEO") and dr.get("MEO") and dr.get("LEO"):
                ws.cell(row=rw, column=7).value = (
                    f"=C{rw}*{dr['GEO']}+D{rw}*{dr['MEO']}+E{rw}*{dr['LEO']}")

    # ---- Exposure-bridge 'Check vs current book' + RPF-adjusted 'Current' -----
    # Both are pure current-book figures — link them so they can never go stale.
    #   Check vs current book = SUM of Per Layer per_sc (the whole book).
    #   RPF-adjusted 'Current' = SUMPRODUCT(per_sc × rpf) by orbit group; the
    #   Per Layer rpf column is itself a live formula, so this ties to the engine.
    n = wb[per_layer].max_row
    L_rng, K_rng, G_rng = (f"{PL}!{PSC}2:{PSC}{n}", f"{PL}!K2:K{n}",
                           f"{PL}!{ORB}2:{ORB}{n}")
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value or "") == "Check vs current book":
            ws.cell(row=r, column=3).value = f"=SUM({L_rng})"
    # RPF-adjusted block: header row has C='Prior', D='Current'
    for r in range(1, ws.max_row + 1):
        if not (str(ws.cell(row=r, column=3).value or "") == "Prior"
                and str(ws.cell(row=r, column=4).value or "") == "Current"):
            continue
        geo_meo = leo = None
        rr = r + 1
        while rr <= ws.max_row:
            lbl = str(ws.cell(row=rr, column=2).value or "")
            if lbl.startswith("GEO"):
                ws.cell(row=rr, column=4).value = (
                    f'=SUMPRODUCT((ISNUMBER(SEARCH("GEO",{G_rng}))'
                    f'+ISNUMBER(SEARCH("MEO",{G_rng})))*{L_rng}*{K_rng})')
                geo_meo = rr
            elif lbl == "LEO":
                ws.cell(row=rr, column=4).value = (
                    f'=SUMPRODUCT(ISNUMBER(SEARCH("LEO",{G_rng}))*{L_rng}*{K_rng})')
                leo = rr
            elif lbl == "Total":
                if geo_meo and leo:
                    ws.cell(row=rr, column=4).value = f"=D{geo_meo}+D{leo}"
                break
            rr += 1
        break

    # ---- Layer-movement exposure split → live (link to Book Movement) --------
    # New business / Renewed / in-progress / Lapsed all come from the same
    # split_movement Book Movement uses, so link each to its at-a-glance cell.
    # Guarded on Book Movement existing, so nothing can #REF.
    def _ch_row(prefix):
        for rr in range(1, ws.max_row + 1):
            if str(ws.cell(row=rr, column=2).value or "").startswith(prefix):
                return rr
        return None

    if "Book Movement" in wb.sheetnames:
        bm = wb["Book Movement"]

        def _bm_ref(needle):
            for rr in range(1, bm.max_row + 1):
                if needle in str(bm.cell(row=rr, column=2).value or ""):
                    return f"'Book Movement'!C{rr}"
            return None

        for ch_label, bm_needle in (
                ("New business exposure", "New business"),
                ("Renewed (rewritten) exposure", "current exposure"),
                ("Renewals in progress (unbound)", "Renewals in progress"),
                ("Lapsed exposure (off the book)", "Lapsed (not renewed")):
            cr, br = _ch_row(ch_label), _bm_ref(bm_needle)
            if cr and br:
                ws.cell(row=cr, column=3).value = f"={br}"


# --------------------------------------------------------------------------- #
#  Cover  →  Summary / Per Layer                                                #
# --------------------------------------------------------------------------- #
def linkify_cover(wb, sheet="Cover", summary="Summary", per_layer="Per Layer"):
    """De-hardcode the Cover 'key facts': on-risk layers, total signed exposure
    and the worst-case FIHL gross headline become live formulas so the cover
    always agrees with the sheets behind it. Values sit in col C beside a col-B
    label; located by label, so tolerant of layout shifts."""
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    pl = f"'{per_layer}'"

    def _val_row(label):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=2).value or "").strip() == label:
                return r
        return None

    if per_layer in wb.sheetnames:
        r = _val_row("On-risk layers")
        if r:
            ws.cell(row=r, column=3).value = f"=COUNTA({pl}!C2:C100000)"
            ws.cell(row=r, column=3).number_format = "#,##0"
        r = _val_row("Total signed exposure")
        if r:
            ws.cell(row=r, column=3).value = (
                f'="$"&TEXT(SUM({pl}!L:L)/1000000,"#,##0.0")&"m"')

    if summary in wb.sheetnames:
        bases = _summary_bases(wb[summary])
        fihl = bases.get("FIHL")
        r = _val_row("Worst-case (FIHL gross)")
        if r and fihl is not None:
            sr = fihl + SCEN_ORDER.index("Space Weather")
            ws.cell(row=r, column=3).value = (
                f'="$"&TEXT(Summary!D{sr}/1000000,"#,##0.0")&"m  ·  Space Weather"')


# --------------------------------------------------------------------------- #
#  Change Narrative  →  Summary (current side) ; prior stays a value            #
# --------------------------------------------------------------------------- #
def linkify_change_narrative(wb, sheet="Change Narrative", summary="Summary"):
    """De-hardcode the Change Narrative's current columns by linking them to the
    canonical Summary cascade (Cur Gross → Summary!D combined/receiver basis,
    Cur Net → Summary!J). Δ / Δ% are already live formulas off these cells.
    Prior stays a value — it's the frozen prior quarter, with no tab to point
    at (same as every other 'Prior' column in the book)."""
    if sheet not in wb.sheetnames or summary not in wb.sheetnames:
        return
    ws = wb[sheet]
    bases = _summary_bases(wb[summary])
    if "FIHL" not in bases:
        return

    def srow(ent, scen):
        b = bases.get(ent)
        return None if (b is None or scen not in SCEN_ORDER) else b + SCEN_ORDER.index(scen)

    cur_scen = None
    for r in range(1, ws.max_row + 1):
        b = str(ws.cell(row=r, column=2).value or "").strip()
        if b in SCEN_ORDER:
            cur_scen = b
            continue
        if b in ("FIHL", "FUL", "FIBL", "FIID") and cur_scen:
            sr = srow(b, cur_scen)
            if sr:
                ws.cell(row=r, column=4).value = f"=Summary!D{sr}"   # Cur Gross
                ws.cell(row=r, column=5).value = f"=Summary!J{sr}"   # Cur Net


# --------------------------------------------------------------------------- #
#  Summary  →  Per Layer (operating-entity additive-scenario gross only)        #
# --------------------------------------------------------------------------- #
def linkify_summary(wb, sheet="Summary", per_layer="Per Layer"):
    """Formula-drive the additive scenario grosses — Proton Flare / Generic
    Defect / Space Debris — for the operating entities FUL & FIID: each becomes
    a SUMIFS on the Per Layer scenario-loss column (itself per_sc × rate × RPF,
    linked to Parameters). Deliberately NOT touched, because they are not plain
    sums of a Per Layer column: FIHL (combined — the scenario loss also lands on
    the S3123/equity add-on exposure), FIBL (internal receiver), Space Weather
    (worst single manufacturer) and Max Risk (largest single spacecraft). Those
    stay the engine's reconciled values."""
    if sheet not in wb.sheetnames or per_layer not in wb.sheetnames:
        return
    ws = wb[sheet]
    pl = wb[per_layer]
    H = _header_cols(pl, 1)
    if not {"entity", "pf_fihl", "gd_fihl", "sd_fihl"} <= set(H):
        return
    PL = f"'{per_layer}'"
    ent_c = _col_letter(H["entity"])
    scen_col = {"Proton Flare": _col_letter(H["pf_fihl"]),
                "Generic Defect": _col_letter(H["gd_fihl"]),
                "Space Debris": _col_letter(H["sd_fihl"])}
    bases = _summary_bases(ws)
    linked = False
    for ent in ("FUL", "FIID"):
        base = bases.get(ent)
        if base is None:
            continue
        for scen, sc in scen_col.items():
            r = base + SCEN_ORDER.index(scen)
            ws.cell(row=r, column=4).value = (
                f'=SUMIFS({PL}!{sc}:{sc},{PL}!{ent_c}:{ent_c},"{ent}")')
            linked = True
    if linked:
        from openpyxl.styles import Font
        nr = ws.max_row + 2
        c = ws.cell(row=nr, column=2, value=(
            "Gross: FUL & FIID Proton Flare / Generic Defect / Space Debris are "
            "live SUMIFS on Per Layer (per_sc x rate x RPF). FIHL (combined), FIBL "
            "(receiver), Space Weather (worst manufacturer) and Max Risk (largest "
            "spacecraft) are the reconciled engine values — not plain per-layer sums."))
        c.font = Font(name="Calibri", size=8, italic=True, color="6B7785")


def linkify(wb):
    linkify_per_layer(wb)
    linkify_netting(wb)
    linkify_exec_summary(wb)
    linkify_portfolio(wb)
    linkify_changes(wb)
    linkify_cover(wb)
    linkify_change_narrative(wb)
    linkify_summary(wb)
