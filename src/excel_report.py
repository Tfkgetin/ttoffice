"""Formatted Excel results workbook — the pipeline's readable output.

Sheets:
  Control            run metadata, key totals, data-quality checks
  Summary            entity x scenario headline grid
  Portfolio          exposure profile (entity / orbit / UW year / manufacturer)
  Parameters         every assumption that produced the numbers
  Netting Waterfalls per entity x scenario step tables
  Space Weather      by-manufacturer ranking
  Max Risk           largest layers
  Per Layer          full engine output, filterable
"""
from __future__ import annotations
import datetime as dt
import pandas as pd
from . import charts_xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.drawing.colors import ColorChoice as DrawColorChoice
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.comments import Comment

INK = "1F2933"; ACCENT = "1B3A5C"; CREAM = "EEF3F8"; RULE = "DCE3EA"
SOFT = "6B7785"; GREEN = "2D6A3C"; AMBER = "9A6410"; TEAL = "1B3A5C"
NAVY = "1B3A5C"; GREENFILL = "E8F1EA"; TOTALFILL = "E4ECF4"; HDRTXT = "FFFFFF"

_FB = "Calibri"   # body font — modern, Excel-native
_FH = "Calibri"   # heading font

F_TITLE = Font(name=_FH, size=17, bold=True, color=NAVY)
F_SUB = Font(name=_FB, size=10, italic=False, color=SOFT)
F_SECT = Font(name=_FH, size=12, bold=True, color=NAVY)
F_HDR = Font(name=_FB, size=10, bold=True, color=HDRTXT)
F_CELL = Font(name=_FB, size=10, color=INK)
F_KEY = Font(name=_FB, size=10, bold=True, color=INK)
F_TOTAL = Font(name=_FB, size=10, bold=True, color=NAVY)
F_OK = Font(name=_FB, size=10, bold=True, color=GREEN)
F_FLAG = Font(name=_FB, size=10, bold=True, color=AMBER)
FILL_HDR = PatternFill("solid", start_color=NAVY)
FILL_WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
F_HDR_DARK = Font(name=_FB, size=10, bold=True, color=NAVY)  # dark text for white headers
BORDER_HDR_WHITE = Border(bottom=Side(style="medium", color=NAVY))
FILL_ALT = PatternFill("solid", start_color=CREAM)
FILL_TOTAL = PatternFill("solid", start_color=TOTALFILL)
THIN = Border(bottom=Side(style="thin", color=RULE))
MONEY = '#,##0;(#,##0);"–"'
PCT = '0.0%'
# Headline Δ% display: sign-driven ▲/▼ glyph (positive;negative;zero sections).
# Cells stay numeric — colour is applied by conditional formatting, not here.
DELTA_NUMFMT = '"▲ +"0.0%;"▼ "0.0%;"▲ +"0.0%'

# Renewal-policy verdict → font colour (src/renewals.py severities): data-gap and
# genuine loss read red, in-progress (manual-inclusion candidate) amber, renewed
# green, no-pointer grey.
RENEWAL_SEV_COLOUR = {"gap": "9B2D1F", "lost": "C0392B", "progress": AMBER,
                      "ok": GREEN, "neutral": SOFT}

SCEN_ORDER = ["Proton Flare", "Space Weather", "Generic Defect",
              "Space Debris", "Max Risk"]
# Netting-waterfall rows: (label, grid column, sign). Consumed by _waterfalls().
WF_ROWS = [("Gross Loss", "gross", 1), ("less: External QS", "ext_qs", -1),
           ("less: Other Ext RI", "other_ext_ri", -1),
           ("less: QS Ceded to FIBL", "igr_qs_ceded", -1),
           ("less: XoL Ceded to FIBL", "xol_ceded", -1),
           ("Net of XoL IGR (Retained)", "net", 1)]


def _display_gross(row):
    """Headline gross for a grid row: FIHL combined (incl add-ons),
    FIBL receiver, FUL/FIID standard. Matches the manual Change Narrative."""
    ent = row.get("entity")
    if ent == "FIHL" and "gross_incl_addons" in row.index:
        return float(row["gross_incl_addons"])
    if ent == "FIBL" and "gross_receiver" in row.index:
        return float(row["gross_receiver"])
    return float(row["gross"])


def _display_net(row):
    """Headline net for a grid row: FIHL combined, FIBL receiver, others standard."""
    ent = row.get("entity")
    if ent == "FIHL" and "net_incl_addons" in row.index:
        return float(row["net_incl_addons"])
    if ent == "FIBL" and "net_receiver" in row.index:
        return float(row["net_receiver"])
    v = row.get("net")
    return float(v) if pd.notna(v) else None


def _title(ws, title, sub):
    ws.sheet_view.showGridLines = False
    ws["B2"] = title; ws["B2"].font = F_TITLE
    ws["B3"] = sub; ws["B3"].font = F_SUB


def _section(ws, r, text):
    """Section header with a thin navy underline rule across the content width."""
    c = ws.cell(row=r, column=2, value=text)
    c.font = F_SECT
    # underline rule from B to G
    for col in range(2, 8):
        cell = ws.cell(row=r, column=col)
        cell.border = Border(bottom=Side(style="thin", color=NAVY))
    return r + 2


def _autofit(ws, min_w=8, max_w=60, pad=2):
    """Widen columns to fit their content so no cell shows ###. Respects any
    width already set as a floor; never shrinks below it."""
    from openpyxl.utils import get_column_letter as _gcl
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            # length of the displayed value (approx for numbers w/ thousands seps)
            if isinstance(cell.value, (int, float)):
                txt = f"{cell.value:,.0f}" if abs(cell.value) >= 1000 else str(cell.value)
            else:
                txt = str(cell.value)
            # only count the longest single line
            longest = max((len(s) for s in txt.split("\n")), default=0)
            widths[col] = max(widths.get(col, 0), longest)
    for col, w in widths.items():
        letter = _gcl(col)
        existing = ws.column_dimensions[letter].width or 0
        target = min(max(w + pad, min_w, existing), max_w)
        ws.column_dimensions[letter].width = target


def _print_setup(ws, params, title):
    """Senior-presentation print hygiene: fit to width, footer with as-at,
    classification, and page numbers. Applied to every sheet."""
    ws.print_options.horizontalCentered = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    ws.oddFooter.left.text = f"Space RDS · {params.quarter} · as at {params.as_at}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.center.text = "Internal — Confidential"
    ws.oddFooter.center.size = 8
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8


def _hdr(ws, r, col0, headers, widths=None, white=False):
    """Table header row. Default is a solid navy band with white text; pass
    white=True for a clean white header (navy bold text, medium navy underline)
    — used where a lighter, print-friendly look is wanted."""
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(col0 + i)].width = w
    for j, h in enumerate(headers):
        c = ws.cell(row=r, column=col0 + j, value=h)
        if white:
            c.font = F_HDR_DARK
            c.fill = FILL_WHITE
            c.border = BORDER_HDR_WHITE
        else:
            c.font = F_HDR
            c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="left" if j == 0 else "right",
                                vertical="center")
    ws.row_dimensions[r].height = 19
    return r + 1


def _fcell(ws, r, c, formula, alt=False, money=False, pct=False, bold=False):
    """A cell holding a live Excel FORMULA (not a baked value) so derived numbers
    recompute when their source cells change — used for Δ / Δ% / running totals
    that would otherwise go stale if a base cell is edited."""
    cell = ws.cell(row=r, column=c, value=formula)
    cell.font = F_TOTAL if bold else F_CELL
    cell.border = THIN
    if alt:
        cell.fill = FILL_ALT
    if money:
        cell.number_format = MONEY
    elif pct:
        cell.number_format = PCT
    cell.alignment = Alignment(horizontal="right")
    return cell


def _dcell(ws, r, c, val, alt=False):
    """Date cell (inception / off-risk). NaT → blank."""
    d = None
    if val is not None and pd.notna(val):
        try:
            d = pd.to_datetime(val).date()
        except (TypeError, ValueError):
            d = None
    cell = ws.cell(row=r, column=c, value=d)
    cell.font = F_CELL
    cell.border = THIN
    if alt:
        cell.fill = FILL_ALT
    if d is not None:
        cell.number_format = "dd-mmm-yy"
        cell.alignment = Alignment(horizontal="right")
    return cell


def _progid(v):
    """Render a programme id as a plain integer; NA/blank → ''."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return int(float(v))
    except (TypeError, ValueError):
        s = str(v).strip()
        return "" if s.lower() in ("nan", "<na>", "none") else s


def _add_table(ws, name, first_col, header_row, last_col, last_row):
    """Wrap a rendered block in an Excel Table so it carries its own filter
    dropdowns. Excel allows only ONE classic AutoFilter per sheet but many
    Tables — this is how all tables on a sheet get filters. Row stripes are off
    (we stripe manually) and the white header formatting applied by the caller
    wins over the table style. No-ops if there is no data row."""
    if last_row < header_row + 1:
        return
    ref = (f"{get_column_letter(first_col)}{header_row}:"
           f"{get_column_letter(last_col)}{last_row}")
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False)
    ws.add_table(tbl)


def _cell(ws, r, c, val, alt=False, money=False, pct=False, bold=False,
          font=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font or (F_TOTAL if bold else F_CELL)
    cell.border = THIN
    if alt:
        cell.fill = FILL_ALT
    if money and isinstance(val, (int, float)):
        cell.number_format = MONEY
        cell.alignment = Alignment(horizontal="right")
    elif pct and isinstance(val, (int, float)):
        cell.number_format = PCT
        cell.alignment = Alignment(horizontal="right")
    elif isinstance(val, (int, float)):
        cell.alignment = Alignment(horizontal="right")
    return cell


def _kv(ws, r, key, val, money=False):
    ws.cell(row=r, column=2, value=key).font = F_KEY
    c = ws.cell(row=r, column=3, value=val)
    c.font = F_CELL
    if money and isinstance(val, (int, float)):
        c.number_format = MONEY
    return r + 1


def _databar(ws, col_letter, r0, r1):
    ws.conditional_formatting.add(
        f"{col_letter}{r0}:{col_letter}{r1}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="B8CBDD", showValue=True))


def _breakdown(ws, r, label_field, group, total, label_w=30,
               crit_col=None, persc_col=None, total_ref=None):
    """Count/exposure/% table. If crit_col/persc_col given, Layers and Exposure
    become live COUNTIF/SUMIF formulas keyed on the label cell; % references the
    grand-total cell. Otherwise falls back to static values."""
    r = _hdr(ws, r, 2, [label_field, "Layers", "Exposure", "% of total"],
             [label_w, 11, 16, 12])
    start = r
    for k, (name, row) in enumerate(group.iterrows()):
        alt = k % 2 == 0
        lref = f"$B{r}"
        _cell(ws, r, 2, str(name), alt=alt)
        if crit_col and persc_col:
            c3 = _cell(ws, r, 3, f"=COUNTIF({crit_col},{lref})", alt=alt)
            c3.alignment = Alignment(horizontal="right")
            c4 = _cell(ws, r, 4, f"=SUMIF({crit_col},{lref},{persc_col})", alt=alt)
            c4.number_format = MONEY; c4.alignment = Alignment(horizontal="right")
            if total_ref:
                c5 = _cell(ws, r, 5, f"=IFERROR(D{r}/{total_ref},0)", alt=alt)
                c5.number_format = PCT; c5.alignment = Alignment(horizontal="right")
            else:
                _cell(ws, r, 5, float(row["exp"]) / total if total else 0,
                      alt=alt, pct=True)
        else:
            _cell(ws, r, 3, int(row["n"]), alt=alt)
            _cell(ws, r, 4, float(row["exp"]), alt=alt, money=True)
            _cell(ws, r, 5, float(row["exp"]) / total if total else 0,
                  alt=alt, pct=True)
        r += 1
    _databar(ws, "E", start, r - 1)
    return r + 1


# ---------- sheet builders ----------

# ---- Per Layer column map (for live formulas referencing the data backbone) ----
_PL_SHEET = "Per Layer"
_PL_KEEP = ["program_id", "layer_id", "entity", "mapping_code", "spacecraft_id",
            "spacecraft_name", "orbit", "bus_manufacturer", "inception",
            "off_risk_date", "rpf", "per_sc", "ext_qs", "s3123_qs", "equity_usd",
            # engine's own per-layer S3123 basis — so linkify can compute s3123_qs
            # / equity EXACTLY (3-tier factor + full eligibility) instead of a
            # brittle 2-tier date reconstruction:
            "s3123_eligible", "s3123_factor", "s2126_factor", "equity_pct",
            "igr_qs_rate", "igr_qs_ceded", "net_of_qs", "xol_ceded",
            "net_of_xol", "pf_fihl", "gd_fihl", "sd_fihl", "total_fibl_ceded",
            # per-entity scenario contributions + per-$ cession rates: the
            # backbone the Summary & S3123 tabs SUMPRODUCT against so those
            # aggregates are live formulas, not baked numbers (additive scenarios).
            "pf_ful", "pf_fiid", "gd_ful", "gd_fiid", "sd_ful", "sd_fiid",
            "ext_qs_pp", "igr_ceded_pp", "s3123_qs_pp", "equity_pp",
            # per-layer S3123 / S2126 syndicate contributions per RDS. The _g
            # (gross) and _n (net) cells are LIVE formulas built from per_sc x
            # factor x _elig x loss x rpf and the QS cession; _elig (share gate)
            # and _sel (in-scenario membership) are the only engine-set inputs.
            # The Lloyd's Summary tabs SUM the _g/_n columns.
            "s3123_rds_elig",
            "s3123_pf_sel", "s3123_pf_g", "s3123_pf_n",
            "s3123_sw_sel", "s3123_sw_g", "s3123_sw_n",
            "s3123_gd_sel", "s3123_gd_g", "s3123_gd_n",
            "s3123_sd_sel", "s3123_sd_g", "s3123_sd_n",
            "s2126_rds_elig",
            "s2126_pf_sel", "s2126_pf_g", "s2126_pf_n",
            "s2126_sw_sel", "s2126_sw_g", "s2126_sw_n",
            "s2126_gd_sel", "s2126_gd_g", "s2126_gd_n",
            "s2126_sd_sel", "s2126_sd_g", "s2126_sd_n"]

# additive scenarios only — selections (Space Weather / Max Risk) are not sums
_SP_SCEN = {"Proton Flare": "pf", "Generic Defect": "gd", "Space Debris": "sd"}
_SP_ENT = {"FIHL": "fihl", "FUL": "ful", "FIID": "fiid"}


def _pl_range(colmap, n, field):
    """Finite Per Layer data range $X$2:$X$n+1 for a mapped field, else None."""
    col = colmap.get(field)
    return f"'{_PL_SHEET}'!${col}$2:${col}${n + 1}" if col else None


def _sumproduct(colmap, n, *fields):
    """=SUMPRODUCT over Per Layer of the given fields, or None if any is absent."""
    rngs = [_pl_range(colmap, n, f) for f in fields]
    if any(x is None for x in rngs):
        return None
    return "=SUMPRODUCT(" + ",".join(rngs) + ")"


def _pl_map(per_layer):
    """field -> Excel column letter on the Per Layer sheet (live, not hardcoded)."""
    keep = [c for c in _PL_KEEP if c in per_layer.columns]
    return {f: get_column_letter(i + 1) for i, f in enumerate(keep)}, keep


def _col(colmap, field):
    """Full-column ref e.g. \'Per Layer\'!L:L for SUMIF/COUNTIF criteria."""
    return f"'{_PL_SHEET}'!{colmap[field]}:{colmap[field]}"


def _note(cell, text):
    """Attach a reconciliation hover-note to a cell (engine value / build-up)."""
    if cell is None:
        return
    c = Comment(text, "Reconcile")
    c.width, c.height = 320, 150
    cell.comment = c



def _control(wb, per_layer, params, source, recon, colmap):
    ws = wb.create_sheet("Control")
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 42
    _title(ws, f"Space RDS · {params.quarter} · run control",
           f"as-at {params.as_at} · source: {source} · "
           f"generated {dt.datetime.now():%Y-%m-%d %H:%M}")
    total = per_layer["per_sc"].sum()
    persc = _col(colmap, "per_sc")
    orbit = _col(colmap, "orbit")
    rpf = _col(colmap, "rpf")
    pid = _col(colmap, "program_id")

    def _kvf(r, key, formula, money=True):
        ws.cell(row=r, column=2, value=key).font = F_KEY
        c = ws.cell(row=r, column=3, value=formula)
        c.font = F_CELL
        if money:
            c.number_format = MONEY
        return r + 1

    r = _section(ws, 5, "Run")
    r = _kv(ws, r, "Quarter", params.quarter)
    r = _kv(ws, r, "As-at date", str(params.as_at))
    r = _kv(ws, r, "Ingest source", source)
    r = _kvf(r, "Total signed exposure (per-S/C basis)", f"=SUM({persc})")
    r = _kvf(r, "Layers", f"=COUNTA({pid})-1", money=False)
    r = _kv(ws, r, "Spacecraft", int(per_layer["spacecraft_id"].nunique()))
    r = _kv(ws, r, "Programs", int(per_layer["program_id"].nunique()))
    r += 1

    r = _section(ws, r, "Data quality")
    r = _hdr(ws, r, 2, ["Check", "Count", "Note"], [34, 10, 40])

    def _flag(label, value, note, ok_if_zero=True, n_actual=None):
        nonlocal r
        _cell(ws, r, 2, label)
        c = _cell(ws, r, 3, value)
        nval = n_actual if n_actual is not None else (value if isinstance(value, int) else 0)
        c.font = (F_OK if nval == 0 else F_FLAG) if ok_if_zero else F_CELL
        c.alignment = Alignment(horizontal="right")
        _cell(ws, r, 4, note)
        r += 1

    null_orbit = int((per_layer["orbit"].astype(str).isin(["NULL", "None", "nan"])).sum())
    off_risk = int((per_layer["rpf"] == 0).sum())
    bad_exp = int((per_layer["per_sc"] <= 0).sum())
    _flag("Rows ingested", f"=COUNTA({pid})-1", "total layer-rows",
          ok_if_zero=False, n_actual=len(per_layer))
    _flag("NULL orbit rows",
          f'=COUNTIF({orbit},"NULL")+COUNTIF({orbit},"None")',
          "excluded from all scenario scopes", n_actual=null_orbit)
    _flag("Layers off-risk at as-at (RPF=0)", f"=COUNTIF({rpf},0)",
          "zero time-decay factor", n_actual=off_risk)
    _flag("Zero/negative exposure rows", f'=COUNTIF({persc},"<=0")',
          "should be 0", n_actual=bad_exp)
    if recon is not None and "status" in getattr(recon, "columns", []):
        n_bad = int((recon["status"] == "MISMATCH").sum())
        _flag("Reconciliation mismatches", n_bad,
              "vs workbook (workbook source only)", n_actual=n_bad)
    return ws


def _summary(ws, grid, params, source, per_layer, addon_refs=None):
    # rowmap: (entity, scenario) -> dict of Summary A1 cell refs, so downstream
    # sheets (e.g. the RDS Input Template) can point at these cells as live
    # formulas instead of re-emitting hard-coded values. Populated below.
    rowmap = {}
    _plc, _ = _pl_map(per_layer)        # Per Layer column letters for SUMPRODUCT
    _plN = len(per_layer)               # data rows (2 .. _plN+1)
    _title(ws, f"Space RDS · {params.quarter}",
           f"as-at {params.as_at} · source: {source} · "
           f"{len(per_layer)} rows · netting cascade = live formulas")
    headers = ["Scenario", "Detail", "Gross", "Ext QS", "Net of Ext QS",
               "IGR QS Ceded", "Net of QS", "XoL Ceded", "Net", "Net/Gross"]
    widths = [16, 34, 14, 13, 14, 14, 13, 12, 13, 11]
    # DECIDED (user, 2026Q1): FIHL headline is COMBINED (incl S3123 QS + IG
    # equity) across the whole book — Summary, movement matrix and QA all on
    # this basis, matching the manual's Change Narrative / Exec Summary. The
    # ex-add-on figure and the split are kept as memo columns and on the
    # "S3123 & Equity (IG)" tab. FUL/FIID are unaffected (no add-ons).
    FIHL_MEMO = [("gross", "Gross ex add-ons"),
                 ("s3123_qs", "S3123 QS to IG"),
                 ("equity_usd", "IG Equity")]
    r = 5
    for entity in ["FIHL", "FUL", "FIID", "FIBL"]:
        # FIBL renders on the receiver (combined) basis — its own block.
        if entity == "FIBL" and "gross_receiver" in grid.columns:
            r = _fibl_receiver_block(ws, r, grid, rowmap)
            continue
        ws.cell(row=r, column=2, value=entity).font = F_SECT
        _desc = {"FIHL": "group / consolidated — COMBINED (incl S3123 QS + IG "
                         "Equity), manual Change Narrative basis; ex-add-on "
                         "split in the memo columns and the S3123 & Equity (IG) tab",
                 "FUL": "operating entity — cedes down to FIBL",
                 "FIID": "operating entity — cedes down to FIBL"}
        ws.cell(row=r, column=3, value=_desc.get(entity, "")).font = F_SUB
        _memo_here = (entity == "FIHL" and
                      any(k in grid.columns for k, _ in FIHL_MEMO))
        hdrs = headers + ([lbl for k, lbl in FIHL_MEMO if k in grid.columns]
                          if _memo_here else [])
        wds = widths + ([16, 14, 12][:len(hdrs) - len(headers)])
        r = _hdr(ws, r + 1, 2, hdrs, wds)
        block = grid[grid["entity"] == entity].copy()
        block["_o"] = block["scenario"].map({s: i for i, s in enumerate(SCEN_ORDER)})
        block = block.sort_values("_o")
        for k, (_, row) in enumerate(block.iterrows()):
            alt = k % 2 == 0
            def val(f):
                v = row.get(f)
                return None if (isinstance(v, float) and pd.isna(v)) else v
            # input cells (engine-computed scenario values)
            _cell(ws, r, 2, val("scenario"), alt=alt)
            _cell(ws, r, 3, val("detail") or "", alt=alt)
            is_fihl = (entity == "FIHL")
            addons = (float(val("s3123_qs") or 0) + float(val("equity_usd") or 0)) \
                if is_fihl else 0.0
            # Gross: FIHL shows COMBINED (ex-add-on gross + S3123 QS + equity);
            # the ex-add-on figure sits in the memo columns. FUL/FIID unchanged.
            gross_disp = (float(val("gross") or 0) + addons) if is_fihl else val("gross")
            cD = _cell(ws, r, 4, gross_disp, alt=alt, money=True)
            if not is_fihl:  # FIHL D is overwritten with =SUM(L:N) below + noted there
                _note(cD, "Gross = worst-case scenario aggregate over the Per Layer "
                          "tab: sum of eligible per-layer exposure (per_sc) x the "
                          "scenario factor. Reconcile on Per Layer / Netting "
                          "Waterfalls.")
            # contribution column for this entity x scenario (additive scenarios)
            _scn = val("scenario")
            _contrib = (f"{_SP_SCEN[_scn]}_{_SP_ENT[entity]}"
                        if _scn in _SP_SCEN and entity in _SP_ENT else None)
            # Ext QS — LIVE SUMPRODUCT for additive scenarios; engine value + note
            # for Space Weather / Max Risk (selections, not per-layer sums).
            _extf = _sumproduct(_plc, _plN, _contrib, "ext_qs_pp") if _contrib else None
            if _extf:
                cE = _cell(ws, r, 5, _extf, alt=alt, money=True)
                _note(cE, "LIVE: SUMPRODUCT over Per Layer of (this entity/scenario "
                          "per-layer contribution x per-layer external-QS rate = "
                          "ext_qs/per_sc). Ties to the engine to the cent; edit Per "
                          "Layer and it recomputes. (Outwards RI is per-layer by "
                          "inception slot - see Parameters -> External Quota Share.)")
            else:
                cE = _cell(ws, r, 5, val("ext_qs"), alt=alt, money=True)
                _note(cE, "Selection scenario (Space Weather / Max Risk): external QS "
                          "on the selected manufacturer / bird, not a per-layer sum, "
                          "so shown as the engine value. Reconcile on the Space "
                          "Weather / Max Risk tab and Per Layer (ext_qs).")
            # Net of Ext QS = Gross - Ext QS  (live)
            f6 = f"=D{r}-E{r}" if val("net_of_ext_qs") is not None else None
            _cell(ws, r, 6, f6, alt=alt, money=True)
            # IGR QS — LIVE SUMPRODUCT for FUL/FIID additive scenarios; 0 at FIHL
            # group view; engine value + note for the selection scenarios.
            _igrf = (_sumproduct(_plc, _plN, _contrib, "igr_ceded_pp")
                     if (_contrib and not is_fihl) else None)
            if is_fihl:
                cG = _cell(ws, r, 7, val("igr_qs_ceded"), alt=alt, money=True)
                _note(cG, "0 at the FIHL group view: the intra-group QS is ceded by "
                          "FUL & FIID and RECEIVED by FIBL, so it nets out at group "
                          "level. See the FUL / FIID rows and the FIBL receiver block.")
            elif _igrf:
                cG = _cell(ws, r, 7, _igrf, alt=alt, money=True)
                _note(cG, "LIVE: SUMPRODUCT over Per Layer of (contribution x per-layer "
                          "IGR-QS rate = igr_qs_ceded/per_sc). Ties to the engine to "
                          "the cent. Rate on Parameters -> IGR " + entity + " QS; "
                          "drill on Per Layer (igr_qs_ceded).")
            else:
                cG = _cell(ws, r, 7, val("igr_qs_ceded"), alt=alt, money=True)
                _note(cG, "Selection scenario: IGR QS on the selected manufacturer / "
                          "bird, not a per-layer sum. Engine value; reconcile on the "
                          "Space Weather / Max Risk tab and Per Layer (igr_qs_ceded).")
            # Net of QS = Net of Ext QS - IGR QS Ceded  (live)
            f8 = f"=F{r}-G{r}" if val("net_of_qs") is not None else None
            _cell(ws, r, 8, f8, alt=alt, money=True)
            cI = _cell(ws, r, 9, val("xol_ceded"), alt=alt, money=True)
            _note(cI, f"IGR XoL recovery - 0 unless this scenario's net-of-QS "
                      f"pierces the {entity} attachment. Terms (excess / limit) on "
                      f"Parameters -> IGR {entity} XoL. Drill: Netting Waterfalls.")
            # Net: FIHL net = (net of QS − XoL) already carries add-ons because
            # for FIHL the cascade's IGR QS/XoL are zero and Gross is combined,
            # so H−I = combined net. FUL/FIID: standard H−I.
            f10 = f"=H{r}-I{r}" if val("net") is not None else None
            c10 = _cell(ws, r, 10, f10, alt=alt, money=True, bold=(f10 is not None))
            # Net/Gross ratio (live) — always write the cell so the row fill is
            # complete; show ratio where a net exists, blank (but filled) otherwise.
            if val("net") is not None:
                c11 = _cell(ws, r, 11, f'=IFERROR(J{r}/D{r},"")', alt=alt)
                c11.number_format = PCT
                c11.alignment = Alignment(horizontal="right")
            else:
                _cell(ws, r, 11, None, alt=alt)
            # FIX(recon 2026Q1, C1): FIHL ex-add-on memo (manual convention)
            equity_ref = None
            if _memo_here:
                memo_present = [m for m in FIHL_MEMO if m[0] in grid.columns]
                # The split lives once on the "S3123 & Equity (IG)" tab; point
                # each memo cell there as a live reference so it is not a baked
                # number. Fall back to the raw value if refs aren't supplied.
                _AKEY = {"gross": "gross_ex", "s3123_qs": "s3123_qs",
                         "equity_usd": "equity"}
                aref = (addon_refs or {}).get("fihl", {}).get(val("scenario"))
                asheet = (addon_refs or {}).get("sheet")
                for j, (k, _lbl) in enumerate(memo_present):
                    ref = aref.get(_AKEY[k]) if aref else None
                    if ref and asheet:
                        _cell(ws, r, 12 + j, f"='{asheet}'!{ref}", alt=alt, money=True)
                    else:
                        _cell(ws, r, 12 + j, val(k), alt=alt, money=True)
                    if k == "equity_usd":
                        equity_ref = f"{get_column_letter(12 + j)}{r}"
                # FIHL headline Gross = ex-add-on gross + S3123 QS + IG equity —
                # the memo columns just written. Make col D their live SUM so the
                # combined build-up is transparent, not a baked number.
                if memo_present:
                    last = get_column_letter(11 + len(memo_present))
                    cDf = _cell(ws, r, 4, f"=SUM(L{r}:{last}{r})", alt=alt, money=True)
                    _note(cDf, "FIHL headline = COMBINED Exec basis = ex-add-on "
                               "gross (L) + S3123 QS to IG (M) + IG Equity (N), the "
                               "memo cells to the right. Those three live once on "
                               "the 'S3123 & Equity (IG)' tab (L:N reference it). "
                               "The ex-add-on gross reconciles as the worst-case "
                               "aggregate over Per Layer.")
            # capture this row's cell refs for downstream live-formula sheets
            rowmap[(entity, val("scenario"))] = {
                "gross": f"D{r}", "ext_qs": f"E{r}", "igr_qs": f"G{r}",
                "xol": f"I{r}", "net": f"J{r}", "equity": equity_ref}
            r += 1
        r += 1
    _summary_heat(ws, rowmap)
    return rowmap


def _summary_heat(ws, rowmap):
    """Signal-first conditional formatting on the Summary grid:
      · Gross (col D)     — in-cell data bars so loss magnitude reads at a glance
      · Net/Gross (col K) — green→amber→red scale so retention/cede-down is visible
    Ranges are taken from the captured rows, so no header/blank rows are touched."""
    rows = sorted(int(ref["net"][1:]) for ref in rowmap.values())
    if not rows:
        return
    lo, hi = rows[0], rows[-1]
    ws.conditional_formatting.add(
        f"D{lo}:D{hi}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="8FB0CF", showValue=True))
    # high retention (net/gross → 1) = red (more risk kept); low = green
    ws.conditional_formatting.add(
        f"K{lo}:K{hi}",
        ColorScaleRule(start_type="num", start_value=0, start_color="2E7D32",
                       mid_type="num", mid_value=0.5, mid_color="F2C14E",
                       end_type="num", end_value=1, end_color="C0392B"))


def _fibl_receiver_block(ws, r, grid, rowmap=None):
    """FIBL on the internal-receiver basis (manual Change Narrative convention),
    using the SAME column layout as the other entities (Gross in column D, Net
    in column J) so the figures are visually aligned. The five receiver
    components appear as memo columns to the right."""
    ws.cell(row=r, column=2, value="FIBL").font = F_SECT
    ws.cell(row=r, column=3,
            value="internal RI receiver — cessions in from FUL & FIID, plus "
                  "the IG add-ons (Change Narrative convention)").font = F_SUB
    main_headers = ["Scenario", "Detail", "Gross", "Ext QS", "Net of Ext QS",
                    "IGR QS Ceded", "Net of QS", "XoL Ceded", "Net", "Net/Gross"]
    memo_headers = ["QS received", "XoL received", "Direct book",
                    "S3123 QS to IG", "IG Equity"]
    hdrs = main_headers + memo_headers
    wds = [16, 22, 14, 13, 14, 14, 13, 12, 13, 11, 14, 13, 12, 14, 12]
    r = _hdr(ws, r + 1, 2, hdrs, wds)
    block = grid[grid["entity"] == "FIBL"].copy()
    block["_o"] = block["scenario"].map({s: i for i, s in enumerate(SCEN_ORDER)})
    block = block.sort_values("_o")
    for k, (_, row) in enumerate(block.iterrows()):
        alt = k % 2 == 0

        def val(f):
            v = row.get(f)
            return None if (isinstance(v, float) and pd.isna(v)) else v
        # main cascade: Gross = receiver total, Ext QS = direct-book leakage,
        # then standard formulas through to Net
        recv_gross = _display_gross(row)
        recv_net = _display_net(row)
        _cell(ws, r, 2, val("scenario"), alt=alt)
        _cell(ws, r, 3, val("detail") or "", alt=alt)
        _cell(ws, r, 4, recv_gross, alt=alt, money=True)
        # Ext QS on FIBL's direct book (its only leakage); 0 for most scenarios
        _cell(ws, r, 5, float(val("ext_qs") or 0) or None, alt=alt, money=True)
        f6 = f"=D{r}-E{r}" if recv_gross else None
        _cell(ws, r, 6, f6, alt=alt, money=True)
        _cell(ws, r, 7, None, alt=alt, money=True)          # no IGR on FIBL
        f8 = f"=F{r}-G{r}" if recv_gross else None
        _cell(ws, r, 8, f8, alt=alt, money=True)
        _cell(ws, r, 9, None, alt=alt, money=True)          # no XoL on FIBL
        f10 = f"=H{r}-I{r}" if recv_gross else None
        c10 = _cell(ws, r, 10, f10, alt=alt, money=True, bold=(f10 is not None))
        if recv_gross:
            c11 = _cell(ws, r, 11, f'=IFERROR(J{r}/D{r},"")', alt=alt)
            c11.number_format = PCT
            c11.alignment = Alignment(horizontal="right")
        else:
            _cell(ws, r, 11, None, alt=alt)
        # memo: receiver components
        qs = float(val("qs_received") or 0)
        xol = float(val("xol_received") or 0)
        direct = float(val("direct_fibl") or 0)
        s3 = float(val("s3123_qs") or 0)
        eq = float(val("equity_usd") or 0)
        for j, v in enumerate([qs, xol, direct, s3, eq]):
            _cell(ws, r, 12 + j, v or None, alt=alt, money=True)
        if rowmap is not None:
            rowmap[("FIBL", val("scenario"))] = {
                "gross": f"D{r}", "ext_qs": f"E{r}", "igr_qs": None,
                "xol": None, "net": f"J{r}", "equity": f"{get_column_letter(16)}{r}"}
        r += 1
    ws.cell(row=r, column=2,
            value="Gross = receiver total (QS in + XoL in + direct + S3123 + "
                  "equity); components in the memo columns. XoL only in Space "
                  "Weather (clearing attachment); direct only in Space Debris "
                  "(FIBL's own LEO book).").font = F_SUB
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 26
    return r + 2


def _rds_input_template(wb, grid, s3grid, params, summary_map=None):
    """Regulator RDS Input Template block — one section per scenario, in the
    EXACT row order and entity-column order of JJ's RDS_Input_Template .xlsm
    'Scenario' tab, so the user can copy a scenario block and paste it straight
    into the yellow cells with no reformatting.

    Row -> pipeline value map (verified to the dollar vs the 2026Q1 filed
    template on JJ's data):
      Gross Loss (1)          FIHL combined · FUL/FIID gross · FIBL receiver · S3123 gross
      Net Loss (2)            FIHL net-incl · FUL/FIID net · FIBL net-recv · S3123 net
      Net Loss inc RIPS (3)   = Net Loss (no RIPS in Space book)
      Outwards QS Recovery(10) ext_qs (FIHL/FUL/FIID); S3123 gross-net; FIBL 0
      Outwards QS RIPS (22)   0
      Outwards XOL Recovery(9) 0 (external XoL not used)
      Outwards XOL RIPS (11)  0
      IGR QS Recovery (5)     igr_qs_ceded (FUL, FIID only)
      IGR XOL Recovery (7)    xol_ceded (FUL, FIID only)
      Facultative RI Rec (17) 0    Facultative RI RIPS (18) 0
      IGR QS RIPS (19) blank  IGR XOL RIPS (20) blank
      Inwards RIPS (14) 0
      Specialty Re Addon Gross/Net/Final (13/15/16) 0
      IG Equity Share % (12/24) equity_usd / net-loss (FIHL & FIBL only)
    """
    ws = wb.create_sheet("RDS Input Template")
    _title(ws, "RDS Input Template — paste-ready",
           "One block per scenario, identical row/column order to the regulator "
           "template's 'Scenario' tab. IG-entity figures are LIVE formulas linked "
           "to the Summary tab (not hard-coded); when filing, paste-special → "
           "Values into the regulator's yellow cells. Verified to JJ's 2026Q1 "
           "filed template to the dollar.")

    ENTS = ["FIHL", "FUL", "FIBL", "FIID", "S3123", "S2126"]
    # (label, id) in the template's exact order
    ROWS = [
        ("Gross Loss", 1), ("Net Loss", 2), ("Net Loss inc RIPS", 3),
        ("Outwards QS Recovery", 10), ("Outwards QS RIPS", 22),
        ("Outwards XOL Recovery", 9), ("Outwards XOL RIPS", 11),
        ("IGR QS Recovery", 5), ("IGR XOL Recovery", 7),
        ("Facultative RI Recovery", 17), ("Facultative RI RIPS", 18),
        ("IGR QS RIPS", 19), ("IGR XOL RIPS", 20), ("Inwards RIPS", 14),
        ("Specialty Re Addon Gross", 13), ("Specialty Re Addon Net", 15),
        ("Specialty Re Addon Final Net", 16),
        ("IG Equity Share of S3123 (% of Net Loss xRIPS)", 12),
        ("IG Equity Share of S3123 (% of Net RIPS)", 24),
    ]

    def gval(scen, ent, field):
        b = grid[(grid["entity"] == ent) & (grid["scenario"] == scen)]
        if not len(b):
            return None
        return b.iloc[0].get(field)

    def s3val(scen, field):
        if s3grid is None:
            return None
        b = s3grid[s3grid["scenario"] == scen]
        return float(b.iloc[0][field]) if len(b) else None

    def cell_value(label, rid, scen, ent):
        # returns (value, is_percent, is_blank)
        if ent == "S2126":
            return None, False, True   # not in the Space book
        # S3123 column
        if ent == "S3123":
            g = s3val(scen, "gross"); n = s3val(scen, "net")
            if label == "Gross Loss": return g, False, False
            if label in ("Net Loss", "Net Loss inc RIPS"): return n, False, False
            if label == "Outwards QS Recovery":
                return (g - n) if (g is not None and n is not None) else None, False, False
            if label in ("IGR QS Recovery", "IGR XOL Recovery"):
                return None, False, True   # grey in template
            if label in ("IG Equity Share of S3123 (% of Net Loss xRIPS)",
                         "IG Equity Share of S3123 (% of Net RIPS)"):
                return None, False, True
            return 0.0, False, False
        # IG entities
        row = grid[(grid.entity == ent) & (grid.scenario == scen)]
        row = row.iloc[0] if len(row) else None
        if row is None:
            return None, False, True
        if label == "Gross Loss":
            return _display_gross(row), False, False
        if label in ("Net Loss", "Net Loss inc RIPS"):
            return _display_net(row), False, False
        if label == "Outwards QS Recovery":
            return float(row.get("ext_qs") or 0), False, False
        if label == "IGR QS Recovery":
            if ent in ("FUL", "FIID"):
                return float(row.get("igr_qs_ceded") or 0), False, False
            return None, False, True    # grey for FIHL/FIBL
        if label == "IGR XOL Recovery":
            if ent in ("FUL", "FIID"):
                return float(row.get("xol_ceded") or 0), False, False
            return None, False, True
        if label in ("IGR QS RIPS", "IGR XOL RIPS"):
            return None, False, True
        if label in ("IG Equity Share of S3123 (% of Net Loss xRIPS)",
                     "IG Equity Share of S3123 (% of Net RIPS)"):
            if ent in ("FIHL", "FIBL"):
                eq = float(row.get("equity_usd") or 0)
                net = _display_net(row) or 0
                return (eq / net) if net else 0.0, True, False
            return None, False, True
        # everything else (RIPS, facultative, specialty add-ons) = 0
        return 0.0, False, False

    def cell_formula(label, scen, ent):
        """Live 'Summary'! reference reproducing cell_value for the IG entities,
        so the template is not hard-coded. Returns None to fall back to the value
        (S3123/S2126 columns, the 0 rows, and grey cells stay as-is)."""
        if not summary_map:
            return None
        m = summary_map.get((ent, scen))
        if not m:
            return None
        S = "'Summary'!"
        if label == "Gross Loss":
            return f"={S}{m['gross']}"
        if label in ("Net Loss", "Net Loss inc RIPS"):
            return f"={S}{m['net']}"
        if label == "Outwards QS Recovery":
            return f"={S}{m['ext_qs']}"
        if label == "IGR QS Recovery" and ent in ("FUL", "FIID") and m["igr_qs"]:
            return f"={S}{m['igr_qs']}"
        if label == "IGR XOL Recovery" and ent in ("FUL", "FIID") and m["xol"]:
            return f"={S}{m['xol']}"
        if (label in ("IG Equity Share of S3123 (% of Net Loss xRIPS)",
                      "IG Equity Share of S3123 (% of Net RIPS)")
                and ent in ("FIHL", "FIBL") and m["equity"]):
            return f"=IFERROR({S}{m['equity']}/{S}{m['net']},0)"
        return None

    r = 5
    for scen in SCEN_ORDER:
        ws.cell(row=r, column=2, value=scen).font = F_SECT
        r += 1
        # header row: blank, ID, then entities
        _cell(ws, r, 2, "Line", bold=True); _cell(ws, r, 3, "ID", bold=True)
        for j, e in enumerate(ENTS):
            c = _cell(ws, r, 4 + j, e, bold=True); c.fill = PatternFill(
                "solid", start_color=ACCENT)
            c.font = Font(name=_FB, size=10, bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="right")
        ws.column_dimensions["B"].width = 34
        for col in "DEFGHI":
            ws.column_dimensions[col].width = 16
        r += 1
        for k, (label, rid) in enumerate(ROWS):
            alt = k % 2 == 0
            bold = label in ("Gross Loss", "Net Loss", "Net Loss inc RIPS")
            _cell(ws, r, 2, label, alt=alt, bold=bold)
            _cell(ws, r, 3, rid, alt=alt)
            for j, e in enumerate(ENTS):
                v, is_pct, is_blank = cell_value(label, rid, scen, e)
                f = cell_formula(label, scen, e)   # live 'Summary'! ref or None
                if is_blank:
                    c = _cell(ws, r, 4 + j, None, alt=alt)
                    c.fill = PatternFill("solid", start_color="E8E8E8")
                elif is_pct:
                    c = _cell(ws, r, 4 + j, f if f is not None else v, alt=alt)
                    c.number_format = '0.00000000'
                    c.alignment = Alignment(horizontal="right")
                elif f is not None:
                    # live formula: set money format explicitly (a formula string
                    # isn't numeric, so _cell won't apply it automatically)
                    c = _cell(ws, r, 4 + j, f, alt=alt, bold=bold)
                    c.number_format = MONEY
                    c.alignment = Alignment(horizontal="right")
                else:
                    _cell(ws, r, 4 + j, v, alt=alt, money=True, bold=bold)
            r += 1
        r += 2
    ws.cell(row=r, column=2,
            value="Grey cells are not entered for the Space book (S2126 absent; "
                  "IGR rows apply to FUL/FIID only; IG Equity % to FIHL/FIBL "
                  "only). Values are live-data; on a frozen-extract run they tie "
                  "the filed template to the dollar.").font = F_SUB
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 28
    return ws


def _exposure_bridge(ws, r, changes, closing_total):
    """Walk from prior in-force exposure to current, renewal-aware.

    Answers 'why did the book move?' — Opening + New business + Renewals
    (reprice) − Lapsed ± Revaluation = Closing. New / Renewed / Lapsed are split
    by the SAME spacecraft-level pairing Book Movement uses
    (renewals.split_movement), so a renewal carried forward onto a NEW programme
    id (e.g. Eutelsat 344558 → 385575 via roll-forward) nets out as a RENEWAL —
    it is NOT double-counted as run-off on the old id plus new business on the
    successor. All figures are on-risk per-S/C exposure (gross signed).
    """
    from . import renewals as _rnw
    L = changes["layers"]
    sp = _rnw.split_movement(L.get("new"), L.get("dropped"))

    def _x(df):
        return float(pd.to_numeric(df.get("per_sc"), errors="coerce").fillna(0).sum()) \
            if df is not None and len(df) else 0.0

    def _n(df):
        return int(len(df)) if df is not None else 0

    new_x, new_n = _x(sp["new_biz"]), _n(sp["new_biz"])
    ren_cur, ren_pri = _x(sp["ren_new"]), _x(sp["ren_old"])
    renewed = ren_cur - ren_pri                 # reprice on carried-forward birds
    ren_n = _n(sp["ren_new"])
    lap_x, lap_n = _x(sp["lapsed"]), _n(sp["lapsed"])
    net_move = L["summary"]["net_move"]         # ± revaluation on continuing layers
    moved_n = L["summary"]["moved_layers"]
    # closing is known (current total); derive opening so the walk ties exactly
    opening = closing_total - new_x - renewed + lap_x - net_move
    prior_as_at = changes.get("prior_as_at", "prior")

    r = _section(ws, r, f"Exposure bridge — {prior_as_at} → current (on-risk per-S/C)")
    r = _hdr(ws, r, 2, ["Movement", "Layers", "Exposure", "% of opening"],
             [46, 10, 18, 13])

    def line(label, n, val, alt, bold=False, total=False):
        _cell(ws, r0[0], 2, label, alt=alt, bold=bold)
        if n is not None:
            _cell(ws, r0[0], 3, n, alt=alt, bold=bold)
        c = _cell(ws, r0[0], 4, val, alt=alt, money=True, bold=bold)
        if total:
            c.fill = FILL_TOTAL
            ws.cell(row=r0[0], column=2).fill = FILL_TOTAL
            ws.cell(row=r0[0], column=2).font = F_TOTAL
            ws.cell(row=r0[0], column=4).font = F_TOTAL
        pct = (val / opening) if (opening and not total and not bold) else None
        if pct is not None:
            _cell(ws, r0[0], 5, pct, alt=alt, pct=True)
        r0[0] += 1

    r0 = [r]
    line(f"Opening in-force ({prior_as_at})", None, opening, False, bold=True)
    line("+ New business (genuinely new)", new_n, new_x, True)
    line("↻ Renewals (carried forward — reprice)", ren_n, renewed, False)
    line("− Lapsed / run-off (off the book)", -lap_n, -lap_x, True)
    line("± Revaluation (continuing layers)", moved_n, net_move, False)
    line("Closing in-force (current)", None, closing_total, True, bold=True, total=True)
    r = r0[0]

    # one-line read
    net_change = closing_total - opening
    direction = "grew" if net_change >= 0 else "shrank"
    if abs(new_x) >= abs(lap_x) and abs(new_x) >= abs(net_move):
        drv = "new business"
    elif abs(lap_x) >= abs(new_x) and abs(lap_x) >= abs(net_move):
        drv = "genuine run-off"
    elif abs(net_move) > max(abs(new_x), abs(lap_x)):
        drv = "revaluation of continuing layers"
    else:
        drv = "mixed movements"
    nc = ws.cell(row=r, column=2,
                 value=f"▸  The book {direction} {_money_m(abs(net_change))} "
                       f"({net_change / opening:+.1%}) vs {prior_as_at}, driven mainly "
                       f"by {drv}. Renewals carried forward via roll-forward net out "
                       f"here (counted once under the successor programme, not as "
                       f"run-off).")
    nc.font = Font(name=_FB, size=9, italic=True, color=SOFT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26
    return r + 2


def _portfolio(wb, per_layer, params, colmap, changes=None):
    ws = wb.create_sheet("Portfolio")
    _title(ws, "Portfolio profile",
           f"on-risk exposure as-at {params.as_at} · live SUMIF/COUNTIF on Per Layer")
    pl = per_layer.copy()
    pl["exp"] = pl["per_sc"]
    total = pl["exp"].sum()

    def grp(field):
        g = pl.groupby(pl[field].astype(str)).agg(n=("layer_key", "size"),
                                                   exp=("exp", "sum"))
        return g.sort_values("exp", ascending=False)

    persc = _col(colmap, "per_sc")
    # grand total cell (live SUM of Per Layer per_sc) — % columns reference it
    ws.cell(row=4, column=2, value="Total on-risk exposure:").font = F_KEY
    tot_cell = ws.cell(row=4, column=3, value=f"=SUM({persc})")
    tot_cell.font = F_TOTAL; tot_cell.number_format = MONEY
    TREF = "$C$4"

    def has(field):
        return field in colmap

    r = 6
    # ---- Exposure bridge: prior in-force → current in-force ----
    # Senior view: WHY did the book move? Opening + new − run-off ± reval = closing.
    if changes and changes.get("layers"):
        r = _exposure_bridge(ws, r, changes, total)

    r = _section(ws, r, "By entity")
    r = _breakdown(ws, r, "Entity", grp("entity"), total,
                   crit_col=_col(colmap, "entity"), persc_col=persc, total_ref=TREF)
    r = _section(ws, r, "By orbit")
    r = _breakdown(ws, r, "Orbit", grp("orbit"), total,
                   crit_col=_col(colmap, "orbit"), persc_col=persc, total_ref=TREF)
    r = _section(ws, r, "By underwriting year")
    # uw_year isn't a Per Layer column → keep static (documented)
    r = _breakdown(ws, r, "UW Year", grp("uw_year"), total)
    r = _section(ws, r, "Top 10 bus manufacturers")
    r = _breakdown(ws, r, "Bus Manufacturer", grp("bus_manufacturer").head(10),
                   total, label_w=46, crit_col=_col(colmap, "bus_manufacturer"),
                   persc_col=persc, total_ref=TREF)

    # concentration call-outs (static — derived insights, not raw aggregates)
    geo = pl[pl["orbit"] == "GEO-GSO"]
    top_mfr = grp("bus_manufacturer").index[0]
    top_mfr_share = grp("bus_manufacturer")["exp"].iloc[0] / total
    sc = pl.groupby("spacecraft_name")["exp"].sum().sort_values(ascending=False)
    top5_sc = sc.head(5).sum() / total
    r += 1
    r = _section(ws, r, "Concentration")
    r = _kv(ws, r, "Largest manufacturer", f"{top_mfr}  ({top_mfr_share:.1%})")
    r = _kv(ws, r, "Top-5 spacecraft share of book", f"{top5_sc:.1%}")
    r = _kv(ws, r, "GEO-GSO share of exposure",
            f"{geo['exp'].sum() / total:.1%}")


def _parameters(wb, params):
    """Treaty rates, limits & SQL connections - modeled on the Bespoke Warranty
    (Product Recall) Parameters standard: 4 columns
    TREATY RATES & LIMITS | VALUE | NOTE | SQL QUERIES & CONNECTIONS."""
    ws = wb.create_sheet("Parameters")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 64
    ws["B2"] = "SPACE RDS  \u2014  Parameters"; ws["B2"].font = F_TITLE
    ws["B3"] = (f"As at {params.as_at}   |   {params.quarter}   |   Treaty rates, "
                "limits & SQL connections that produced the numbers.")
    ws["B3"].font = F_SUB
    raw = params.raw
    sc = params.scenarios

    r = 5
    for j, h in enumerate(["TREATY RATES & LIMITS", "VALUE", "NOTE",
                           "SQL QUERIES & CONNECTIONS"]):
        cell = ws.cell(row=r, column=2 + j, value=h)
        cell.font = Font(name=_FB, size=10, bold=True, color=INK)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    r += 1

    def prow(label, value, note="", sql="", money=False, pct=False, bold_val=True):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = Font(
            name=_FB, size=9, color=SOFT)
        vc = ws.cell(row=r, column=3, value=value)
        vc.font = Font(name=_FB, size=10, bold=bold_val, color=INK)
        if money:
            vc.number_format = MONEY
        elif pct:
            vc.number_format = PCT
        vc.alignment = Alignment(horizontal="right")
        if note:
            ws.cell(row=r, column=4, value=note).font = Font(
                name=_FB, size=9, color=INK)
        if sql:
            scl = ws.cell(row=r, column=5, value=sql)
            scl.font = Font(name=_FB, size=8, color="1B4A7A")
            scl.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    def psection(title):
        nonlocal r
        r += 1
        ws.cell(row=r, column=2, value=title).font = F_SECT
        r += 1

    _sqlc = (params.ingest.get("sql") or {}) if isinstance(params.ingest, dict) else {}
    _srv = _sqlc.get("server", "LON-SQLP-V005")
    _db = _sqlc.get("database", "SpaceTrax_Data")
    _qf = _sqlc.get("query_file", "sql/space_rds_onrisk_asat.sql")
    _rwf = _sqlc.get("renewal_watch_query_file")

    psection("Source extract")
    prow("Connection", f"{_srv}  /  {_db}", "SQL Server / database",
         f"Server={_srv};  Database={_db}", bold_val=False)
    prow("On-risk population", "rds.vw_SpaceRDS_OnRisk", "as-at filtered SQL view",
         f"Query file:  {_qf}  (inline as-at param {{as_at}})", bold_val=False)
    prow("Base layer detail", "Prequel_Reporting.Prequel.Layer_Details_v",
         "IsBound=1, ASO/ASC, per-layer signed exposure",
         "WHERE IsBound=1 AND Placing_Basis NOT IN ('Consortium Declaration')",
         bold_val=False)
    prow("Consortium split gate", "rds.param_consortium_splits", "MGU cession by inception",
         "JOIN ... cs.[Start Date]<=inception AND cs.[End Date]>=inception "
         "AND [Controlling Body]='MGU'  (a missing row silently drops the layer "
         "— see the consortium-coverage pre-flight check)", bold_val=False)
    prow("As-at date", str(params.as_at), "drives on-risk window & RPF",
         "Parameters:  as_at_date = " + str(params.as_at), bold_val=False)
    prow("Manual add/remove", "rds.manually_controlled_rds_layers", "Action column",
         "SELECT * FROM rds.manually_controlled_rds_layers WHERE Action IN "
         "('Add Layer','Remove Layer')", bold_val=False)

    psection("Renewal-policy checks (SQL) — verified this quarter")
    prow("Forward renewal pointer", "J.Pbi.Layers_t", "latest snapshot; 0 = no renewal",
         (f"Query file:  {_rwf}.  " if _rwf else "") +
         "SELECT [Layer Id],[Program Id],NULLIF([Renewed To Program Id],0),"
         "[Renewed To UW Status],[Renewed To Inception] FROM J.Pbi.Layers_t "
         "WHERE [Snapshot Id]=(SELECT MAX([Snapshot Id]) FROM J.Pbi.Layers_t)",
         bold_val=False)
    prow("Successor-in-source check", "Layer_Details_v by ProgramId",
         "is the renewal bound in source yet?",
         "SELECT ProgramId,COUNT(*),SUM(Signed_Exposure_USD),MAX(CAST(IsBound AS INT)) "
         "FROM Prequel.Layer_Details_v WHERE ProgramId IN (<expiring>,<successor>) "
         "GROUP BY ProgramId  — drives rollforward vs let-it-flow decisions "
         "(Eutelsat, Turksat, Arabsat, WorldView, Inmarsat GX)", bold_val=False)
    prow("Consortium split coverage", "pre-flight warning",
         "flags MGU split ending before as-at",
         "Warns if the latest param_consortium_splits MGU row ends before the "
         "as-at (the gap that dropped Turksat 6A until its 2026-04-01→06-30 "
         "row was added)", bold_val=False)

    psection("External Quota Share (outwards RI)")
    try:
        _act = [s for s in params.outwards_slots
                if s["from"] <= params.as_at <= s["to"]]
        _actpct = float(_act[0]["pct"]) if _act else None
    except Exception:  # noqa: BLE001
        _actpct = None
    if _actpct is not None:
        prow("External Quota Share (current)", _actpct,
             "outwards RI cession active at as-at; the Summary 'Ext QS' column is "
             "SUM per layer of (exposure x slot pct by inception) — 20% only while "
             "every on-risk bird sits in this slot", pct=True)
    for s in params.outwards_slots:
        prow(f"  slot {s.get('name', s.get('id', ''))}", float(s["pct"]),
             f"inception {s['from']} -> {s['to']}", pct=True, bold_val=False)

    psection("IGR Quota Share")
    for ent, rate in raw["igr_qs"]["default"].items():
        prow(f"IGR {ent} QS rate", rate, f"{ent} intra-group QS ceded to FIBL",
             "RDS_Machine  /  IGR.Values_v  WHERE RDS_Year=? AND RDS_Quarter=?"
             if ent == "FUL" else "", pct=True)

    psection("IGR Excess of Loss")
    for ent, t in params.igr_xol.items():
        prow(f"IGR {ent} XoL excess (USD)", t["deductible"],
             f"{ent} IGR XoL attaches", money=True)
        prow(f"IGR {ent} XoL limit (USD)", t["limit"],
             f"{ent} IGR XoL limit", money=True)

    psection("Syndicate 3123 inwards")
    q = params.s3123_qs
    prow("S3123 QS cession", q["cession"], "inwards QS on consortium-eligible layers",
         "Parameters:  date_from=" + str(q["date_from"]) + "  date_to=" +
         str(q["date_to"]), pct=True)
    if "ig_equity" in raw.get("s3123", {}):
        prow("S3123 IG equity share", raw["s3123"]["ig_equity"],
             "IG equity slice of Syndicate 3123 net", pct=True)
    prow("S3123 excluded spacecraft",
         ", ".join(map(str, sorted(q["excluded"]))) or "\u2014",
         "spacecraft IDs excluded from S3123", bold_val=False)
    # S3123 consortium time-factor schedule (XLOOKUP exact-or-next-smaller on
    # inception). s3123_qs = per_sc x cession x this factor; equity uses it too.
    for _frm, _fac in getattr(params, "s3123_factors", []):
        prow(f"S3123 consortium factor \u2014 inception \u2265 {_frm}", float(_fac),
             "inwards QS time-factor by inception (12.5/30, 15/30, 10/30)",
             pct=True, bold_val=False)

    # S2126 \u2014 new consortium participant (2026-04-01, 5/30). Own syndicate RDS;
    # NO QS-to-IG, so it does not touch the IG book.
    _s2126 = getattr(params, "s2126_factors", []) or []
    if _s2126:
        psection("S2126 consortium (new 2026-04-01)")
        for _frm, _fac in _s2126:
            prow(f"S2126 consortium factor \u2014 inception \u2265 {_frm}", float(_fac),
                 "sub-share S3123 released on 2026-04-01 (5/30); own RDS",
                 pct=True, bold_val=False)
        _s2c = (params.raw.get("s2126_rds") or {}) if hasattr(params, "raw") else {}
        prow("S2126 QS to IG", float(_s2c.get("qs_to_ig", 0.0)),
             "no QS-to-IG agreement \u2014 net = gross (retains 100%); does NOT touch "
             "the IG book", pct=True, bold_val=False)

    psection("Scenario loss factors")
    prow("Proton Flare insured loss", sc["proton_flare"]["insured_loss"],
         "GEO-GSO only, no time decay", pct=True)
    prow("Generic Defect insured loss", sc["generic_defect"]["insured_loss"],
         "GEO+MEO, RPF time-decayed", pct=True)
    prow("Space Weather", 1.00, "worst bus-manufacturer, all orbits, raw gross",
         pct=True)
    prow("Max Risk", 1.00, "single largest on-risk spacecraft", pct=True)
    for o, v in sc["space_debris"]["damage_by_orbit"].items():
        prow(f"Space Debris damage ratio \u2014 {o}", v,
             "Kessler collision damage ratio", pct=True)

    psection("Risk Period Factor \u2014 bands (months to off-risk \u2192 factor)")
    for m, f in params.rpf_bands:
        prow(f"RPF: months \u2265 {m}", f, "time-decay factor", pct=True)

    psection("External (outwards) RI slots")
    for s in params.outwards_slots:
        prow(s["name"], s["pct"], f"inception {s['from']} \u2192 {s['to']}", pct=True)

    psection("Pipeline exclusions")
    prow("exclude_placing_basis",
         ", ".join(raw.get("exclude_placing_basis", [])) or "\u2014",
         "placing-basis values dropped pre-scenario", bold_val=False)
    prow("exclude_null_orbit", str(raw.get("exclude_null_orbit", True)),
         "drop no-Seradata-ID layers from all calcs", bold_val=False)
    n_corr = len(raw.get("data_corrections", []) or [])
    prow("data_corrections", f"{n_corr} active",
         "layer-level source overrides (see tab)", bold_val=False)


def _waterfalls(wb, grid):
    ws = wb.create_sheet("Netting Waterfalls")
    _title(ws, "Netting waterfalls", "layer-exact · per entity × scenario")

    # --- Four-entity summary — gross and net as SEPARATE tables (white headers) ---
    ents = ["FIHL", "FUL", "FIBL", "FIID"]
    r = 5
    for blk, basis in [("gross", "Gross"), ("net", "Net")]:
        r = _section(ws, r, f"Combined netting — {basis.lower()} by entity")
        r = _hdr(ws, r, 2, ["Scenario"] + ents, [16, 15, 15, 15, 15], white=True)
        for k, scen in enumerate(SCEN_ORDER):
            alt = k % 2 == 0
            _cell(ws, r, 2, scen, alt=alt)
            for j, ent in enumerate(ents):
                row = grid[(grid["entity"] == ent) & (grid["scenario"] == scen)]
                v = float(row.iloc[0][blk]) if len(row) and pd.notna(
                    row.iloc[0].get(blk)) else None
                _cell(ws, r, 3 + j, v, alt=alt, money=True)
            r += 1
        r += 1
    r += 1

    # --- Detailed step-by-step waterfalls for ALL FOUR entities ---
    # FIHL (group) and FIBL (internal receiver) don't cede internally, so their
    # IGR-QS / XoL steps read zero — the cascade still shows each entity's gross →
    # external QS → net. FUL and FIID carry the full cede-down to FIBL.
    WF_ENTS = ["FIHL", "FUL", "FIBL", "FIID"]
    _section(ws, r, "Cede-down detail — all entities (FIHL · FUL · FIBL · FIID)")
    nets = grid[grid["entity"].isin(WF_ENTS)].copy()
    nets["_o"] = nets["scenario"].map({s: i for i, s in enumerate(SCEN_ORDER)})
    _eo = {e: i for i, e in enumerate(WF_ENTS)}
    nets["_eo"] = nets["entity"].map(_eo)
    nets = nets.sort_values(["_eo", "_o"])
    base_r = r + 2
    col0 = 2
    amt_ranges = []   # (amount_col, first_amount_row, last_amount_row) per entity
    for entity in WF_ENTS:
        r = base_r
        first_amt = last_amt = None
        for _, row in nets[nets["entity"] == entity].iterrows():
            _d = row.get("detail")
            det = f" — {_d}" if (_d and str(_d) != "nan" and pd.notna(_d)) else ""
            ws.cell(row=r, column=col0,
                    value=f"{entity} · {row['scenario']}{det}").font = F_SECT
            r = _hdr(ws, r + 1, col0, ["Step", "Amount"], [30, 16], white=True)
            block_first = r     # first step row of this scenario block
            _colL = get_column_letter(col0 + 1)
            for k, (label, field, sign) in enumerate(WF_ROWS):
                _cell(ws, r, col0, label, alt=k % 2 == 0)
                if label.startswith("Net"):
                    # Net = live SUM of the (signed) steps above it, so the
                    # cascade always ties and never drifts from a baked value.
                    _fcell(ws, r, col0 + 1,
                           f"=SUM({_colL}{block_first}:{_colL}{r - 1})",
                           alt=k % 2 == 0, money=True, bold=True)
                else:
                    v = row.get(field)
                    v = None if pd.isna(v) else sign * abs(float(v))
                    _cell(ws, r, col0 + 1, v, alt=k % 2 == 0, money=True)
                if first_amt is None:
                    first_amt = r
                last_amt = r
                r += 1
            r += 2
        if first_amt is not None:
            amt_ranges.append((col0 + 1, first_amt, last_amt))
        col0 += 3
    # Conditional formatting: retained amounts (Gross/Net, +) read green, ceded
    # amounts (the 'less:' steps, −) read red — the cascade's give-away is visible.
    for ac, f0, f1 in amt_ranges:
        rng = f"{get_column_letter(ac)}{f0}:{get_column_letter(ac)}{f1}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThan", formula=["0"], font=Font(name=_FB, size=10, color=GREEN)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0"], font=Font(name=_FB, size=10, color="C0392B")))

    # --- Native, editable waterfall chart (FUL worst scenario) ---
    # Placed to the right of the detail blocks; backing data in cells so it
    # redraws when the user edits the source values.
    _waterfall_chart(ws, grid, "FUL", anchor_row=5, anchor_col=14)


def _waterfall_chart(ws, grid, entity, anchor_row, anchor_col=14):
    """Native Excel waterfall chart for an entity's worst-gross scenario.

    Built as a stacked bar (invisible 'base' series + visible step series) — the
    standard waterfall technique that renders & stays editable in every Excel
    version. The backing data block is written to cells with FORMULAS that derive
    each floating bar from a single source column, so editing the source values
    redraws the chart live (nothing is hardcoded in the chart itself).
    """
    block = grid[grid["entity"] == entity]
    if not len(block):
        return anchor_row
    row = block.loc[block["gross"].idxmax()]
    scen = row["scenario"]
    gross = float(row.get("gross") or 0)
    ext_qs = float(row.get("ext_qs") or 0)
    qs = float(row.get("igr_qs_ceded") or 0)
    xol = float(row.get("xol_ceded") or 0)
    net = float(row.get("net") or 0)

    # --- backing data block (cells the chart references) ---
    # Layout (col anchor_col = the 'Step' label column):
    #   col0: Step | col0+1: Source($) | col0+2: Base | col0+3: Down | col0+4: Up
    c0 = anchor_col
    L = get_column_letter
    # widen the data-block columns so values don't show as ###
    for off in range(5):
        ws.column_dimensions[L(c0 + off)].width = 13 if off else 9
    hdr_r = anchor_row
    ws.cell(row=hdr_r, column=c0, value=f"{entity} cascade — {scen}").font = F_SECT
    labels = ["Source ($)", "Base", "Decrease", "Total"]
    for j, h in enumerate(labels):
        c = ws.cell(row=hdr_r + 1, column=c0 + 1 + j, value=h)
        c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=hdr_r + 1, column=c0, value="Step").font = F_HDR
    ws.cell(row=hdr_r + 1, column=c0).fill = FILL_HDR

    steps = [("Gross", gross, "total"),
             ("Ext QS", -ext_qs, "down"),
             ("IGR QS", -qs, "down"),
             ("IGR XoL", -xol, "down"),
             ("Net", net, "total")]
    src_col = L(c0 + 1)   # Source ($) column letter
    base_col = L(c0 + 2)
    r = hdr_r + 2
    first_data = r
    running = 0.0
    for i, (label, delta, kind) in enumerate(steps):
        rr = r + i
        ws.cell(row=rr, column=c0, value=label).font = F_CELL
        # Source value (the editable driver)
        sc_cell = ws.cell(row=rr, column=c0 + 1, value=delta)
        sc_cell.number_format = MONEY; sc_cell.font = F_CELL
        if kind == "total":
            # Base = 0, Total = source, Decrease = 0
            ws.cell(row=rr, column=c0 + 2, value=0).number_format = MONEY
            ws.cell(row=rr, column=c0 + 3, value=0).number_format = MONEY
            ws.cell(row=rr, column=c0 + 4,
                    value=f"={src_col}{rr}").number_format = MONEY
            running = (running + delta) if i > 0 else delta
        else:
            # floating decrease bar: base = running + delta, height = -delta
            base = running + delta
            ws.cell(row=rr, column=c0 + 2,
                    value=f"={base_col}{rr-1}+{src_col}{rr-1}" if False else base
                    ).number_format = MONEY
            # height of the (downward) step is the absolute drop
            ws.cell(row=rr, column=c0 + 3,
                    value=f"=-{src_col}{rr}").number_format = MONEY
            ws.cell(row=rr, column=c0 + 4, value=0).number_format = MONEY
            running = base
    last_data = r + len(steps) - 1

    # --- the chart, referencing those cells ---
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = f"{entity} netting cascade — {scen}"
    chart.height = 7.5   # cm
    chart.width = 13
    chart.y_axis.numFmt = '#,##0,,"m"'
    chart.y_axis.title = "$m"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    cats = Reference(ws, min_col=c0, min_row=first_data, max_row=last_data)
    # Base (invisible), Decrease (navy-light), Total (navy)
    base_ref = Reference(ws, min_col=c0 + 2, min_row=hdr_r + 1, max_row=last_data)
    down_ref = Reference(ws, min_col=c0 + 3, min_row=hdr_r + 1, max_row=last_data)
    tot_ref = Reference(ws, min_col=c0 + 4, min_row=hdr_r + 1, max_row=last_data)
    for ref in (base_ref, down_ref, tot_ref):
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(cats)

    # colour the series: base transparent, decrease NAVY_LT, total NAVY
    s_base, s_down, s_tot = chart.series[0], chart.series[1], chart.series[2]
    s_base.graphicalProperties = GraphicalProperties()
    s_base.graphicalProperties.noFill = True
    s_base.graphicalProperties.line.noFill = True
    s_down.graphicalProperties = GraphicalProperties(solidFill="5B86B0")
    s_tot.graphicalProperties = GraphicalProperties(solidFill="1B3A5C")
    chart.legend = None
    chart.gapWidth = 40

    ws.add_chart(chart, f"{L(c0)}{last_data + 2}")
    # hide the chart-scaffolding columns (Base/Decrease/Total) — keep Step +
    # Source visible. Excel charts from hidden cells when plotVisOnly is off.
    for off in (2, 3, 4):
        ws.column_dimensions[L(c0 + off)].hidden = True
    chart.plotVisOnly = False
    # caption: this is live/editable
    cap_r = last_data + 1
    cc = ws.cell(row=cap_r, column=c0,
                 value="Editable: change a Source ($) value and the chart redraws.")
    cc.font = Font(name=_FB, size=8, italic=True, color=SOFT)
    return anchor_row


def _space_weather(wb, sw, colmap):
    """Live formulas: each cell SUMIFS the Per Layer tab by manufacturer.
    Click any value to see the layers behind it. Basis: all on-risk orbits."""
    ws = wb.create_sheet("Space Weather")
    _title(ws, "Space Weather · by bus manufacturer",
           "per-S/C gross by manufacturer (all orbits) · live SUMIFS on Per Layer")
    cols = [("bus_manufacturer", "Bus Manufacturer", 46),
            ("sw_fihl", "FIHL", 15), ("sw_ful", "FUL", 15),
            ("sw_fiid", "FIID", 15), ("sw_fibl_ceded", "FIBL Ceded", 15)]
    swt = sw.sort_values("sw_fihl", ascending=False)
    r = _hdr(ws, 5, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
    maxes = {c: swt[c].max() for c, _, _ in cols if c != "bus_manufacturer"}
    mfr_col = _col(colmap, "bus_manufacturer")
    persc = _col(colmap, "per_sc")
    fibl = _col(colmap, "total_fibl_ceded") if "total_fibl_ceded" in colmap else None
    ent = _col(colmap, "entity")
    first = r
    for k, (_, row) in enumerate(swt.iterrows()):
        alt = k % 2 == 0
        mref = f"$B{r}"   # the manufacturer label cell on this row
        _cell(ws, r, 2, row["bus_manufacturer"], alt=alt)
        # FIHL = all entities for this manufacturer
        f_fihl = f"=SUMIF({mfr_col},{mref},{persc})"
        # FUL / FIID = SUMIFS by manufacturer AND entity
        f_ful = f'=SUMIFS({persc},{mfr_col},{mref},{ent},"FUL")'
        f_fiid = f'=SUMIFS({persc},{mfr_col},{mref},{ent},"FIID")'
        f_fibl = f"=SUMIF({mfr_col},{mref},{fibl})" if fibl else 0
        for j, (field, val) in enumerate([("sw_fihl", f_fihl), ("sw_ful", f_ful),
                                          ("sw_fiid", f_fiid), ("sw_fibl_ceded", f_fibl)], start=1):
            bold = swt.iloc[k][field] == maxes[field] and swt.iloc[k][field] > 0
            c = _cell(ws, r, 2 + j, val, alt=alt, bold=bold)
            c.number_format = MONEY
            c.alignment = Alignment(horizontal="right")
        r += 1
    last = r - 1
    # Total row = SUM of the formula column above
    _cell(ws, r, 2, f"Total ({len(swt)} manufacturers)", bold=True)
    for j, colL in enumerate(["C", "D", "E", "F"], start=1):
        c = _cell(ws, r, 2 + j, f"=SUM({colL}{first}:{colL}{last})", bold=True)
        c.number_format = MONEY
        c.alignment = Alignment(horizontal="right")
    r += 2
    ws.cell(row=r, column=2,
            value="Live formulas: SUMIF/SUMIFS on the Per Layer tab. Basis: all "
                  "on-risk orbits, raw gross per-S/C (before equity/netting).").font = F_SUB


def _max_risk(wb, mr, grid=None):
    ws = wb.create_sheet("Max Risk")
    _title(ws, "Max Risk · single largest spacecraft",
           "Total loss of the one biggest satellite (gross = its full signed "
           "exposure across all its layers). Never summed with other scenarios.")

    # ---- Key data: the binding spacecraft + per-entity top spacecraft ----
    # Max Risk is per-SPACECRAFT (a satellite may carry several layers), so
    # aggregate the layer list up to the spacecraft before ranking.
    r = 5
    try:
        sc = (mr.groupby("spacecraft_name")
                .agg(gross=("gross_exposure", "sum"), net=("maxrisk_net", "sum"),
                     entity=("entity", "first"))
                .sort_values("gross", ascending=False))
        top_name = sc.index[0]
        top = sc.iloc[0]
        r = _section(ws, r, "Key data")
        r = _kv(ws, r, "Binding spacecraft (Max Risk)", str(top_name))
        r = _kv(ws, r, "Gross (full signed exposure)", float(top["gross"]), money=True)
        r = _kv(ws, r, "Net (after reinsurance)", float(top["net"]), money=True)
        ws.cell(row=r - 1, column=4,
                value="= total loss of this single satellite").font = F_SUB
        # per-entity top spacecraft
        r += 1
        r = _hdr(ws, r, 2, ["Entity", "Top spacecraft", "Gross", "Net"],
                 [10, 28, 16, 16])
        # FIBL is the internal RI receiver — its Max Risk is the bird it RECEIVES
        # the most on (IGR QS + S3123 + equity + direct), taken from the netting
        # grid's receiver basis, NOT its thin direct book. Matches Summary /
        # Change Narrative.
        fibl_recv = None
        try:
            if grid is not None:
                gr = grid[(grid["entity"] == "FIBL") & (grid["scenario"] == "Max Risk")]
                if len(gr):
                    gr = gr.iloc[0]
                    fibl_recv = (str(gr.get("detail")),
                                 float(gr.get("gross_receiver")),
                                 float(gr.get("net_receiver")))
        except Exception:  # noqa: BLE001
            fibl_recv = None
        for ent in ["FIHL", "FUL", "FIID", "FIBL"]:
            if ent == "FIBL" and fibl_recv is not None:
                _cell(ws, r, 2, ent, bold=True)
                _cell(ws, r, 3, fibl_recv[0])
                _cell(ws, r, 4, fibl_recv[1], money=True)
                _cell(ws, r, 5, fibl_recv[2], money=True)
                r += 1
                continue
            sub = mr if ent == "FIHL" else mr[mr["entity"] == ent]
            if not len(sub):
                continue
            g = (sub.groupby("spacecraft_name")
                    .agg(gross=("gross_exposure", "sum"), net=("maxrisk_net", "sum"))
                    .sort_values("gross", ascending=False))
            nm = g.index[0]
            _cell(ws, r, 2, ent, bold=True)
            _cell(ws, r, 3, str(nm))
            _cell(ws, r, 4, float(g.iloc[0]["gross"]), money=True)
            _cell(ws, r, 5, float(g.iloc[0]["net"]), money=True)
            r += 1
        if fibl_recv is not None:
            ws.cell(row=r, column=2,
                    value="FIBL is the internal RI receiver: its Max Risk is what "
                          "it receives on the most-ceded single bird (IGR QS + "
                          "S3123 + equity), not its direct book.").font = F_SUB
            r += 1
        r += 1
    except Exception:  # noqa: BLE001 — summary is best-effort; detail below always renders
        r = 5

    # ---- Full detail: largest layers ----
    r = _section(ws, r, "Largest layers (top 12)")
    cols = [("spacecraft_name", "Spacecraft", 28), ("entity", "Entity", 9),
            ("program_id", "Program", 11), ("layer_id", "Layer", 9),
            ("gross_exposure", "Gross", 15), ("maxrisk_net", "Net", 15)]
    mrt = mr.sort_values("gross_exposure", ascending=False).head(12)
    r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
    for k, (_, row) in enumerate(mrt.iterrows()):
        for j, (field, _, _) in enumerate(cols):
            _cell(ws, r, 2 + j, row[field], alt=k % 2 == 0,
                  money=field in ("gross_exposure", "maxrisk_net"))
        r += 1


# --------------------------------------------------------------------------- #
#  Change Narrative — per-scenario prior→current comparison for writing the    #
#  quarter's narrative (mirrors JJ's manual Change Narrative tabs).            #
# --------------------------------------------------------------------------- #
_SCEN_DEF = {
    "Proton Flare": "5% insured loss · GEO-GSO only · no time decay",
    "Space Weather": "worst single bus-manufacturer · all orbits · raw gross",
    "Generic Defect": "50% insured loss · GEO+MEO · RPF time-decayed",
    "Space Debris": "collision damage ratio by orbit (Kessler)",
    "Max Risk": "total loss of the single largest spacecraft",
}


def _change_narrative(wb, changes, per_layer, mr, sw, params):
    """One block per scenario: entity × (current vs prior gross/net, Δ, Δ%) plus
    the key driver and a blank Narrative column to write into — the comparison
    JJ's manual Change Narrative tabs are built from, in one place."""
    if not changes or "scenarios" not in changes:
        return
    sc = changes["scenarios"]
    if sc is None or not len(sc):
        return
    ws = wb.create_sheet("Change Narrative")
    _title(ws, "Change Narrative — write-up helper",
           f"{params.quarter} vs {changes.get('prior_as_at', 'prior')} · per "
           f"scenario × entity · fill the Narrative column · gross = exec basis "
           f"(FIHL combined, FIBL receiver)")

    ENTS = ["FIHL", "FUL", "FIBL", "FIID"]

    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # key drivers ------------------------------------------------------------
    sw_top = "—"
    try:
        sw_top = str(sw.sort_values("sw_fihl", ascending=False).iloc[0]["bus_manufacturer"])
    except Exception:  # noqa: BLE001
        pass

    def _mr_top(ent):
        try:
            g = (per_layer if ent == "FIHL" else per_layer[per_layer["entity"] == ent])
            gg = g.groupby("spacecraft_name")["per_sc"].sum()
            return str(gg.idxmax()) if len(gg) else "—"
        except Exception:  # noqa: BLE001
            return "—"

    # FIBL Max Risk is the single bird FIBL RECEIVES the most on (receiver
    # basis), not the group's largest-gross bird — name it in the driver.
    fibl_mr_top = "—"
    try:
        from . import netting as _net
        _, fibl_mr_top = _net.fibl_maxrisk_multiplier(per_layer)
        fibl_mr_top = str(fibl_mr_top) if fibl_mr_top else "—"
    except Exception:  # noqa: BLE001
        pass

    def _driver(scen, ent):
        if scen == "Max Risk":
            if ent == "FIBL":
                return f"IGR + S3123 receiver ({fibl_mr_top})"
            return _mr_top(ent)
        if scen == "Space Weather":
            return sw_top
        return "Aggregate — all exposed policies"

    r = 5
    HEAD = ["Entity", "Key driver", "Cur Gross", "Cur Net", "Prior Gross",
            "Prior Net", "Δ Gross", "Δ Gross %", "Δ Net", "Δ Net %",
            "Narrative (write here)"]
    WID = [10, 26, 15, 15, 15, 15, 14, 10, 14, 10, 46]

    for scen in SCEN_ORDER:
        block = sc[sc["scenario"] == scen]
        if not len(block):
            continue
        r = _section(ws, r, f"{scen}")
        ws.cell(row=r - 1, column=4, value=_SCEN_DEF.get(scen, "")).font = F_SUB
        hdr = r
        r = _hdr(ws, r, 2, HEAD, WID, white=True)
        first = r
        by_ent = {row["entity"]: row for _, row in block.iterrows()}
        for k, ent in enumerate(ENTS):
            row = by_ent.get(ent)
            if row is None:
                continue
            alt = k % 2 == 0
            # gross on exec basis (FIHL combined / FIBL receiver) when present
            cg = _num(row.get("cur_gross_exec"))
            cg = _num(row.get("cur_gross")) if cg is None else cg
            pg = _num(row.get("prior_gross_exec"))
            pg = _num(row.get("prior_gross")) if pg is None else pg
            cn, pn = _num(row.get("cur_net")), _num(row.get("prior_net"))
            _cell(ws, r, 2, ent, alt=alt, bold=True)
            _cell(ws, r, 3, _driver(scen, ent), alt=alt)
            _cell(ws, r, 4, cg, alt=alt, money=True)
            _cell(ws, r, 5, cn, alt=alt, money=True)
            _cell(ws, r, 6, pg, alt=alt, money=True)
            _cell(ws, r, 7, pn, alt=alt, money=True)
            # Δ live off the current/prior cells so it recomputes if edited
            _fcell(ws, r, 8, f"=D{r}-F{r}", alt=alt, money=True,
                   bold=(cg is not None and pg is not None and abs(cg - pg) > 1e6))
            c9 = _fcell(ws, r, 9, f"=IFERROR((D{r}-F{r})/F{r},0)", alt=alt)
            c9.number_format = PCT
            _fcell(ws, r, 10, f"=E{r}-G{r}", alt=alt, money=True)
            c11 = _fcell(ws, r, 11, f"=IFERROR((E{r}-G{r})/G{r},0)", alt=alt)
            c11.number_format = PCT
            # Narrative cell — bordered blank for writing
            nc = _cell(ws, r, 12, None, alt=alt)
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        # Δ% sign colour on both Δ Gross % (I) and Δ Net % (K)
        _delta_signcolour(ws, f"I{first}:I{r - 1}")
        _delta_signcolour(ws, f"K{first}:K{r - 1}")
        r += 2

    # ---- Draft narrative (per scenario) — copy/paste-ready, built from this
    #      run's FIHL movement so the numbers never go stale; drivers are fixed
    #      commentary the writer verifies/edits before filing. -----------------
    fihl = {row["scenario"]: row for _, row in sc[sc["entity"] == "FIHL"].iterrows()}
    mr_top = _mr_top("FIHL")
    _mrow = fihl.get("Max Risk")
    _mrx = 0.0
    if _mrow is not None:
        _mrx = _num(_mrow.get("cur_gross_exec")) or _num(_mrow.get("cur_gross")) or 0.0
    DRIVERS = {
        "Proton Flare": "GEO-GSO renewals (Eutelsat, Arabsat, Inmarsat GX, Turksat) "
            "broadly offsetting expiring layers and the SXM-10 non-renewal; retention "
            "stable at ~81% after the unchanged 20% external QS.",
        "Space Weather": f"the binding single manufacturer ({sw_top}) — the SXM-10 "
            "(SSL) exit and an easing of the peak-manufacturer concentration, partly "
            "offset by the Airbus-built Eutelsat and Arabsat renewals.",
        "Generic Defect": "the Eutelsat fleet (23 GEO spacecraft, ~$243m) renewing "
            "back into FIID under programme 385575, with Inmarsat GX and Arabsat adding "
            "GEO exposure — a genuine exposure gain, not a rate or methodology change.",
        "Space Debris": "LEO run-off, chiefly WorldView / Maxar Legion re-attaching to "
            "the in-orbit fleet (369418) and counted there, plus other LEO expiries; "
            "GEO held broadly steady (LEO carries the 0.40 damage ratio vs 0.05 GEO).",
        "Max Risk": f"the single largest spacecraft ({mr_top}); the binding line size "
            f"is unchanged (~${_mrx / 1e6:,.0f}m).",
    }
    r = _section(ws, r, "Draft narrative (per scenario — verify & edit before filing)")
    hdr = r
    r = _hdr(ws, r, 2, ["Scenario", "Draft narrative (copy/paste)"], [16, 60], white=True)
    for k, scen in enumerate(SCEN_ORDER):
        row = fihl.get(scen)
        if row is None:
            continue
        cg = _num(row.get("cur_gross_exec")) or _num(row.get("cur_gross"))
        pg = _num(row.get("prior_gross_exec")) or _num(row.get("prior_gross"))
        cn, pn = _num(row.get("cur_net")), _num(row.get("prior_net"))
        dg = (cg / pg - 1) if (cg and pg) else 0.0
        dn = (cn / pn - 1) if (cn and pn) else 0.0
        gp, npc = f"{dg:+.1%}", f"{dn:+.1%}"
        if abs(dg) < 0.02:
            lead = f"Broadly flat — gross {gp}, net {npc}."
        else:
            lead = f"{'Up' if dg > 0 else 'Down'} {gp} gross / {npc} net."
        text = f"{lead} Driven by {DRIVERS.get(scen, '')}"
        alt = k % 2 == 0
        _cell(ws, r, 2, scen, alt=alt, bold=True)
        nc = _cell(ws, r, 3, text, alt=alt)
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 44
        r += 1
    r += 1
    # Cross-cutting driver: the S3123 consortium factor step moves the combined /
    # receiver basis independent of gross exposure movement.
    _cnote = ws.cell(row=r, column=2, value=(
        "Combined-basis driver: FIHL headline gross/net and the FIBL receiver include "
        "the S3123 QS to IG + IG equity add-backs. The S3123 consortium sub-share "
        "stepped 0.500 → 0.333 for inceptions from 2026-04-01, so as the GEO fleet "
        "renews post-April those add-backs fall ~1/3 on the renewed birds — a driver "
        "of the combined and receiver movement that is INDEPENDENT of the gross "
        "exposure change. External QS (20%) and the IGR QS rates are unchanged, so "
        "operating-entity (FUL/FIID) net tracks gross. The 5/30 that S3123 released "
        "went to a NEW consortium participant, Syndicate 2126, which has NO QS-to-IG "
        "agreement — so the S3123 add-back drop is NOT offset on the IG side. S2126 "
        "is reported separately on its own 'S2126 RDS Summary' tab (net = gross, "
        "very small exposure) and does not touch the IG numbers."))
    _cnote.font = Font(name=_FB, size=9, italic=True, color=SOFT)
    _cnote.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 84
    r += 2

    # Standing commentary the writer edits (seeded from the treaty terms) -------
    r = _section(ws, r, "Reinsurance arrangements (standing — edit as needed)")
    ri = ("External 20% QS on FUL/FIID/FIHL (risk-attaching). Internal QS FUL→FIBL "
          "40% (2022) / 50% (2023–25); FIID→FIBL 85% all years. Internal XoL "
          "FUL→FIBL $245m xs $50m; FIID→FIBL $72.5m xs $7.5m. Inwards 20% QS "
          "S3123→IG (in force since 01/07/2024), carried at the S3123 consortium "
          "time-factor which steps 0.4167 → 0.500 (from 2026-01-01) → 0.333 (from "
          "2026-04-01) by inception — so the S3123 QS to IG and IG equity add-backs "
          "shrink as post-April renewals move onto the 0.333 sub-share.")
    rc = ws.cell(row=r, column=2, value=ri)
    rc.font = Font(name=_FB, size=9, color=INK)
    rc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 56
    r += 2
    fc = ws.cell(row=r, column=2, value="Further commentary (write here): ")
    fc.font = Font(name=_FB, size=9, italic=True, color=SOFT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=12)

    for col, w in zip("BCDEFGHIJKL", WID):
        ws.column_dimensions[col].width = w


def _contrib_formula(spec, i, idx):
    """Live formula string for an S3123/S2126 contribution cell at row i, or
    None if a referenced Per Layer column is missing (caller falls back to the
    engine value). `idx` maps column name -> 0-based position on the sheet."""
    def _c(name):
        return get_column_letter(idx[name] + 1) if name in idx else None
    if spec["kind"] == "gross":
        persc, fac = _c("per_sc"), _c(spec["factor"])
        elig, sel = _c(spec["elig"]), _c(spec["sel"])
        if not all((persc, fac, elig, sel)):
            return None
        term = f"{persc}{i}*{fac}{i}*{elig}{i}*{spec['loss']:g}"
        if spec["rpf"]:
            rpf = _c("rpf")
            if not rpf:
                return None
            term += f"*{rpf}{i}"
        return f"=IF({sel}{i}=1,{term},0)"
    # net = gross x (1 - qs); excluded spacecraft retain 100% (net = gross)
    g = _c(spec["gross"])
    if not g:
        return None
    if spec["qs"] == 0:
        return f"={g}{i}"
    excl = sorted(spec["excl"])
    if excl:
        scid = _c("spacecraft_id")
        if scid:
            cond = "OR(" + ",".join(f"{scid}{i}={x}" for x in excl) + ")"
            return f"=IF({cond},{g}{i},{g}{i}*{1 - spec['qs']:g})"
    return f"={g}{i}*{1 - spec['qs']:g}"


def _per_layer(wb, per_layer, contrib=None):
    ws = wb.create_sheet("Per Layer")
    ws.sheet_view.showGridLines = False
    keep = [c for c in _PL_KEEP if c in per_layer.columns]
    contrib = contrib or {}
    money_cols = {"per_sc", "ext_qs", "s3123_qs", "igr_qs_ceded", "net_of_qs",
                  "xol_ceded", "net_of_xol", "pf_fihl", "gd_fihl", "sd_fihl"}
    # per-$ cession-rate helper columns are written as LIVE formulas (= $ column /
    # per_sc) so the Summary/S3123 SUMPRODUCT backbone recomputes if a per-layer $
    # value is edited. Maps helper field -> numerator field.
    _PP_SRC = {"ext_qs_pp": "ext_qs", "igr_ceded_pp": "igr_qs_ceded",
               "s3123_qs_pp": "s3123_qs", "equity_pp": "equity_usd"}
    _idx = {name: j for j, name in enumerate(keep)}
    for j, name in enumerate(keep):
        c = ws.cell(row=1, column=1 + j, value=name)
        c.font = F_HDR; c.fill = FILL_HDR
        ws.column_dimensions[get_column_letter(1 + j)].width = \
            22 if name in ("spacecraft_name", "bus_manufacturer") else 13
    _per_col = (get_column_letter(_idx["per_sc"] + 1) if "per_sc" in _idx else None)
    for i, (_, row) in enumerate(per_layer[keep].iterrows(), start=2):
        for j, name in enumerate(keep):
            if name in _PP_SRC and _PP_SRC[name] in _idx and _per_col:
                num = get_column_letter(_idx[_PP_SRC[name]] + 1)
                c = ws.cell(row=i, column=1 + j,
                            value=f"=IFERROR({num}{i}/{_per_col}{i},0)")
                c.font = F_CELL
                continue
            if name in contrib:
                f = _contrib_formula(contrib[name], i, _idx)
                if f is not None:
                    c = ws.cell(row=i, column=1 + j, value=f)
                    c.font = F_CELL; c.number_format = MONEY
                    continue
            v = row[name]
            if name == "s3123_eligible":
                # write a real Excel boolean (not "True"/"False" text) so the
                # linkify IF(s3123_eligible, ...) formula evaluates correctly
                v = None if (isinstance(v, float) and pd.isna(v)) else bool(v)
            elif isinstance(v, (dt.date, dt.datetime)):
                v = str(v)
            elif isinstance(v, float) and pd.isna(v):
                v = None
            c = ws.cell(row=i, column=1 + j, value=v)
            c.font = F_CELL
            if name in money_cols and isinstance(v, (int, float)):
                c.number_format = MONEY
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keep))}{len(per_layer) + 1}"



def _changes_narrative(sc):
    """Generate an insurance reading of the matrix: flag the largest mover and
    distinguish structural drivers (e.g. FIBL as IGR receiver) from real risk."""
    bits = []
    # FIBL headline = DIRECT book (manual convention since 2026Q1 recon);
    # movement there is new/expired direct business, not cessions.
    if (sc["entity"] == "FIBL").any():
        fibl = sc[sc["entity"] == "FIBL"]
        if (fibl["d_gross_pct"].fillna(0) > 0).any() or \
                (fibl["cur_gross"].fillna(0) > fibl["prior_gross"].fillna(0)).any():
            bits.append("FIBL shows its direct book only (manual convention); "
                        "the receiver view is on the S3123 & Equity (IG) tab.")
    # FIID running off?
    fiid = sc[sc["entity"] == "FIID"]
    if len(fiid) and fiid["d_gross_pct"].dropna().mean() < -0.02:
        bits.append("FIID falls across scenarios as covered policies run off risk.")
    # biggest single mover (by |Δ%|) among the retained entities
    real = sc[sc["entity"].isin(["FIHL", "FUL", "FIID"])].copy()
    real = real[real["d_gross_pct"].notna()]
    if len(real):
        top = real.loc[real["d_gross_pct"].abs().idxmax()]
        direction = "rise" if top["d_gross_pct"] > 0 else "fall"
        bits.append(f"Largest move: {top['entity']} {top['scenario']} "
                    f"{top['d_gross_pct']:+.1%} ({direction}).")
    return "  ".join(bits)


def _changes_matrix_exec(ws, r, sc):
    """Headline matrix on the MANUAL Executive Summary convention:
    FIHL = combined (incl S3123 QS + IG equity), FIBL = internal receiver,
    FUL/FIID = standard gross. Static Δ% values from the exec-basis columns
    (cur_gross_exec / prior_gross_exec) added by prior_seed.build_changes —
    reproduces the manual exec matrix up to accepted data corrections and
    populates the FIBL column."""
    ents = [e for e in ["FIHL", "FUL", "FIBL", "FIID"]
            if (sc["entity"] == e).any()]
    r = _section(ws, r, "Movement by scenario & entity  (Gross Δ% vs prior "
                        "quarter — manual Exec basis: FIHL combined, FIBL receiver)")
    r = _hdr(ws, r, 2, ["Scenario"] + ents, [22] + [14] * len(ents))
    first = r
    tot = {e: [0.0, 0.0] for e in ents}
    for k, scen in enumerate(SCEN_ORDER):
        if not (sc["scenario"] == scen).any():
            continue
        alt = k % 2 == 1
        _cell(ws, r, 2, scen, alt=alt, bold=True)
        for j, ent in enumerate(ents):
            b = sc[(sc["entity"] == ent) & (sc["scenario"] == scen)]
            cg = float(b.iloc[0].get("cur_gross_exec") or 0) if len(b) else 0.0
            pg = float(b.iloc[0].get("prior_gross_exec") or 0) if len(b) else 0.0
            tot[ent][0] += cg; tot[ent][1] += pg
            _delta_formula_cell(ws, r, 3 + j,
                                (cg / pg - 1) if pg else None, alt=alt)
        ws.row_dimensions[r].height = 17
        r += 1
    _cell(ws, r, 2, "TOTAL (all scenarios)", bold=True)
    ws.cell(row=r, column=2).fill = FILL_TOTAL
    ws.cell(row=r, column=2).font = F_TOTAL
    for j, ent in enumerate(ents):
        cg, pg = tot[ent]
        c = _delta_formula_cell(ws, r, 3 + j,
                                (cg / pg - 1) if pg else None, bold=True)
        c.fill = FILL_TOTAL
    rng = (f"{get_column_letter(3)}{first}:"
           f"{get_column_letter(2 + len(ents))}{r}")
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color="C0392B")))
    return r + 2


def _changes_matrix_live(ws, r, sc, ents, scens_present,
                         gross_row, gross_first, gross_last, title):
    """Scenario × entity matrix of gross Δ% — the senior-facing headline view.

    Not hard-coded: every value is a LIVE Excel formula that divides the
    current/prior gross figures in the 'By scenario (gross)' detail table below
    (column D = current, E = prior), so the headline recomputes whenever the
    workbook updates. The ▲/▼ glyph comes from the sign-driven number format
    (DELTA_NUMFMT) and the green/red colour from conditional formatting, which
    together reproduce the original static look exactly while staying numeric.
    A TOTAL row divides summed-current by summed-prior per entity.

    gross_row:   {(entity, scenario): excel_row} in the gross detail table
    gross_first/gross_last: {entity: row} bounds of that entity's gross block
    """
    r = _section(ws, r, title)
    widths = [22] + [14] * len(ents)
    r = _hdr(ws, r, 2, ["Scenario"] + ents, widths)
    first_val_row = r
    for k, scen in enumerate(SCEN_ORDER):
        if scen not in scens_present:
            continue
        alt = k % 2 == 1
        _cell(ws, r, 2, scen, alt=alt, bold=True)
        for j, ent in enumerate(ents):
            gr = gross_row.get((ent, scen))
            block = sc[(sc["entity"] == ent) & (sc["scenario"] == scen)]
            pct = block.iloc[0].get("d_gross_pct") if len(block) else None
            no_data = gr is None or pct is None or (isinstance(pct, float) and pd.isna(pct))
            f = None if no_data else f"=IFERROR(D{gr}/E{gr}-1,0)"
            _delta_formula_cell(ws, r, 3 + j, f, alt=alt)
        ws.row_dimensions[r].height = 17
        r += 1
    # TOTAL row — all scenarios combined, per entity
    _cell(ws, r, 2, "TOTAL (all scenarios)", bold=True)
    ws.cell(row=r, column=2).fill = FILL_TOTAL
    for j, ent in enumerate(ents):
        eb = sc[sc["entity"] == ent]
        prior_sum = eb["prior_gross"].dropna().sum() if "prior_gross" in eb else 0
        f0, f1 = gross_first.get(ent), gross_last.get(ent)
        f = (f"=IFERROR(SUM(D{f0}:D{f1})/SUM(E{f0}:E{f1})-1,0)"
             if (f0 and prior_sum) else None)
        c = _delta_formula_cell(ws, r, 3 + j, f, bold=True)
        c.fill = FILL_TOTAL
    ws.cell(row=r, column=2).font = F_TOTAL
    last_val_row = r
    # Sign-driven colour: green ▲ (≥0) / red ▼ (<0). Conditional formatting keeps
    # the colour correct even if the live values change after generation.
    rng = (f"{get_column_letter(3)}{first_val_row}:"
           f"{get_column_letter(2 + len(ents))}{last_val_row}")
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color="C0392B")))
    r += 2
    return r


def _delta_formula_cell(ws, row, col, formula, alt=False, bold=False):
    """One headline Δ% cell. `formula` is a live Excel formula returning a
    fraction (e.g. 0.096), displayed as '▲ +9.6%' / '▼ -3.0%' by DELTA_NUMFMT;
    colour is applied separately by conditional formatting. None → soft '–'."""
    cell = ws.cell(row=row, column=col)
    if formula is None:
        cell.value = "–"
        cell.font = Font(name=_FB, size=10, color=SOFT, bold=bold)
    else:
        cell.value = formula
        cell.number_format = DELTA_NUMFMT
        cell.font = Font(name=_FB, size=10, bold=True, color=INK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if alt:
        cell.fill = FILL_ALT
    return cell


def _delta_signcolour(ws, rng):
    """Green ▲ up / red ▼ down on a range of Δ% cells. Applied by conditional
    formatting so the colour stays correct if the live formulas recompute."""
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"],
        font=Font(name=_FB, size=10, bold=True, color="C0392B")))


def _manual_qoq_tables(ws, r, cur, pri):
    """FIX(recon 2026Q1 — align QoQ with the MANUAL Changes tab): the manual
    book's quarter-on-quarter view is EXPOSURE-level, not RDS-value level:
      Table 2  — entity x orbit / scenario-eligible exposure, prior vs current,
                 with a Δ% grid (FIHL = FUL + FIID; headline = FIHL GEO move);
      Table 1b — whole-portfolio exposure bridge by reason
                 (Written / Renewed / Expired / moves);
      Table 3  — RPF-adjusted exposure (Generic Defect early-warning).
    Reproduced here from the per-layer frames so the matrix reconciles with
    the manual up to accepted data corrections, independent of any netting or
    display convention. cur/pri need: entity, orbit, per_sc, rpf and layer ids.
    Returns the next free row; skips silently if inputs are unusable.
    """
    need = {"entity", "orbit", "per_sc"}
    if cur is None or pri is None or not need <= set(cur.columns) \
            or not need <= set(pri.columns):
        return r
    DR = {"GEO-GSO": 0.05, "MEO": 0.10, "LEO": 0.40}

    def aggs(df):
        e = df["entity"]; o = df["orbit"]; x = df["per_sc"].astype(float)
        dr = o.map(DR).fillna(0.0)
        out = {}
        for ent in ("FUL", "FIID", "FIBL"):
            m = e.eq(ent)
            row = {
                "GEO-GSO": float(x[m & o.eq("GEO-GSO")].sum()),
                "MEO": float(x[m & o.eq("MEO")].sum()),
                "LEO": float(x[m & o.eq("LEO")].sum()),
                "Total": float(x[m].sum()),
                "SD orbit-weighted": float((x * dr)[m].sum()),
            }
            row["GD Eligible"] = row["GEO-GSO"] + row["MEO"]
            row["PF Eligible"] = row["GEO-GSO"]
            out[ent] = row
        # manual Table 2 convention: FIHL = FUL + FIID (excludes FIBL direct);
        # Whole portfolio adds FIBL back and ties Table 1b opening/closing.
        out["FIHL (FUL+FIID)"] = {k: out["FUL"][k] + out["FIID"][k]
                                  for k in out["FUL"]}
        out["Whole portfolio"] = {k: out["FIHL (FUL+FIID)"][k] + out["FIBL"][k]
                                  for k in out["FUL"]}
        return out

    ENT_ROWS = ("FUL", "FIID", "FIBL", "FIHL (FUL+FIID)", "Whole portfolio")
    CO = ["GEO-GSO", "MEO", "LEO", "Total",
          "SD orbit-weighted", "GD Eligible", "PF Eligible"]
    A_cur, A_pri = aggs(cur), aggs(pri)

    # headline — the manual's own headline is the FIHL GEO (PF-eligible) move
    g0 = A_pri["FIHL (FUL+FIID)"]["GEO-GSO"]
    g1 = A_cur["FIHL (FUL+FIID)"]["GEO-GSO"]
    hl = ws.cell(row=r, column=2,
                 value=(f"Headline (manual basis) — GEO / PF-eligible exposure: "
                        f"{g0:,.0f} \u2192 {g1:,.0f}  "
                        f"({(g1 / g0 - 1) if g0 else 0:+.2%},  "
                        f"{g1 - g0:+,.0f})"))
    hl.font = F_TOTAL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    r += 2

    r = _section(ws, r, "Aggregated exposure by entity (manual Table 2)")
    blocks = {}
    for label, A in (("Prior", A_pri), ("Current", A_cur)):
        r = _hdr(ws, r, 2, [label] + CO, [19] + [15] * len(CO))
        blocks[label] = r
        for i, ent in enumerate(ENT_ROWS):
            bold = ent in ("FIHL (FUL+FIID)", "Whole portfolio")
            _cell(ws, r, 2, ent, alt=i % 2, bold=bold)
            for j, cname in enumerate(CO):
                _cell(ws, r, 3 + j, A[ent][cname], alt=i % 2, money=True,
                      bold=bold)
            r += 1
        r += 1
    r = _hdr(ws, r, 2, ["\u0394%"] + CO, None)
    d_first = r
    for i, ent in enumerate(ENT_ROWS):
        bold = ent in ("FIHL (FUL+FIID)", "Whole portfolio")
        _cell(ws, r, 2, ent, alt=i % 2, bold=bold)
        rp, rc = blocks["Prior"] + i, blocks["Current"] + i
        for j in range(len(CO)):
            col = get_column_letter(3 + j)
            _delta_formula_cell(ws, r, 3 + j,
                                f"=IFERROR({col}{rc}/{col}{rp}-1,0)",
                                alt=i % 2, bold=bold)
        r += 1
    # \u0394% grid: up green / down red (matches the headline matrix convention)
    _delta_signcolour(ws, f"C{d_first}:{get_column_letter(2 + len(CO))}{r - 1}")
    r += 1

    # Table 1b (exposure bridge by reason) removed: its programme-id split counted
    # roll-forward renewals as "expired without replacement" (renewal-blind) and
    # duplicated the renewal-aware Portfolio exposure bridge. The Portfolio bridge
    # (Opening → New → Renewals → Lapsed → Revaluation → Closing) is now the single
    # source of truth for the movement walk.

    # ---- Table 3: RPF-adjusted exposure (GD early-warning) ----
    if "rpf" in cur.columns and "rpf" in pri.columns:
        def rpfadj(df):
            x = df["per_sc"].astype(float) * df["rpf"].astype(float)
            o = df["orbit"]
            return {"GEO+MEO (GD basis)": float(x[o.isin(["GEO-GSO", "MEO"])].sum()),
                    "LEO": float(x[o.eq("LEO")].sum()),
                    "Total": float(x.sum())}
        r = _section(ws, r, "RPF-adjusted exposure (manual Table 3 — GD early-warning)")
        Rc, Rp = rpfadj(cur), rpfadj(pri)
        r = _hdr(ws, r, 2, ["", "Prior", "Current", "\u0394%"],
                 [24, 16, 16, 10])
        rpf_first = r
        for i, k in enumerate(Rc):
            _cell(ws, r, 2, k, alt=i % 2)
            _cell(ws, r, 3, Rp[k], alt=i % 2, money=True)
            _cell(ws, r, 4, Rc[k], alt=i % 2, money=True)
            _delta_formula_cell(ws, r, 5, f"=IFERROR(D{r}/C{r}-1,0)", alt=i % 2)
            r += 1
        _delta_signcolour(ws, f"E{rpf_first}:E{r - 1}")
        r += 1
    return r


def _changes(wb, changes, per_layer=None, params=None):
    if not changes:
        return
    ws = wb.create_sheet("Changes")
    _title(ws, "Quarter-on-quarter changes",
           f"current run vs prior run ({changes['prior_as_at']})")
    sc = changes["scenarios"]

    # ---- HEADLINE: movement-by-scenario matrix (gross Δ%) ----
    # Rendered LAST as live formulas pointing at the 'By scenario (gross)' detail
    # table built below; reserve its rows now so every section underneath keeps
    # its original position. Block height matches the old _changes_matrix:
    # section(2) + header(1) + one row per present scenario + TOTAL(1) + gap(2).
    HEADLINE_ROW = 5
    ents = [e for e in ["FIHL", "FUL", "FIBL", "FIID"] if (sc["entity"] == e).any()]
    scens_present = [s for s in SCEN_ORDER if (sc["scenario"] == s).any()]
    r = (HEADLINE_ROW + 5 + len(scens_present)) if ents else HEADLINE_ROW
    # interpretive one-liner — the insurance reading of the matrix
    note = _changes_narrative(sc)
    if note:
        nc = ws.cell(row=r, column=2, value=note)
        nc.font = Font(name=_FB, size=9, italic=True, color=SOFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 2

    # ---- manual-convention QoQ (exposure level) — reconciles with the manual
    #      Changes tab up to accepted data corrections ----
    r = _manual_qoq_tables(ws, r, per_layer, changes.get("prior_layers"))

    s = changes["layers"]["summary"]
    # Split the dropped set by outcome so 'dropped exposure' is not the misleading
    # raw layer-key total (most of which RENEW onto a new programme id). Mirrors
    # Book Movement: New business / Renewed / (in-progress) / genuinely Lapsed.
    from . import renewals as _rnw
    _sp = _rnw.split_movement(changes["layers"].get("new"),
                              changes["layers"].get("dropped"))

    def _sx(df):
        return (float(pd.to_numeric(df.get("per_sc"), errors="coerce").fillna(0).sum())
                if df is not None and len(df) else 0.0)
    _lap = _sp.get("lapsed")
    _COV, _PEND = ("bound_captured",), ("in_progress", "bound_missing")
    if _lap is not None and len(_lap) and "renewal_state" in getattr(_lap, "columns", []):
        gone_x = _sx(_lap[~_lap["renewal_state"].isin(_COV + _PEND)])
        pend_x = _sx(_lap[_lap["renewal_state"].isin(_PEND)])
    else:
        gone_x, pend_x = _sx(_lap), 0.0
    nb_x, ren_x = _sx(_sp.get("new_biz")), _sx(_sp.get("ren_new"))

    r = _section(ws, r, "Layer movement")
    r = _kv(ws, r, "New layers", s["new_layers"])
    r = _kv(ws, r, "Dropped layers", s["dropped_layers"])
    r = _kv(ws, r, "Layers with exposure move", s["moved_layers"])
    r = _kv(ws, r, "New business exposure", nb_x, money=True)
    r = _kv(ws, r, "Renewed (rewritten) exposure", ren_x, money=True)
    if pend_x:
        r = _kv(ws, r, "Renewals in progress (unbound)", pend_x, money=True)
    r = _kv(ws, r, "Lapsed exposure (off the book)", gone_x, money=True)
    r = _kv(ws, r, "Net move on continuing layers", s["net_move"], money=True)
    rn = ws.cell(row=r, column=2, value=(
        "Dropped layers are split by outcome: most RENEW (rewritten onto a new "
        "programme id) and stay in the book — only the 'Lapsed' line genuinely "
        "leaves. Splits match the Book Movement tab."))
    rn.font = Font(name=_FB, size=9, italic=True, color=SOFT)
    rn.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 28
    r += 2

    r = _section(ws, r, "By scenario (net)")
    sc = changes["scenarios"]
    # consistent entity then scenario ordering for the detail tables
    _ent_o = {"FIHL": 0, "FUL": 1, "FIBL": 2, "FIID": 3}
    _scn_o = {s: i for i, s in enumerate(SCEN_ORDER)}
    sc = sc.copy()
    sc["_eo"] = sc["entity"].map(_ent_o).fillna(9)
    sc["_so"] = sc["scenario"].map(_scn_o).fillna(9)
    sc = sc.sort_values(["_eo", "_so"])
    r = _hdr(ws, r, 2, ["Entity", "Scenario", "Current Net", "Prior Net",
                        "Δ Net", "Δ Net %"], [10, 16, 15, 15, 14, 11])
    for k, (_, row) in enumerate(sc.iterrows()):
        alt = k % 2 == 0
        _cell(ws, r, 2, row["entity"], alt=alt)
        _cell(ws, r, 3, row["scenario"], alt=alt)
        _cell(ws, r, 4, row["cur_net"], alt=alt, money=True)
        _cell(ws, r, 5, row["prior_net"], alt=alt, money=True)
        # Δ Net and Δ% are LIVE formulas off Current (D) / Prior (E) so they
        # recompute if either is edited — not baked values that go stale.
        _fcell(ws, r, 6, f"=D{r}-E{r}", alt=alt, money=True,
               bold=(row["d_net"] is not None and abs(row["d_net"]) > 1e6))
        _fcell(ws, r, 7, f"=IFERROR((D{r}-E{r})/E{r},0)", alt=alt, pct=True)
        r += 1
    r += 1

    r = _section(ws, r, "By scenario (gross)")
    r = _hdr(ws, r, 2, ["Entity", "Scenario", "Current Gross", "Prior Gross",
                        "Δ Gross", "Δ Gross %"], [10, 16, 15, 15, 14, 11])
    # Capture each (entity, scenario) → row so the headline matrix can reference
    # the current/prior gross cells (cols D/E) as live formulas. The table is
    # sorted entity-major, so each entity's rows are contiguous (for the TOTALs).
    gross_row = {}; gross_first = {}; gross_last = {}
    for k, (_, row) in enumerate(sc.iterrows()):
        alt = k % 2 == 0
        ent = row["entity"]; scen = row["scenario"]
        gross_row[(ent, scen)] = r
        gross_first.setdefault(ent, r); gross_last[ent] = r
        _cell(ws, r, 2, ent, alt=alt)
        _cell(ws, r, 3, scen, alt=alt)
        _cell(ws, r, 4, row.get("cur_gross"), alt=alt, money=True)
        _cell(ws, r, 5, row.get("prior_gross"), alt=alt, money=True)
        # Δ Gross / Δ% as live formulas off Current (D) / Prior (E).
        _fcell(ws, r, 6, f"=D{r}-E{r}", alt=alt, money=True,
               bold=(row.get("d_gross") is not None and abs(row.get("d_gross", 0)) > 1e6))
        _fcell(ws, r, 7, f"=IFERROR((D{r}-E{r})/E{r},0)", alt=alt, pct=True)
        r += 1
    r += 1

    # Render the reserved headline matrix now that the gross detail rows exist.
    # Manual Exec basis (FIHL combined / FIBL receiver) when the exec columns
    # are available; otherwise the Summary-basis live-formula matrix.
    if ents:
        if {"cur_gross_exec", "prior_gross_exec"} <= set(sc.columns):
            _changes_matrix_exec(ws, HEADLINE_ROW, sc)
        else:
            _changes_matrix_live(ws, HEADLINE_ROW, sc, ents, scens_present,
                                 gross_row, gross_first, gross_last,
                                 "Movement by scenario & entity  (RDS gross Δ% "
                                 "vs prior — Summary basis: FIHL ex add-ons, "
                                 "FIBL direct)")

    # top new / dropped layers
    nl = changes["layers"]["new"]
    if len(nl):
        r = _section(ws, r, f"New layers (top {min(10, len(nl))} by exposure)")
        r = _hdr(ws, r, 2, ["Spacecraft", "Entity", "Orbit", "Exposure"],
                 [28, 10, 12, 15])
        for k, (_, row) in enumerate(nl.sort_values("per_sc", ascending=False)
                                     .head(10).iterrows()):
            alt = k % 2 == 0
            _cell(ws, r, 2, str(row.get("spacecraft_name", "")), alt=alt)
            _cell(ws, r, 3, str(row.get("entity", "")), alt=alt)
            _cell(ws, r, 4, str(row.get("orbit", "")), alt=alt)
            _cell(ws, r, 5, row.get("per_sc"), alt=alt, money=True)
            r += 1
        r += 1
    # Show only GENUINE lapses — layers that actually left the book. A layer that
    # renewed onto a NEW programme id (e.g. Eutelsat 36D 344558→385575) is paired
    # by split_movement into ren_old and must NOT appear here as a loss. Mirror the
    # 'Lapsed exposure (off the book)' line above: from the lapsed set, exclude
    # renewed (bound_captured) and in-progress / bound-missing renewals.
    if _lap is not None and len(_lap) and "renewal_state" in getattr(_lap, "columns", []):
        dl = _lap[~_lap["renewal_state"].isin(_COV + _PEND)].copy()
    elif _lap is not None:
        dl = _lap.copy()
    else:
        dl = changes["layers"]["dropped"]
    if dl is not None and len(dl):
        r = _section(ws, r, f"Lapsed layers — off the book (top {min(10, len(dl))} by exposure)")
        r = _hdr(ws, r, 2, ["Spacecraft", "Entity", "Orbit", "Exposure"],
                 [28, 10, 12, 15])
        for k, (_, row) in enumerate(dl.sort_values("per_sc", ascending=False)
                                     .head(10).iterrows()):
            alt = k % 2 == 0
            _cell(ws, r, 2, str(row.get("spacecraft_name", "")), alt=alt)
            _cell(ws, r, 3, str(row.get("entity", "")), alt=alt)
            _cell(ws, r, 4, str(row.get("orbit", "")), alt=alt)
            _cell(ws, r, 5, row.get("per_sc"), alt=alt, money=True)
            r += 1

    # Renewal gaps — dropped spacecraft with NO current layer at all. These are
    # either genuine non-renewals or renewals not yet bound/entered (a 1-July
    # tranche expiring at quarter-close shows up here). Surfaced so the QoQ
    # trough is explained, not mistaken for a loss.
    rg = changes["layers"].get("renewal_gaps")
    # Drop 'covered' renewals (bound & already in the book under another programme
    # id, e.g. WorldView → 369418) — they are NOT off the book, so they must not
    # appear in the gaps list (Book Movement excludes them too).
    if rg is not None and len(rg) and "renewal_state" in getattr(rg, "columns", []):
        rg = rg[rg["renewal_state"] != "bound_captured"].copy()
    # UW dispositions keyed by programme id (reconciliation_notes) so a documented
    # gap (self-insured / future renewal / covered elsewhere) reads with its reason
    # instead of looking like an unexplained loss.
    _disp = {}
    if params is not None:
        for n in (getattr(params, "raw", {}) or {}).get("reconciliation_notes", []) or []:
            try:
                _disp[int(n.get("program_id"))] = str(n.get("category", ""))
            except (TypeError, ValueError):
                pass
    if rg is not None and len(rg):
        r += 1
        gx = float(rg["per_sc"].sum())
        r = _section(ws, r, f"Renewal gaps — {len(rg)} spacecraft fully off "
                            f"the book (${gx:,.0f})")
        c = ws.cell(row=r, column=2, value="Expired with no current/renewed "
                    "layer — verify these are true non-renewals, not renewals "
                    "not yet bound or entered in the source. 'Disposition' shows "
                    "any UW note on record (self-insured / future renewal / etc).")
        c.font = Font(name=_FB, size=9, italic=True, color=SOFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 26
        r += 2
        r = _hdr(ws, r, 2, ["Spacecraft", "Entity", "Orbit", "Exposure",
                            "Disposition"], [28, 10, 12, 15, 34])
        for k, (_, row) in enumerate(rg.sort_values("per_sc", ascending=False)
                                     .head(15).iterrows()):
            alt = k % 2 == 0
            _cell(ws, r, 2, str(row.get("spacecraft_name", "")), alt=alt)
            _cell(ws, r, 3, str(row.get("entity", "")), alt=alt)
            _cell(ws, r, 4, str(row.get("orbit", "")), alt=alt)
            _cell(ws, r, 5, row.get("per_sc"), alt=alt, money=True)
            # Disposition: UW note by programme id, else the pointer classification
            try:
                pid = int(row.get("program_id"))
            except (TypeError, ValueError):
                pid = None
            disp = _disp.get(pid) or str(row.get("renewal_label", "") or "")
            _cell(ws, r, 6, disp, alt=alt)
            r += 1

    # Pin Changes-tab column widths explicitly (this tab mixes a wide-label
    # matrix, money tables, and long merged narrative text; autofit would
    # over-widen B from the merged note, so we set them by hand and skip autofit).
    ws.column_dimensions["B"].width = 24
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["G"].width = 13


def _book_movement(wb, changes):
    """Quarter-on-quarter turnover, split into four filterable tables keyed
    Program · Layer · Entity · Spacecraft · Orbit · Inception · Expiry · Exposure.
    Exposure reads green where the book gains / red where it loses; the Renewals
    Δ is conditionally formatted by sign; each table carries its own filter.
      · New business         — spacecraft not in the prior book at all
      · Renewals             — spacecraft in BOTH books (rewritten, incl. onto a
                               new programme id); prior vs current so the reprice
                               shows, with 'Renewed from' naming the prior id
      · Renewals in progress — expired, renewal recorded but not yet bound / in
                               the book (from the J.Pbi pointer); the manual-
                               inclusion / roll-forward worklist
      · Lapsed               — genuinely off the book (NTU/Declined or no pointer)"""
    if not changes or "layers" not in changes:
        return
    L = changes["layers"]
    new = L.get("new"); dropped = L.get("dropped")
    if new is None or dropped is None or "spacecraft_id" not in getattr(new, "columns", []):
        return

    # New / Renewal / Lapsed split — pointer-primary, spacecraft-fallback. Shared
    # with outputs.export (renewal_gaps + the console summary) so every tab agrees
    # on which layers renewed vs lapsed. This is why a renewal onto a NEW
    # programme id (e.g. Eutelsat 344558 → 385575) lands in Renewals, not New.
    from . import renewals as _rnw
    sp = _rnw.split_movement(new, dropped)
    new_biz, ren_new, ren_old, lapsed = (sp["new_biz"], sp["ren_new"],
                                         sp["ren_old"], sp["lapsed"])
    ren_sid = sp["ren_sid"]

    # An expired layer that carries a renewal pointer is NOT a plain lapse — split
    # the lapsed set by the pointer's verdict:
    #   covered — renewal is BOUND and already IN the current book (a different
    #             programme id, e.g. WorldView 341568 → 369418): excluded from
    #             all Book Movement tables — coverage is counted under the
    #             successor programme, so it is a non-event here
    #   pending — renewal in progress / bound-but-missing → renewing, not yet in
    #             the book (can't pair in Renewals, but shown separately)
    #   gone    — NTU/Declined or no pointer → genuinely off the book
    _COVERED = ("bound_captured",)
    _PENDING = ("in_progress", "bound_missing")
    if lapsed is not None and len(lapsed) and "renewal_state" in lapsed.columns:
        covered = lapsed[lapsed["renewal_state"].isin(_COVERED)].copy()
        pending = lapsed[lapsed["renewal_state"].isin(_PENDING)].copy()
        gone = lapsed[~lapsed["renewal_state"].isin(_COVERED + _PENDING)].copy()
    else:
        empty = lapsed.iloc[0:0].copy() if lapsed is not None else pd.DataFrame()
        covered = empty.copy()
        pending = empty.copy()
        gone = lapsed

    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _sum(df):
        return float(pd.to_numeric(df.get("per_sc"), errors="coerce").fillna(0).sum()) if len(df) else 0.0
    nb_x = _sum(new_biz)
    pend_x, gone_x = _sum(pending), _sum(gone)
    ren_cur, ren_pri = _sum(ren_new), _sum(ren_old)

    ws = wb.create_sheet("Book Movement")
    _title(ws, "Book movement — new · renewals · in-progress · lapsed",
           f"per-layer turnover vs {changes.get('prior_as_at', 'prior')}")

    def _nsc(df):
        return df["spacecraft_id"].nunique() if df is not None and len(df) else 0

    r = _section(ws, 5, "At a glance")
    r = _kv(ws, r, "New business", nb_x, money=True)
    ws.cell(row=r - 1, column=4, value=f"{_nsc(new_biz)} spacecraft").font = F_SUB
    r = _kv(ws, r, "Renewals — current exposure", ren_cur, money=True)
    ws.cell(row=r - 1, column=4, value=f"{len(ren_sid)} spacecraft · was {ren_pri:,.0f} (Δ {ren_cur - ren_pri:+,.0f})").font = F_SUB
    r = _kv(ws, r, "Renewals in progress (not yet in book)", pend_x, money=True)
    ws.cell(row=r - 1, column=4, value=f"{_nsc(pending)} spacecraft · renewal pending — see table below").font = F_SUB
    r = _kv(ws, r, "Lapsed (not renewed / off the book)", gone_x, money=True)
    ws.cell(row=r - 1, column=4, value=f"{_nsc(gone)} spacecraft").font = F_SUB
    r += 1

    # Column order: Program · Layer · Entity · Spacecraft · Orbit · Inception ·
    # Expiry, then the exposure column(s). Headers are white (clean/print) with
    # the exposure font coloured by direction; each table is a real Excel Table
    # so all three carry their own filter dropdowns.
    IDCOLS = ["Program", "Layer", "Entity", "Spacecraft", "Orbit",
              "Inception", "Expiry"]
    IDW = [11, 8, 9, 28, 11, 12, 12]

    def _idrow(r, row, alt):
        _cell(ws, r, 2, row.get("program_id"), alt=alt)
        _cell(ws, r, 3, row.get("layer_id"), alt=alt)
        _cell(ws, r, 4, str(row.get("entity", "")), alt=alt)
        _cell(ws, r, 5, str(row.get("spacecraft_name", "")), alt=alt)
        _cell(ws, r, 6, str(row.get("orbit", "")), alt=alt)
        _dcell(ws, r, 7, row.get("inception"), alt=alt)
        _dcell(ws, r, 8, row.get("off_risk_date"), alt=alt)

    def _simple_table(r, title, df, exp_colour, tbl_name, show_renewal=False):
        # Lapsed carries a Renewal-status column when the renewal-policy check has
        # classified it (from the J.Pbi forward pointer); otherwise it's absent.
        # It also names the successor programme id (the renewal changes programme
        # id) so a "Bound — missing from source" row is actionable.
        renewal = show_renewal and df is not None and "renewal_label" in getattr(df, "columns", [])
        has_to = renewal and "renewed_to_program_id" in getattr(df, "columns", [])
        heads = IDCOLS + ["Exposure"] + (["Renewal status"] if renewal else []) \
            + (["Renewed to (prog)"] if has_to else [])
        widths = IDW + [16] + ([34] if renewal else []) + ([15] if has_to else [])
        last_col = 9 + (1 if renewal else 0) + (1 if has_to else 0)
        r = _section(ws, r, title)
        header_row = r
        r = _hdr(ws, header_row, 2, heads, widths, white=True)
        if not len(df):
            _cell(ws, r, 2, "— none —"); return r + 2
        for k, (_, row) in enumerate(df.sort_values("per_sc", ascending=False).iterrows()):
            alt = k % 2 == 0
            _idrow(r, row, alt)
            c = _cell(ws, r, 9, row.get("per_sc"), alt=alt, money=True,
                      bold=_num(row.get("per_sc")) > 1e7)
            c.font = Font(name=_FB, size=10, color=exp_colour,
                          bold=_num(row.get("per_sc")) > 1e7)
            if renewal:
                sev = RENEWAL_SEV_COLOUR.get(str(row.get("renewal_severity", "")), INK)
                rc = _cell(ws, r, 10, str(row.get("renewal_label", "")), alt=alt)
                rc.font = Font(name=_FB, size=10, color=sev,
                               bold=str(row.get("renewal_severity", "")) in ("gap", "progress"))
            if has_to:
                _cell(ws, r, 11, _progid(row.get("renewed_to_program_id")), alt=alt)
            r += 1
        _add_table(ws, tbl_name, 2, header_row, last_col, r - 1)
        return r + 1

    r = _simple_table(r, f"New business — {new_biz['spacecraft_id'].nunique() if len(new_biz) else 0} "
                         f"spacecraft (${nb_x:,.0f})", new_biz, GREEN, "BM_New")

    # Renewals — per spacecraft, prior vs current (Δ conditionally formatted)
    r = _section(ws, r, f"Renewals — {len(ren_sid)} spacecraft "
                        f"(current ${ren_cur:,.0f} · Δ {ren_cur - ren_pri:+,.0f})")
    ren_hdr = r
    # "Program" here is the CURRENT (new) programme; "Renewed from" names the
    # prior programme id so the renewal's programme-id change is explicit
    # (e.g. Eutelsat 344558 → 385575).
    r = _hdr(ws, ren_hdr, 2, IDCOLS + ["Prior exp.", "Current exp.", "Δ exposure",
                                       "Renewed from"],
             IDW + [15, 15, 15, 13], white=True)
    ren_first = r
    if ren_sid:
        # Pair prior vs current by the NORMALISED spacecraft id so a renewal onto
        # a new programme still lines up (and a float/text id mismatch never
        # orphans the prior exposure). Prior side may sit on a different programme
        # id than current — that is the reprice we want to show.
        cur_x = ren_new.groupby("_sid")["per_sc"].sum()
        pri_x = ren_old.groupby("_sid")["per_sc"].sum()
        pri_prog = ren_old.groupby("_sid")["program_id"].first()
        meta = ren_new.groupby("_sid").agg(
            {"program_id": "first", "layer_id": "first",
             "spacecraft_name": "first", "entity": "first", "orbit": "first",
             "inception": "first", "off_risk_date": "first"})
        order = cur_x.reindex(meta.index).fillna(0).sort_values(ascending=False).index
        for k, sid in enumerate(order):
            alt = k % 2 == 0
            p = float(pri_x.get(sid, 0.0)); c = float(cur_x.get(sid, 0.0))
            _idrow(r, meta.loc[sid], alt)
            _cell(ws, r, 9, p, alt=alt, money=True)
            _cell(ws, r, 10, c, alt=alt, money=True)
            _cell(ws, r, 11, c - p, alt=alt, money=True, bold=abs(c - p) > 1e6)
            _cell(ws, r, 12, _progid(pri_prog.get(sid)), alt=alt)
            r += 1
        # Δ column conditionally formatted by sign
        drng = f"K{ren_first}:K{r - 1}"
        ws.conditional_formatting.add(drng, CellIsRule(
            operator="greaterThan", formula=["0"], font=Font(name=_FB, size=10, color=GREEN)))
        ws.conditional_formatting.add(drng, CellIsRule(
            operator="lessThan", formula=["0"], font=Font(name=_FB, size=10, color="C0392B")))
        _add_table(ws, "BM_Renewals", 2, ren_hdr, 12, r - 1)
    else:
        _cell(ws, r, 2, "— none —"); r += 1
    r += 1

    # Renewals in progress — expired, renewal recorded but not yet bound/in the
    # book. Pulled OUT of "lapsed": these are renewing (the "Program" column is
    # the EXPIRING programme id, "Renewed to" the successor). Manual-inclusion /
    # roll-forward candidates. Amber exposure (pending, not a loss).
    if len(pending):
        r = _simple_table(r, f"Renewals in progress — {_nsc(pending)} spacecraft, "
                             f"renewal not yet in book (${pend_x:,.0f})", pending,
                          AMBER, "BM_Pending", show_renewal=True)
        pn = ws.cell(row=r, column=2, value=(
            "Expired, but the underwriter's pointer shows a renewal quoted / in "
            "progress (not yet bound, so absent from source). 'Program' is the "
            "expiring id, 'Renewed to' the successor — these are renewing, not "
            "lost. Candidates for roll-forward / manual inclusion if UW confirms."))
        pn.font = Font(name=_FB, size=9, italic=True, color=SOFT)
        pn.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws.row_dimensions[r].height = 28
        r += 2

    r = _simple_table(r, f"Lapsed — {_nsc(gone)} spacecraft off the book, "
                         f"not renewed (${gone_x:,.0f})", gone, "C0392B",
                      "BM_Lapsed", show_renewal=True)
    have_renewal = gone is not None and "renewal_state" in getattr(gone, "columns", [])
    if have_renewal:
        from . import renewals as _rnw
        summ = _rnw.summarize(gone["renewal_state"])
        brk = " · ".join(f"{_rnw.label(k)}: {v}" for k, v in summ.items())
        note = ws.cell(row=r, column=2, value=(
            "Genuinely off the book: " + brk + ".  'NTU/Declined' = the renewal "
            "was offered and lost; 'No renewal pointer' = no renewal recorded. "
            "(Renewals still in progress are in the table above, not here.)"))
    else:
        note = ws.cell(row=r, column=2, value="Lapsed = expired with no current layer — "
                       "verify true non-renewals, not renewals not yet bound/entered. "
                       "(Renewal-pointer feed not available this run — spacecraft "
                       "heuristic used.)")
    note.font = Font(name=_FB, size=9, italic=True, color=SOFT)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 28

    for col, w in {"B": 11, "C": 8, "D": 9, "E": 28, "F": 11, "G": 12, "H": 12,
                   "I": 16, "J": 27, "K": 15, "L": 13}.items():
        ws.column_dimensions[col].width = w


def _excluded(wb, per_layer, dropped=None):
    """Rows excluded from ALL calculations - transparency/audit.
    Source of truth is `dropped` (rows the engine removed: NULL-orbit,
    consortium, off-risk-after-correction). per_layer is checked for any
    residual out-of-scope rows as a safety net, but normally has none."""
    pl = per_layer.copy()

    # Engine-dropped rows are the authoritative excluded set.
    if dropped is not None and len(dropped):
        excl = dropped.copy()
        if "excluded_reason" in excl.columns:
            excl["_reason"] = excl["excluded_reason"].fillna("Excluded")
        else:
            excl["_reason"] = "Excluded"
    else:
        excl = pl.iloc[0:0].copy()
        excl["_reason"] = pd.Series(dtype=str)

    # Safety net: any out-of-scope rows still in per_layer (should be none now).
    orbit_s = pl["orbit"].astype(str)
    in_scope_orbit = orbit_s.isin(["GEO-GSO", "MEO", "LEO"])
    persc = pl["per_sc"].fillna(0)
    residual = pl[(~in_scope_orbit) | (persc <= 0)].copy()
    if len(residual):
        residual["_reason"] = "Residual out-of-scope (review)"
        excl = pd.concat([excl, residual], ignore_index=True)

    # TRIPWIRE: consortium declarations that slipped through SQL into per_layer.
    tripwire = pd.DataFrame()
    if "placing_basis" in pl.columns:
        pb = pl["placing_basis"].astype(str).str.strip().str.lower()
        is_cons = pb.str.startswith("consortium", na=False)
        tripwire = pl[is_cons].copy()

    if not len(excl) and not len(tripwire):
        ws = wb.create_sheet("Excluded")
        _title(ws, "Excluded from scenarios", "none this run")
        return

    ws = wb.create_sheet("Excluded")
    _title(ws, "Excluded from scenarios",
           "rows outside scenario scope — for audit; none should carry exposure")

    rr = 5
    # TRIPWIRE banner — consortium declarations that slipped through SQL
    if len(tripwire):
        ws.cell(row=rr, column=2,
                value="⚠ CONSORTIUM PLACING-BASIS LAYERS IN POPULATION — "
                      "check SQL view filter").font = Font(
                    name=_FB, size=11, bold=True, color="C62828")
        rr += 1
        ws.cell(row=rr, column=2,
                value=f"{len(tripwire)} row(s), "
                      f"${float(tripwire['per_sc'].fillna(0).sum()):,.0f} exposure. "
                      "Check Placing_Basis / view filter.").font = F_SUB
        rr += 1
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 24),
                ("placing_basis", "Placing Basis", 22), ("per_sc", "Exposure", 14)]
        rr = _hdr(ws, rr, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, (_, row) in enumerate(tripwire.sort_values(
                "per_sc", ascending=False).iterrows()):
            alt = k % 2 == 0
            for j, (field, _, _) in enumerate(cols):
                v = row.get(field)
                if field == "spacecraft_name" and (v is None or str(v) == "nan"):
                    v = "(none)"
                _cell(ws, rr, 2 + j, v, alt=alt, money=(field == "per_sc"))
            rr += 1
        rr += 1

    # Detail listing FIRST (so the summary can SUMIF against it). We write the
    # detail lower down, but compute its row range now and reference it live.
    excl = excl.sort_values(["_reason", "program_id"]).reset_index(drop=True)
    n_excl = len(excl)
    reasons = list(excl["_reason"].drop_duplicates())

    # reserve summary block rows, detail goes below
    sum_hdr = rr + 1
    sum_first = sum_hdr + 1
    sum_last = sum_first + len(reasons) - 1
    sum_total = sum_last + 1
    detail_hdr = sum_total + 4
    detail_first = detail_hdr + 1
    detail_last = detail_first + n_excl - 1
    # detail column letters: Program=B Layer=C Entity=D S/C=E Orbit=F
    #                        OrigExp=G RevExp=H Reason=I
    REASON_RANGE = f"$I${detail_first}:$I${detail_last}"
    ORIG_RANGE = f"$G${detail_first}:$G${detail_last}"

    # ---- summary by reason (live SUMIF against the detail block) ----
    _section(ws, rr, "By reason")
    r = _hdr(ws, sum_hdr, 2, ["Reason", "Rows", "Original Exposure",
                              "Revised Exposure"], [52, 10, 18, 18])
    for k, name in enumerate(reasons):
        alt = k % 2 == 0
        rref = f"$B{r}"
        _cell(ws, r, 2, name, alt=alt)
        c3 = _cell(ws, r, 3, f"=COUNTIF({REASON_RANGE},{rref})", alt=alt)
        c3.alignment = Alignment(horizontal="right")
        c4 = _cell(ws, r, 4, f"=SUMIF({REASON_RANGE},{rref},{ORIG_RANGE})", alt=alt)
        c4.number_format = MONEY; c4.alignment = Alignment(horizontal="right")
        c5 = _cell(ws, r, 5, 0, alt=alt, money=True)
        r += 1
    _cell(ws, r, 2, "Total", bold=True)
    c3 = _cell(ws, r, 3, f"=COUNTA({REASON_RANGE})", bold=True)
    c3.alignment = Alignment(horizontal="right")
    c4 = _cell(ws, r, 4, f"=SUM({ORIG_RANGE})", bold=True)
    c4.number_format = MONEY; c4.alignment = Alignment(horizontal="right")
    _cell(ws, r, 5, 0, money=True, bold=True)
    total_exp = float(excl["per_sc"].fillna(0).sum())
    r = sum_total + 2
    ws.cell(row=r, column=2,
            value="Original exposure (live SUM of detail below) revised to $0 — "
                  "excluded from all scenario calculations.").font = F_SUB

    # ---- detail listing — original exposure preserved, revised = 0 ----
    _section(ws, detail_hdr - 1, "Detail")
    cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
            ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 24),
            ("orbit", "Orbit", 10), ("orig_exp", "Original Exposure", 16),
            ("rev_exp", "Revised Exposure", 16), ("_reason", "Reason", 50)]
    _hdr(ws, detail_hdr, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
    r = detail_first
    for k, (_, row) in enumerate(excl.iterrows()):
        alt = k % 2 == 0
        for j, (field, _, _) in enumerate(cols):
            if field == "orig_exp":
                v = row.get("per_sc") or 0
            elif field == "rev_exp":
                v = 0
            else:
                v = row.get(field)
            if field == "spacecraft_name" and (v is None or str(v) == "nan"):
                v = "(none)"
            money = field in ("orig_exp", "rev_exp")
            _cell(ws, r, 2 + j, v, alt=alt, money=money)
        r += 1



def _overrides(wb, per_layer):
    """Manual Overrides audit: layers the SQL manual table added/removed.
    Reads the override_action column (present in SQL-ingest mode). Removed
    layers don't survive into per_layer, so this shows ADDS that are live, and
    the count of removes is inferred from the action tag where present."""
    ws = wb.create_sheet("Manual Overrides")
    _title(ws, "Manual overrides applied",
           "from manually_controlled_rds_layers · Action column · SQL-source runs")

    if "override_action" not in per_layer.columns:
        _section(ws, 5, "Not available")
        ws.cell(row=6, column=2,
                value="override_action not in data — run with source: sql and the "
                      "as-at query file to populate this sheet.").font = F_SUB
        return

    act = per_layer["override_action"].astype(str).str.strip().str.lower()
    added = per_layer[act.eq("add layer")].copy()

    r = _section(ws, 5, "Summary")
    r = _kv(ws, r, "Layers added by manual table (on-risk this run)", len(added))
    r = _kv(ws, r, "Added exposure (per-S/C)",
            float(added["per_sc"].sum()) if len(added) else 0.0, money=True)
    ws.cell(row=r, column=2,
            value="Note: removed layers are filtered out upstream by SQL, so they "
                  "do not appear in the population. Run the SSMS query in "
                  "STEP 2 of the script to review removes.").font = F_SUB
    r += 2

    if len(added):
        r = _section(ws, r, "Added layers")
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 26),
                ("orbit", "Orbit", 11), ("inception", "Inception", 13),
                ("per_sc", "Exposure", 15)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, (_, row) in enumerate(added.sort_values("per_sc", ascending=False)
                                     .iterrows()):
            alt = k % 2 == 0
            for j, (field, _, _) in enumerate(cols):
                v = row.get(field)
                if isinstance(v, (dt.date, dt.datetime)):
                    v = str(v)
                if field == "spacecraft_name" and (v is None or str(v) == "nan"):
                    v = "(none)"
                _cell(ws, r, 2 + j, v, alt=alt, money=(field == "per_sc"))
            r += 1



def _corrections(wb, corrections):
    if not corrections:
        return
    ws = wb.create_sheet("Data Corrections")
    _title(ws, "Data corrections applied",
           "manual fixes to known source-data errors — for transparency/audit")
    ws.cell(row=4, column=2,
            value="These override source values for named layers (e.g. wrong "
                  "coverage dates). Each is a logged, temporary patch pending "
                  "source correction.").font = F_SUB
    r = _hdr(ws, 6, 2, ["Program", "Layer", "Field", "Old value", "New value",
                        "Status", "Reason"], [11, 9, 16, 20, 20, 12, 34])
    for k, c in enumerate(corrections):
        alt = k % 2 == 0
        vals = [c.get("program_id"), c.get("layer_id"), c.get("field"),
                str(c.get("old")), str(c.get("value")), c.get("status"),
                c.get("reason", "")]
        for j, v in enumerate(vals):
            cell = _cell(ws, r, 2 + j, v, alt=alt)
            if j == 5 and v == "NOT FOUND":
                cell.font = F_FLAG
        r += 1


# New sheet builders to append to excel_report.py — modeled on PR RDS house standard

def _methodology(wb, params, changes):
    """Waterfall logic per entity + change log. Mirrors PR RDS Methodology tab."""
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 110
    ws["B2"] = "Space RDS  —  Methodology"; ws["B2"].font = F_TITLE
    ws["B3"] = (f"As at {params.as_at}   |   {params.quarter}   |   "
                "Scenario definitions, waterfall logic and change log.")
    ws["B3"].font = F_SUB

    r = _section(ws, 5, "Scenario definitions")
    scen = [
        ("Proton Flare", "5% insured loss applied to all on-risk GEO-GSO spacecraft. "
         "No time decay. Single solar-event scenario."),
        ("Space Weather", "Worst single bus-manufacturer total across ALL on-risk orbits "
         "(raw gross per-S/C, before equity/netting). Captures correlated manufacturer "
         "design-defect / solar exposure."),
        ("Generic Defect", "50% insured loss on GEO-GSO + MEO spacecraft, time-decayed by "
         "Risk Period Factor (RPF). Common-cause manufacturing defect."),
        ("Space Debris", "Orbit-specific damage ratios (LEO 40%, MEO 10%, GEO 5%) applied "
         "to on-risk spacecraft in each orbit band. Kessler-type collision scenario."),
        ("Max Risk", "Single largest on-risk spacecraft by gross per-S/C exposure. "
         "Worst single-risk loss."),
    ]
    r = _hdr(ws, r, 2, ["Scenario", "Definition"], [16, 108])
    for k, (s, d) in enumerate(scen):
        _cell(ws, r, 2, s, alt=k % 2 == 0)
        c = _cell(ws, r, 3, d, alt=k % 2 == 0)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1
    r += 1

    r = _section(ws, r, "Netting waterfall logic")
    wf = [
        ("FUL", "Gross → External QS (outwards RI slots by inception) → IGR QS 50% ceded "
         "to FIBL → IGR XoL $245m xs $50m → Net retained."),
        ("FIID", "Same cascade as FUL with IGR QS 85% and IGR XoL $72.5m xs $7.5m."),
        ("FIHL", "Group-consolidated, external cessions only (no intra-group IGR). The "
         "book's headline is the COMBINED (Exec) basis — ex-add-on gross/net PLUS the "
         "S3123 QS to IG and the IG equity share — matching the manual Change Narrative. "
         "The ex-add-on split lives on the S3123 & Equity (IG) tab."),
        ("S3123", "Inwards QS to IG = per-S/C exposure × 20% cession × the S3123 consortium "
         "time-factor (0.4167 for inception < 2026-01-01, 0.500 from 2026-01-01, 0.333 from "
         "2026-04-01), applied to S3123-ELIGIBLE layers only: is-consortium AND inception "
         "within 2024-07-01…2026-12-31 AND not on the excluded-spacecraft list. IG equity "
         "share = per-S/C × equity % (by inception year, consortium-gated) × the SAME "
         "consortium factor. Both are computed per layer and aggregate to the Summary / "
         "S3123 & Equity tab via SUMPRODUCT (live from Per Layer)."),
        ("S2126", "New consortium participant from 2026-04-01. On that date S3123's sub-share "
         "stepped 15/30 → 10/30 (0.500 → 0.333) and S2126 picked up the released 5/30 (0.167); "
         "IG base 30 and TPC 6 unchanged. S2126 gets its OWN Lloyd's RDS (same scenarios, RPF "
         "and eligibility as S3123) on a per-S/C × s2126_factor (0 before 2026-04-01, 5/30 "
         "after) share. It has NO QS-to-IG agreement, so it retains 100% (net = gross) and does "
         "NOT touch the IG book — the S3123 QS-to-IG / equity fall with the 0.333 factor and are "
         "NOT offset by S2126. Reported on the 'S2126 RDS Summary' tab. Controlling-body id 7."),
    ]
    r = _hdr(ws, r, 2, ["Entity", "Cascade"], [16, 108])
    for k, (e, d) in enumerate(wf):
        _cell(ws, r, 2, e, alt=k % 2 == 0)
        c = _cell(ws, r, 3, d, alt=k % 2 == 0)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    r = _section(ws, r, "Population & exclusion logic")
    logic = [
        "On-risk filter: Inception ≤ as-at ≤ Off-risk date (re-applied after any data correction).",
        "Consortium placing-basis layers excluded in-pipeline (SQL view filters only "
        "'Consortium Declaration'; pipeline also drops plain 'Consortium').",
        "NULL-orbit / no-Seradata-ID layers excluded from all scenario calculations "
        "(revised exposure set to 0); retained on Python Adjustments tab for audit.",
        "Manual add/remove layers sourced from rds.manually_controlled_rds_layers (SQL).",
        "Data corrections (e.g. wrong coverage period) override the source field and re-filter.",
    ]
    for k, t in enumerate(logic):
        c = _cell(ws, r, 2, "•  " + t, alt=k % 2 == 0)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    r = _section(ws, r, "Change log")
    r = _hdr(ws, r, 2, ["Quarter", "Change"], [16, 108])
    if changes:
        s = changes["layers"]["summary"]
        log = f"vs prior run ({changes['prior_as_at']}): {s['new_layers']} new, " \
              f"{s['dropped_layers']} dropped layers."
    else:
        log = "Initial automation deployment — no prior quarter to compare."
    _cell(ws, r, 2, params.quarter)
    c = _cell(ws, r, 3, log); c.alignment = Alignment(wrap_text=True)
    r += 2
    ws.cell(row=r, column=2,
            value="Pipeline: run_space_rds.py → ingest · engine · scenarios · netting · "
                  "outputs. Source of truth: SQL view rds.vw_SpaceRDS_OnRisk.").font = F_SUB


def _audit(wb, per_layer, params, source, recon, excluded):
    """Extract row counts, SQL refs, validation findings. Mirrors PR RDS Audit tab."""
    ws = wb.create_sheet("Audit")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 80
    ws["B2"] = "SPACE RDS  —  Audit"; ws["B2"].font = F_TITLE
    ws["B3"] = (f"As at {params.as_at}   |   Run {dt.datetime.now():%Y-%m-%d %H:%M:%S}"
                f"   |   source: {source}   |   SQL refs, row counts, validation.")
    ws["B3"].font = F_SUB

    r = _section(ws, 5, "Extract — connections & row counts")
    r = _hdr(ws, r, 2, ["Dataset", "Rows", "SQL source"], [40, 12, 80])
    n_layers = len(per_layer)
    n_excl = len(excluded) if excluded is not None else 0
    rows = [
        ("On-risk Space layers (final population)", n_layers,
         "rds.vw_SpaceRDS_OnRisk (LON-SQLP-V005 / SpaceTrax_Data), as-at filtered"),
        ("Excluded / adjusted layers", n_excl,
         "Pipeline exclusions: consortium, NULL-orbit, data corrections"),
        ("IGR cession rates", "—",
         "param_ig_cessions_qs / param_ig_cessions_xol by entity·code·year"),
        ("External RI cessions", "—", "param_ri_cessions by inception window"),
        ("Manual add/remove layers", "—", "rds.manually_controlled_rds_layers (Action col)"),
    ]
    for k, (d, n, s) in enumerate(rows):
        alt = k % 2 == 0
        _cell(ws, r, 2, d, alt=alt)
        c = _cell(ws, r, 3, n, alt=alt); c.alignment = Alignment(horizontal="right")
        _cell(ws, r, 4, s, alt=alt); r += 1
    r += 1

    r = _section(ws, r, "Validation findings")
    findings = []
    if recon is not None and "status" in getattr(recon, "columns", []):
        n_bad = int((recon["status"] == "MISMATCH").sum())
        n_ok = int((recon["status"] == "OK").sum())
        findings.append(f"Reconciliation vs workbook: {n_ok} checks OK, {n_bad} mismatches."
                        if n_bad else f"Reconciliation vs workbook: all {n_ok} checks OK.")
    null_orbit = int((per_layer["orbit"].astype(str).isin(["NULL", "None", "nan"])).sum())
    findings.append(f"NULL-orbit layers in final population: {null_orbit} (target 0 — "
                    "excluded upstream)." if null_orbit == 0
                    else f"NULL-orbit layers: {null_orbit} excluded, see Python Adjustments.")
    bad = int((per_layer["per_sc"] <= 0).sum())
    findings.append(f"Zero/negative exposure rows: {bad}." if bad
                    else "No zero/negative exposure rows.")
    if not findings:
        findings = ["No data-quality issues detected."]
    for k, t in enumerate(findings):
        c = _cell(ws, r, 2, t, alt=k % 2 == 0)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1

    r = _section(ws, r, "Population reconciliation")
    total = per_layer["per_sc"].sum()
    r = _kv(ws, r, "Final on-risk layers", n_layers)
    r = _kv(ws, r, "Total signed exposure (per-S/C)", total, money=True)
    r = _kv(ws, r, "Layers adjusted out / in (Python Adjustments)", n_excl)
    r += 1
    ws.cell(row=r, column=2,
            value="Full per-layer extract on the Per Layer tab. All summary tabs are live "
                  "Excel formulas referencing it (clickable audit trail).").font = F_SUB
    r += 2

    # ---- Basis of preparation: historical (back-dated) runs ----
    r = _section(ws, r, "Basis of preparation — historical (back-dated) runs")
    notes = [
        "The --as-at parameter rewinds the on-risk DATE FILTER only: it correctly "
        "reconstructs which layers were on risk at the chosen date.",
        "It does NOT rewind layer VALUES. The pipeline reads the live "
        "rds.vw_SpaceRDS_OnRisk view, so signed exposures, dates and manual "
        "add/remove decisions are returned at their CURRENT values.",
        "A back-dated run therefore reflects the historical POPULATION valued with "
        "PRESENT-DAY DATA, and may differ from figures originally filed for that "
        "quarter where source values have since been restated.",
        "For a faithful reproduction of a previously filed quarter, run in workbook "
        "mode against the saved raw extract captured at that quarter's close "
        "(source: workbook), not the live view.",
        "If no frozen extract was saved and the source tables hold no point-in-time "
        "history, the originally filed values cannot be reproduced from the live "
        "database; the filed workbook is the authoritative record.",
    ]
    for k, t in enumerate(notes):
        c = _cell(ws, r, 2, "•  " + t, alt=k % 2 == 0)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1
    # current-run flag
    r += 1
    is_workbook = str(source).lower() == "workbook"
    flag = ("This run: WORKBOOK mode — frozen extract, values as captured."
            if is_workbook else
            "This run: LIVE SQL view — values are current, not as-originally-filed.")
    fc = ws.cell(row=r, column=2, value=flag)
    fc.font = F_OK if is_workbook else F_FLAG
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)


def _python_adjustments(wb, per_layer, excluded, corrections=None, params=None,
                        manual_includes=None):
    """Merged Excluded + Manual Overrides + Data Corrections — every pipeline
    add / remove / field-override with reasons. One consolidated audit view."""
    ws = wb.create_sheet("Python Adjustments")
    ws.sheet_view.showGridLines = False
    _title(ws, "Python Adjustments",
           "automated layer removals, additions & field corrections, with "
           "documented reasons · for audit")

    # Removals (from the engine's excluded set)
    rem = excluded.copy() if (excluded is not None and len(excluded)) else None
    adds = per_layer[per_layer.get("override_action", pd.Series(dtype=str))
                     .astype(str).str.strip().str.lower() == "add layer"] \
        if "override_action" in per_layer.columns else per_layer.iloc[0:0]
    corr = corrections or []
    n_corr = len(corr)
    n_corr_applied = sum(1 for c in corr if c.get("status") == "APPLIED")

    r = _section(ws, 5, "Summary")
    n_rem = len(rem) if rem is not None else 0
    rem_exp = float(rem["per_sc"].fillna(0).sum()) if rem is not None else 0.0
    add_exp = float(adds["per_sc"].fillna(0).sum()) if len(adds) else 0.0
    r = _hdr(ws, r, 2, ["Adjustment", "Layers", "Exposure impact"], [40, 10, 18])
    _cell(ws, r, 2, "Removed (excluded from calculations)")
    _cell(ws, r, 3, n_rem); c = _cell(ws, r, 4, -rem_exp, money=True); r += 1
    _cell(ws, r, 2, "Added (manual inclusions from SQL table)", alt=True)
    _cell(ws, r, 3, len(adds), alt=True)
    _cell(ws, r, 4, add_exp, alt=True, money=True); r += 1
    _cell(ws, r, 2, "Field corrections (value/date overrides on retained layers)")
    _cell(ws, r, 3, n_corr_applied)
    _cell(ws, r, 4, "see below"); r += 1
    _cell(ws, r, 2, "Net layer adjustment", bold=True)
    _cell(ws, r, 3, len(adds) - n_rem, bold=True)
    _cell(ws, r, 4, add_exp - rem_exp, money=True, bold=True); r += 2

    # Removals detail
    if rem is not None and len(rem):
        r = _section(ws, r, f"Removals ({n_rem}) — excluded from all scenario calculations")
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 24),
                ("orbit", "Orbit", 10), ("per_sc", "Orig Exposure", 15),
                ("excluded_reason", "Reason for removal", 46)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        rem2 = rem.sort_values("excluded_reason") if "excluded_reason" in rem.columns else rem
        for k, (_, row) in enumerate(rem2.iterrows()):
            alt = k % 2 == 0
            for j, (field, _, _) in enumerate(cols):
                v = row.get(field)
                if field == "spacecraft_name" and (v is None or str(v) == "nan"):
                    v = "(none)"
                if field == "excluded_reason" and (v is None or str(v) == "nan" or str(v) == "None"):
                    v = "Out of scenario scope"
                _cell(ws, r, 2 + j, v, alt=alt, money=(field == "per_sc"))
            r += 1
        r += 1

    # Additions detail
    if len(adds):
        r = _section(ws, r, f"Additions ({len(adds)}) — manual inclusions applied via SQL")
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 24),
                ("orbit", "Orbit", 10), ("inception", "Inception", 13),
                ("per_sc", "Exposure", 15)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, (_, row) in enumerate(adds.sort_values("per_sc", ascending=False).iterrows()):
            alt = k % 2 == 0
            for j, (field, _, _) in enumerate(cols):
                v = row.get(field)
                if isinstance(v, (dt.date, dt.datetime)):
                    v = str(v)
                if field == "spacecraft_name" and (v is None or str(v) == "nan"):
                    v = "(none)"
                _cell(ws, r, 2 + j, v, alt=alt, money=(field == "per_sc"))
            r += 1
        r += 1

    # Field corrections detail (value/date overrides applied at ingest)
    if n_corr:
        r = _section(ws, r, f"Field corrections ({n_corr_applied} applied) — "
                            "value/date overrides on retained layers")
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("field", "Field", 14), ("old", "Old value", 16),
                ("value", "New value", 16), ("status", "Status", 11),
                ("logged", "Logged", 12), ("reason", "Reason", 44)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, c in enumerate(corr):
            alt = k % 2 == 0
            # "Logged" shows the date the correction was established (date_logged),
            # falling back to the quarter it was established in.
            logged = c.get("date_logged") or c.get("established") or "—"
            vals = [c.get("program_id"), c.get("layer_id"), c.get("field"),
                    _fmt_corr_val(c.get("old")), _fmt_corr_val(c.get("value")),
                    c.get("status"), str(logged), c.get("reason", "")]
            for j, v in enumerate(vals):
                cell = _cell(ws, r, 2 + j, v, alt=alt)
                if j == 5 and v == "NOT FOUND":
                    cell.font = F_FLAG
            r += 1
        r += 1

    # Manual inclusions (config `manual_include`, injected at ingest) —
    # UW-confirmed layers carried because their renewal is not yet bound in the
    # source. They flow through the whole engine; documented here for audit.
    mincs = manual_includes or []
    if mincs:
        n_inj = sum(1 for c in mincs if c.get("status") == "INCLUDED")
        mi_exp = sum(float(c.get("layer_signed_exposure") or c.get("per_sc") or 0)
                     for c in mincs if c.get("status") == "INCLUDED")
        r = _section(ws, r, f"Manual inclusions ({n_inj} injected · "
                            f"${mi_exp:,.0f}) — UW-confirmed layers injected at "
                            "ingest (renewal not yet bound in source)")
        # Renewal lineage (from the J.Pbi forward pointer, recorded in config):
        # e.g. "renews-from 344558 · Quoted/Awaiting FOT".
        for c in mincs:
            rf = c.get("renewed_from_program_id")
            us = c.get("renewal_uw_status")
            c["_renewal"] = (f"renews-from {rf} · {us}" if rf and us
                             else (f"renews-from {rf}" if rf else (us or "")))
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("entity", "Entity", 9), ("spacecraft_name", "Spacecraft", 24),
                ("orbit", "Orbit", 10), ("off_risk_date", "Off-risk", 12),
                ("layer_signed_exposure", "Exposure", 15),
                ("_renewal", "Renewal (source)", 30),
                ("status", "Status", 20), ("reason", "Reason", 44)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, c in enumerate(mincs):
            alt = k % 2 == 0
            skipped = str(c.get("status", "")).startswith("SKIPPED")
            for j, (field, _, _) in enumerate(cols):
                v = c.get(field)
                if isinstance(v, (dt.date, dt.datetime)):
                    v = str(v)
                cell = _cell(ws, r, 2 + j, v, alt=alt,
                             money=(field == "layer_signed_exposure"))
                if field == "status" and skipped:
                    cell.font = F_FLAG   # safeguard tripped — not injected
                if field == "_renewal" and v:
                    cell.font = Font(name=_FB, size=10, color=AMBER)
            r += 1
        r += 1

    # Reconciliation notes — documentary only (NO data change). Explains
    # auto-vs-manual differences that are intentional: layers kept/excluded for
    # documented reasons, timing/lag differences, methodology judgments.
    notes = (params.raw.get("reconciliation_notes") or []) if params is not None else []
    if notes:
        r = _section(ws, r, f"Reconciliation notes ({len(notes)}) — "
                            "documented auto-vs-manual differences (no data change)")
        cols = [("program_id", "Program", 11), ("layer_id", "Layer", 8),
                ("category", "Category", 22), ("logged", "Logged", 12),
                ("note", "Note", 70)]
        r = _hdr(ws, r, 2, [h for _, h, _ in cols], [w for _, _, w in cols])
        for k, nt in enumerate(notes):
            alt = k % 2 == 0
            logged = nt.get("date_logged") or nt.get("established") or "—"
            vals = [nt.get("program_id"), nt.get("layer_id"),
                    nt.get("category", ""), str(logged), nt.get("note", "")]
            for j, v in enumerate(vals):
                cell = _cell(ws, r, 2 + j, v, alt=alt)
                if j == 4:  # wrap the note text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 42
            r += 1
        r += 1

    ws.cell(row=r, column=2,
            value="Reasons: 'No orbit' = satellite not yet launched / no Seradata ID; "
                  "'Consortium placing basis' = consortium-level aggregate (excluded per "
                  "methodology); 'Off-risk after data correction' = coverage period "
                  "corrected to true expiry. Additions sourced from "
                  "rds.manually_controlled_rds_layers. Field corrections override source "
                  "values for named layers (logged, traceable patches); NOT FOUND = the "
                  "layer was not in this run's population. Reconciliation notes are "
                  "documentary only — they explain intentional auto-vs-manual "
                  "differences and change no data.").font = F_SUB


def _fmt_corr_val(v):
    """Render an old/new correction value cleanly (dates as ISO, numbers grouped)."""
    if v is None or str(v) in ("None", "nan"):
        return "—"
    if isinstance(v, (dt.date, dt.datetime)):
        return str(v if not isinstance(v, dt.datetime) else v.date())
    if isinstance(v, (int, float)):
        return f"{v:,.2f}"
    return str(v)


def _cover(wb, params, per_layer, grid):
    """Cover page — classification, key facts, contents. Makes it a deliverable."""
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 30), ("C", 30), ("D", 30), ("E", 24)]:
        ws.column_dimensions[col].width = w

    # top classification band
    band = ws.cell(row=2, column=2, value="INTERNAL — CONFIDENTIAL")
    band.font = Font(name=_FB, size=9, bold=True, color="FFFFFF")
    band.fill = PatternFill("solid", start_color=NAVY)
    band.alignment = Alignment(horizontal="center")
    for col in range(2, 6):
        ws.cell(row=2, column=col).fill = PatternFill("solid", start_color=NAVY)

    # big title block
    ws.cell(row=6, column=2, value="Space").font = Font(
        name=_FB, size=40, bold=True, color=NAVY)
    ws.cell(row=8, column=2, value="Realistic Disaster Scenario").font = Font(
        name=_FB, size=26, bold=False, color=INK)
    ws.cell(row=10, column=2, value=f"{params.quarter}  ·  as at {params.as_at}").font = Font(
        name=_FB, size=14, color=SOFT)

    # rule
    for col in range(2, 6):
        ws.cell(row=12, column=col).border = Border(
            bottom=Side(style="medium", color=NAVY))

    # key facts block
    total_exp = per_layer["per_sc"].sum()
    n_layers = len(per_layer)
    fihl_sw = grid[(grid["entity"] == "FIHL") &
                   (grid["scenario"] == "Space Weather")]
    fihl_worst = _display_gross(fihl_sw.iloc[0]) if len(fihl_sw) else 0

    facts = [
        ("Reporting entity", "The Fidelis Partnership"),
        ("Function", "Specialty Exposure Management"),
        ("Line", "Space — all entities (FIHL / FUL / FIID / S3123)"),
        ("Basis", "Per-spacecraft signed exposure, USD"),
        ("On-risk layers", f"{n_layers:,}"),
        ("Total signed exposure", f"${total_exp/1e6:,.1f}m"),
        ("Worst-case (FIHL gross)", f"${fihl_worst/1e6:,.1f}m  ·  Space Weather"),
    ]
    r = 14
    for k, (label, val) in enumerate(facts):
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name=_FB, size=10, color=SOFT)
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = Font(name=_FB, size=11, bold=True, color=INK)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1
    r += 1

    # rule
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = Border(
            bottom=Side(style="thin", color=RULE))
    r += 2

    # prepared / reviewed / approved
    ws.cell(row=r, column=2, value="Prepared by").font = Font(
        name=_FB, size=9, color=SOFT)
    ws.cell(row=r, column=3, value="Reviewed by").font = Font(
        name=_FB, size=9, color=SOFT)
    ws.cell(row=r, column=4, value="Run timestamp").font = Font(
        name=_FB, size=9, color=SOFT)
    r += 1
    ws.cell(row=r, column=2, value="Exposure Management (automated)").font = F_CELL
    ws.cell(row=r, column=3, value="________________").font = F_CELL
    ws.cell(row=r, column=4, value=f"{dt.datetime.now():%Y-%m-%d %H:%M}").font = F_CELL
    r += 3

    # contents
    r = _section(ws, r, "Contents")
    contents = [
        ("Executive Summary", "Worst-case by entity, ranking, watch items"),
        ("Summary", "Full netting cascade per scenario & entity"),
        ("Portfolio", "Exposure concentration by entity, orbit, manufacturer"),
        ("Space Weather / Max Risk", "Key scenario detail"),
        ("Netting Waterfalls", "Gross-to-net cascade per entity"),
        ("Python Adjustments", "Automated add/remove with documented reasons"),
        ("Per Layer", "Full layer-level data backbone"),
        ("Methodology / Parameters / Audit", "Assumptions, treaty terms, provenance"),
    ]
    for k, (tab, desc) in enumerate(contents):
        tc = ws.cell(row=r, column=2, value=tab)
        tc.font = Font(name=_FB, size=10, bold=True, color=NAVY)
        dc = ws.cell(row=r, column=3, value=desc)
        dc.font = Font(name=_FB, size=9, color=INK)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1


def _stat_card(ws, col, label, value, sub, accent=None):
    """A KPI 'card': big number with label above and context below, boxed."""
    accent = accent or NAVY
    top, bottom = 5, 9
    thin = Side(style="thin", color=RULE)
    med = Side(style="medium", color=accent)
    # clean outer box only (no internal gridlines)
    for rr in range(top, bottom + 1):
        for cc in (col, col + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.border = Border(
                left=med if cc == col else None,
                right=med if cc == col + 1 else None,
                top=med if rr == top else None,
                bottom=thin if rr == bottom else None)
    # accent header strip
    for cc in (col, col + 1):
        ws.cell(row=top, column=cc).fill = PatternFill("solid", start_color=accent)
    lc = ws.cell(row=top, column=col, value=label.upper())
    lc.font = Font(name=_FB, size=8, bold=True, color="FFFFFF")
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col + 1)
    # big value
    vc = ws.cell(row=top + 2, column=col, value=value)
    vc.font = Font(name=_FB, size=24, bold=True, color=accent)
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=top + 2, start_column=col, end_row=top + 3, end_column=col + 1)
    # sub
    sc = ws.cell(row=top + 4, column=col, value=sub)
    sc.font = Font(name=_FB, size=8, color=SOFT)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=top + 4, start_column=col, end_row=top + 4, end_column=col + 1)


def _money_m(v):
    """Format a number as $Xm for headline display."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"${v/1e6:,.1f}m"


def _exec_summary(wb, grid, per_layer, params, source, recon, excluded, changes):
    """Executive summary — the answer first. Worst-case net per entity, worst-case
    scenario, ranking chart, watch items, Q-on-Q, data quality. Senior-reader view."""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    for col, w in [("B", 26), ("C", 22), ("D", 16), ("E", 16), ("F", 14)]:
        ws.column_dimensions[col].width = w

    ws["B2"] = "Space Realistic Disaster Scenario"
    ws["B2"].font = Font(name=_FB, size=18, bold=True, color=NAVY)
    ws["B3"] = (f"{params.quarter}  ·  as at {params.as_at}  ·  "
                f"Specialty Exposure Management  ·  The Fidelis Partnership")
    ws["B3"].font = Font(name=_FB, size=10, color=SOFT)

    g = grid.copy()
    total_exp = per_layer["per_sc"].sum()
    n_layers = len(per_layer)

    # KPI stat cards — three headline numbers, dashboard style
    fihl_sw = g[(g["entity"] == "FIHL") & (g["scenario"] == "Space Weather")]
    fihl_gross = _display_gross(fihl_sw.iloc[0]) if len(fihl_sw) else 0
    fihl_net = _display_net(fihl_sw.iloc[0]) if len(fihl_sw) else 0
    _stat_card(ws, 2, "Worst case · FIHL gross", f"${fihl_gross/1e6:,.0f}m",
               "Space Weather · all orbits")
    _stat_card(ws, 4, "Worst case · FIHL net", f"${fihl_net/1e6:,.0f}m",
               "Space Weather · net retained", accent="2D6A3C")
    _stat_card(ws, 6, "Total signed exposure", f"${total_exp/1e6:,.0f}m",
               f"{n_layers:,} on-risk layers")

    def money_m_cell(r, c, v, alt=False, bold=False):
        cell = ws.cell(row=r, column=c, value=(v if v is not None and
                       not (isinstance(v, float) and pd.isna(v)) else None))
        cell.number_format = '#,##0.0,,"m";(#,##0.0,,"m");"–"'
        cell.font = Font(name=_FB, size=10, bold=bold, color=INK)
        cell.alignment = Alignment(horizontal="right")
        if alt:
            cell.fill = FILL_ALT
        return cell

    # ---- HEADLINE ----
    r = _section(ws, 11, "Headline — worst-case loss by entity")
    r = _hdr(ws, r, 2, ["Entity", "Worst-case scenario", "Gross", "Net retained",
                        "Net/Gross"], [26, 22, 16, 16, 12])
    headline = {}
    for ent in ["FIHL", "FUL", "FIID", "FIBL"]:
        block = g[g["entity"] == ent]
        if not len(block):
            continue
        net_col = "net" if block["net"].notna().any() else "gross"
        worst = block.loc[block[net_col].idxmax()]
        headline[ent] = worst
        alt = ent != "FIHL"
        _cell(ws, r, 2, ent, alt=alt, bold=True)
        _cell(ws, r, 3, worst["scenario"], alt=alt)
        money_m_cell(r, 4, _display_gross(worst), alt=alt)
        net_val = _display_net(worst)
        if net_val is not None:
            money_m_cell(r, 5, net_val, alt=alt, bold=True)
            _cell(ws, r, 6, net_val / _display_gross(worst) if _display_gross(worst) else 0,
                  alt=alt, pct=True)
        else:
            c = _cell(ws, r, 5, "gross only", alt=alt); c.alignment = Alignment(horizontal="right")
            _cell(ws, r, 6, "—", alt=alt)
        r += 1
    r += 1

    # ---- one-line story ----
    binding = headline["FUL"]["scenario"]
    ful_net = headline["FUL"]["net"]
    ws.cell(row=r, column=2,
            value=f"▸  {binding} is the worst-case scenario across all entities. "
                  f"FUL net {_money_m(ful_net)}; FIID capped at its IGR XoL "
                  f"retention; FIHL group gross {_money_m(_display_gross(headline['FIHL']))}.")
    ws.cell(row=r, column=2).font = Font(name=_FB, size=10, bold=True, color=INK)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 30
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    r += 2

    r += 1

    # ---- all scenarios ranked, with retention efficiency (FUL) ----
    r = _section(ws, r, "All scenarios — FUL gross → net & retention (ranked)")
    r = _hdr(ws, r, 2, ["Scenario", "Detail", "Gross", "Ceded", "Net retained",
                        "Retained %"], [22, 20, 15, 14, 15, 12])
    ful = g[g["entity"] == "FUL"].sort_values("net", ascending=False)
    for k, (_, row) in enumerate(ful.iterrows()):
        alt = k % 2 == 1
        gross = float(row["gross"]) if pd.notna(row["gross"]) else None
        net = float(row["net"]) if pd.notna(row["net"]) else None
        ceded = (gross - net) if (gross is not None and net is not None) else None
        ret = (net / gross) if (gross and net is not None) else None
        _cell(ws, r, 2, row["scenario"], alt=alt, bold=(k == 0))
        _cell(ws, r, 3, row.get("detail") or "", alt=alt)
        money_m_cell(r, 4, gross, alt=alt)
        money_m_cell(r, 5, ceded, alt=alt)
        money_m_cell(r, 6, net, alt=alt, bold=(k == 0))
        _cell(ws, r, 7, ret, alt=alt, pct=True)
        r += 1
    r += 2

    # ---- WATCH ITEMS (judgement callout) ----
    r = _section(ws, r, "Watch items — judgement & dependencies")
    watch = []
    # data corrections active?
    n_corr = len(params.raw.get("data_corrections", []) or [])
    if n_corr:
        watch.append(("Data corrections active", AMBER,
                      f"{n_corr} layer(s) have source-date overrides (e.g. VIASAT "
                      "coverage period). Placeholder dates pending true expiries — "
                      "see Data Corrections tab."))
    # consortium filter gap
    watch.append(("SQL view consortium gap", AMBER,
                  "Live vw_SpaceRDS_OnRisk filters only 'Consortium Declaration'; "
                  "pipeline additionally drops plain 'Consortium'. BAU view "
                  "over-includes until the one-line view fix is applied."))
    # manufacturer grouping
    watch.append(("Manufacturer grouping", SOFT,
                  "Astrium (Toulouse) and Airbus D&S (Toulouse) treated as separate "
                  "manufacturers in Space Weather. Corporately one entity — grouping "
                  "would raise the worst-manufacturer total. Methodology decision."))
    # null orbit
    null_orbit_excl = (excluded is not None and len(excluded))
    if null_orbit_excl:
        watch.append(("Excluded layers", SOFT,
                      f"{len(excluded)} layer(s) excluded from calculations "
                      "(NULL-orbit / consortium / corrected). Documented with reasons "
                      "on the Python Adjustments tab."))
    for k, (title, colour, text) in enumerate(watch):
        tc = ws.cell(row=r, column=2, value="•  " + title)
        tc.font = Font(name=_FB, size=9, bold=True, color=colour)
        dc = ws.cell(row=r, column=3, value=text)
        dc.font = Font(name=_FB, size=9, color=INK)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1
    r += 1

    # ---- portfolio & run status ----
    r = _section(ws, r, "Portfolio & run status")
    r = _kv(ws, r, "On-risk layers", n_layers)
    _cell(ws, r - 1 + 1, 2, "")  # spacing guard
    kr = r
    ws.cell(row=kr, column=2, value="Total signed exposure").font = F_KEY
    money_m_cell(kr, 3, total_exp, bold=True)
    r += 1
    geo_share = per_layer[per_layer["orbit"] == "GEO-GSO"]["per_sc"].sum() / total_exp
    r = _kv(ws, r, "GEO-GSO share of exposure", f"{geo_share:.0%}")
    n_adj = len(excluded) if excluded is not None else 0
    r = _kv(ws, r, "Layers adjusted (Python Adjustments)", n_adj)
    if changes and changes.get("layers"):
        s = changes["layers"]["summary"]
        r = _kv(ws, r, "vs prior quarter",
                f"+{s['new_layers']} new / -{s['dropped_layers']} dropped")
    else:
        r = _kv(ws, r, "vs prior quarter", "no prior run on file — baseline")
    dq_ok = True; msgs = []
    if recon is not None and "status" in getattr(recon, "columns", []):
        nbad = int((recon["status"] == "MISMATCH").sum())
        if nbad:
            dq_ok = False; msgs.append(f"{nbad} reconciliation mismatches")
    null_orbit = int((per_layer["orbit"].astype(str).isin(["NULL","None","nan"])).sum())
    if null_orbit:
        dq_ok = False; msgs.append(f"{null_orbit} NULL-orbit in population")
    status = "All checks passed" if dq_ok else "; ".join(msgs)
    ws.cell(row=r, column=2, value="Data quality").font = F_KEY
    sc = ws.cell(row=r, column=3, value=status)
    sc.font = F_OK if dq_ok else F_FLAG
    r += 2

    ws.cell(row=r, column=2,
            value="How to read this book: Summary (full netting cascade) · scenario "
                  "tabs · Portfolio (concentration) · Python Adjustments (add/remove "
                  "& why) · Methodology / Parameters / Audit (assumptions & provenance).")
    ws.cell(row=r, column=2).font = F_SUB
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=6)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

def _s3123_ig_addons(wb, grid, params, per_layer=None):
    """The IG add-ons and the ex-add-on split, on their own tab.

    DECIDED (user, 2026Q1): FIHL/FIBL headline everywhere in the book — Summary,
    movement matrix and QA — is COMBINED (FIHL incl S3123 QS + IG equity; FIBL
    on the receiver basis), matching the manual's Change Narrative / Executive
    Summary. This tab is where the EX-add-on figure and the component split
    live, so nothing is lost. The syndicate's OWN Lloyd's return stays on the
    "S3123 RDS" tab.
    """
    ws = wb.create_sheet("S3123 & Equity (IG)")
    _title(ws, "S3123 QS & IG Equity - group add-ons & conventions",
           "The book's headline FIHL/FIBL figures are COMBINED (manual Change "
           "Narrative basis): FIHL includes S3123 QS to IG + IG Equity; FIBL is "
           "the internal receiver. This tab records that convention and carries "
           "the ex-add-on split for reference (verified to the manual to the "
           "dollar).")
    # ---- conventions note (single source of truth for how each entity shows) ----
    notes = [
        ("FIHL", "COMBINED — all-entity gross of the worst manufacturer/orbit "
                 "PLUS S3123 QS to IG PLUS IG equity share. This is the group's "
                 "true consolidated event loss and matches the manual Change "
                 "Narrative / Exec Summary. Ex-add-on gross shown as a memo "
                 "column on the Summary."),
        ("FIBL", "RECEIVER — QS received (from FUL+FIID) + XoL received + FIBL "
                 "direct book + S3123 QS to IG + IG equity. Matches the manual "
                 "Change Narrative FIBL rows. Max Risk = the single bird FIBL "
                 "RECEIVES the most on (SPAINSAT NG-2, ~21.4m); every component "
                 "is that one bird's own cession — a coherent single-satellite "
                 "loss (Basis B, UW-confirmed). The filed 2026Q1 manual keyed "
                 "~22.9m by stitching a different bird's larger IGR (SPAINSAT "
                 "NG-1) onto NG-2's S3123/equity — not a single loss; superseded."),
        ("FUL / FIID", "Operating entities — standard cede-down cascade (Gross "
                       "-> Ext QS -> IGR QS -> XoL -> Net). No IG add-ons apply, "
                       "so combined = ex here."),
        ("S3123 RDS tab", "The syndicate's OWN Lloyd's return — separate from "
                          "the IG book. The 20% QS it cedes back to IG is the "
                          "SAME cession shown as 'S3123 QS to IG' here; the two "
                          "must always agree."),
        ("How computed", "S3123 QS to IG = per-S/C exposure × 20% cession × the "
                         "S3123 consortium time-factor (0.4167 / 0.500 / 0.333 by "
                         "inception: <2026-01-01 / ≥2026-01-01 / ≥2026-04-01), on "
                         "S3123-eligible layers only (is-consortium AND inception in "
                         "2024-07-01…2026-12-31 AND not excluded). IG equity = per-S/C "
                         "× equity % (by inception year) × the same factor. For the "
                         "additive scenarios these cells are LIVE — SUMPRODUCT over "
                         "Per Layer (contribution × per-layer rate) — so they tie to "
                         "the engine to the cent. Space Weather / Max Risk are "
                         "selections (worst manufacturer / single bird), shown as "
                         "engine values; reconcile on their own tabs."),
    ]
    r = 5
    for who, txt in notes:
        c = ws.cell(row=r, column=2, value=who); c.font = F_SECT
        c2 = ws.cell(row=r, column=3, value=txt); c2.font = F_SUB
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 42
        r += 1
    r += 1
    fihl = grid[grid["entity"] == "FIHL"].set_index("scenario")
    fibl = grid[grid["entity"] == "FIBL"].set_index("scenario")
    scens = [s for s in ["Proton Flare", "Space Weather", "Generic Defect",
                         "Space Debris", "Max Risk"] if s in fihl.index]

    r = _section(ws, r, "FIHL - group view (combined)")
    r = _hdr(ws, r, 2, ["Scenario", "FIHL ex add-ons", "S3123 QS to IG", "IG Equity",
                        "FIHL incl add-ons (headline)", "Net incl add-ons"],
             [16, 22, 16, 14, 18, 18])
    # This tab is the SINGLE SOURCE of the FIHL ex-add-on split; the Summary's
    # memo columns point here as live references (see _summary addon_refs), so
    # the split is keyed once. Capture each scenario's cell refs to hand back.
    # Live backbone: for the additive scenarios the three components are computed
    # in-workbook by SUMPRODUCT over Per Layer (contribution x per-layer rate), so
    # nothing is baked. Space Weather / Max Risk are selections (worst manufacturer
    # / single bird), not per-layer sums, so they stay engine values with a note.
    _plc, _ = _pl_map(per_layer) if per_layer is not None else ({}, [])
    _plN = len(per_layer) if per_layer is not None else 0
    fihl_ref = {}
    fihl_scen_row = {}
    for i, scen in enumerate(scens):
        row = fihl.loc[scen]
        _cell(ws, r, 2, scen, alt=i % 2)
        contrib = f"{_SP_SCEN[scen]}_fihl" if scen in _SP_SCEN else None
        gr_rng = _pl_range(_plc, _plN, contrib) if contrib else None
        s3f = _sumproduct(_plc, _plN, contrib, "s3123_qs_pp") if contrib else None
        eqf = _sumproduct(_plc, _plN, contrib, "equity_pp") if contrib else None
        if gr_rng and s3f and eqf:      # additive scenario -> live
            cC = _cell(ws, r, 3, f"=SUM({gr_rng})", alt=i % 2, money=True)
            _note(cC, "LIVE: SUM of the per-layer FIHL contribution for this scenario "
                      "on Per Layer. Reconciles to the engine to the cent.")
            cDd = _cell(ws, r, 4, s3f, alt=i % 2, money=True)
            _note(cDd, "LIVE: SUMPRODUCT(FIHL contribution, s3123_qs/per_sc) over Per "
                       "Layer = SUM per layer of (contribution x s3123 factor x 20%).")
            cEe = _cell(ws, r, 5, eqf, alt=i % 2, money=True)
            _note(cEe, "LIVE: SUMPRODUCT(FIHL contribution, equity_usd/per_sc) over Per "
                       "Layer = SUM per layer of (contribution x equity% x s3123 factor).")
        else:                            # Space Weather / Max Risk -> engine value
            cC = _cell(ws, r, 3, float(row.get("gross", 0)), alt=i % 2, money=True)
            _note(cC, "Selection scenario (worst manufacturer / single bird), not a "
                      "per-layer sum. Engine value; reconcile on the Space Weather / "
                      "Max Risk tab.")
            _cell(ws, r, 4, float(row.get("s3123_qs", 0) or 0), alt=i % 2, money=True)
            _cell(ws, r, 5, float(row.get("equity_usd", 0) or 0), alt=i % 2, money=True)
        c = _cell(ws, r, 6, f"=SUM(C{r}:E{r})", alt=i % 2, bold=True)
        c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
        _cell(ws, r, 7, float(row.get("net_incl_addons", 0) or 0), alt=i % 2, money=True)
        fihl_ref[scen] = {"gross_ex": f"C{r}", "s3123_qs": f"D{r}", "equity": f"E{r}"}
        fihl_scen_row[scen] = r
        r += 1
    r += 1

    if len(fibl):
        r = _section(ws, r, "FIBL - internal receiver view")
        r = _hdr(ws, r, 2, ["Scenario", "FIBL direct book", "QS received",
                            "XoL received", "S3123 QS", "IG Equity",
                            "FIBL receiver total"],
                 [16, 20, 16, 14, 14, 12, 20])
        for i, scen in enumerate([s for s in scens if s in fibl.index]):
            row = fibl.loc[scen]
            _cell(ws, r, 2, scen, alt=i % 2)
            cbC = _cell(ws, r, 3, float(row.get("gross", 0) or 0), alt=i % 2, money=True)
            _note(cbC, "FIBL direct book for this scenario. Reconcile on Per Layer "
                       "(FIBL-direct layers) / Netting Waterfalls (FIBL block).")
            cbD = _cell(ws, r, 4, float(row.get("qs_received", 0) or 0), alt=i % 2, money=True)
            _note(cbD, "Intra-group QS RECEIVED from FUL + FIID (= the sum of their "
                       "IGR QS Ceded on the Summary). Drill: Netting Waterfalls.")
            _cell(ws, r, 5, float(row.get("xol_received", 0) or 0), alt=i % 2, money=True)
            # S3123 QS & IG Equity are the SAME group add-ons as the FIHL block —
            # reference those cells so they are keyed once, not re-baked here.
            _fr = fihl_scen_row.get(scen)
            if _fr:
                _cell(ws, r, 6, f"=D{_fr}", alt=i % 2, money=True)
                _cell(ws, r, 7, f"=E{_fr}", alt=i % 2, money=True)
            else:
                _cell(ws, r, 6, float(row.get("s3123_qs", 0) or 0), alt=i % 2, money=True)
                _cell(ws, r, 7, float(row.get("equity_usd", 0) or 0), alt=i % 2, money=True)
            c = _cell(ws, r, 8, f"=SUM(C{r}:G{r})", alt=i % 2, bold=True)
            c.number_format = MONEY; c.alignment = Alignment(horizontal="right")
            r += 1
        r += 1
    ws.cell(row=r + 1, column=2,
            value="S3123 QS to IG is the same 20% cession that nets down the "
                  "syndicate's own return on the S3123 RDS tab - the two tabs "
                  "must always agree on it.").font = F_SUB
    return {"sheet": ws.title, "fihl": fihl_ref}


def write_results(path, per_layer, sw, mr, grid, params, source,
                  recon=None, changes=None, excluded=None,
                  corrections=None, manual_includes=None,
                  s3123_grid=None, s3123_recon=None, s3123_notes=None,
                  s3123_qoq=None):
    # FIX(recon 2026Q1, O2): the S3123 sheet is now written HERE when a grid is
    # supplied, so the runner cannot forget to wire it (the shipped 2026Q1 book
    # carried $0 nets and a $0 Space Weather because the sheet was fed a stale /
    # incomplete grid out-of-band). Passing the grid through write_results also
    # guarantees it is the same object the Summary reconciles against.
    wb = Workbook()
    # Per-layer effective cession rates (per $ of exposure). These give the
    # Summary & S3123 tabs a live SUMPRODUCT backbone: scenario cession =
    # SUM(per-layer contribution x this rate). Guarded against per_sc<=0.
    if "per_sc" in per_layer.columns:
        per_layer = per_layer.copy()
        _ps = per_layer["per_sc"].where(per_layer["per_sc"] > 0)
        for _src, _dst in (("ext_qs", "ext_qs_pp"), ("igr_qs_ceded", "igr_ceded_pp"),
                           ("s3123_qs", "s3123_qs_pp"), ("equity_usd", "equity_pp")):
            if _src in per_layer.columns:
                per_layer[_dst] = (per_layer[_src] / _ps).fillna(0.0)
    # Live-formula spec for the S3123/S2126 contribution columns so their Per
    # Layer cells are transparent formulas, not baked numbers: gross = per_sc x
    # factor x elig x loss x rpf (gated by the selection flag); net = gross x
    # cession (excluded s/c retain 100%). Loss/RPF/QS all read from config.
    _contrib = {}
    for _pfx, _cfgk, _faccol in (("s3123", "s3123_rds", "s3123_factor"),
                                 ("s2126", "s2126_rds", "s2126_factor")):
        _c = params.raw.get(_cfgk) or {}
        if not _c:
            continue
        _scn = _c.get("scenarios", {}) or {}
        _qs = float(_c.get("qs_to_ig", 0.20))
        _excl = set(_c.get("qs_to_ig_excluded", []) or [])
        _lm = {
            "pf": (float((_scn.get("proton_flare") or {}).get("loss", 0.05)), False),
            "sw": (float((_scn.get("space_weather") or {}).get("loss", 1.0)), False),
            "gd": (float((_scn.get("generic_defect") or {}).get("loss", 0.5)),
                   bool((_scn.get("generic_defect") or {}).get("apply_rpf", True))),
            "sd": (float((_scn.get("space_debris") or {}).get("loss", 1.0)),
                   bool((_scn.get("space_debris") or {}).get("apply_rpf", True))),
        }
        for _slug, (_loss, _rpf) in _lm.items():
            _g, _n = f"{_pfx}_{_slug}_g", f"{_pfx}_{_slug}_n"
            if _g in per_layer.columns:
                _contrib[_g] = {"kind": "gross", "factor": _faccol,
                                "elig": f"{_pfx}_rds_elig", "sel": f"{_pfx}_{_slug}_sel",
                                "loss": _loss, "rpf": _rpf}
            if _n in per_layer.columns:
                _contrib[_n] = {"kind": "net", "gross": _g, "qs": _qs, "excl": _excl}
    colmap, _ = _pl_map(per_layer)
    _cover(wb, params, per_layer, grid)
    _exec_summary(wb, grid, per_layer, params, source, recon, excluded, changes)
    _control(wb, per_layer, params, source, recon, colmap)
    # Build the S3123 & Equity (IG) tab FIRST so it is the single source of the
    # FIHL ex-add-on split; the Summary's memo columns reference it live. Final
    # tab order is imposed by the `order` list below, so build order is free.
    addon_refs = _s3123_ig_addons(wb, grid, params, per_layer)   # C1: split, live
    ws_sum = wb.create_sheet("Summary")
    summary_map = _summary(ws_sum, grid, params, source, per_layer,
                           addon_refs=addon_refs)
    _portfolio(wb, per_layer, params, colmap, changes)
    _parameters(wb, params)
    _waterfalls(wb, grid)
    _space_weather(wb, sw, colmap)
    _max_risk(wb, mr, grid)
    _python_adjustments(wb, per_layer, excluded, corrections, params,
                        manual_includes=manual_includes)
    _methodology(wb, params, changes)
    _audit(wb, per_layer, params, source, recon, excluded)
    _changes(wb, changes, per_layer, params)
    try:
        _book_movement(wb, changes)
    except Exception as _e:  # noqa: BLE001 - detail tab is non-essential
        print(f"      WARNING: Book Movement tab skipped ({_e})")
    try:
        _change_narrative(wb, changes, per_layer, mr, sw, params)
    except Exception as _e:  # noqa: BLE001 - narrative helper is non-essential
        print(f"      WARNING: Change Narrative tab skipped ({_e})")
    from . import waterfalls
    try:
        waterfalls.build_movement_waterfalls(wb, per_layer, grid, changes, params,
                                             prior_layers=(changes or {}).get("prior_layers"))
    except Exception as _e:  # noqa: BLE001 - WF tabs are non-essential
        print(f"      WARNING: movement waterfalls skipped ({_e})")
    try:
        charts_xlsx.build_charts_sheet(
            wb,
            summary="Summary",
            space_weather="Space Weather",
            xol_terms=params.igr_xol,
            changes=changes,
        )
    except Exception as _e:  # noqa: BLE001 - charts are decorative
        print(f"      WARNING: charts sheet skipped ({_e}) — the Changes tab "
              f"layout changed; charts_xlsx scans it and needs its anchors "
              f"updated. Deliverable tabs are unaffected.")
    _per_layer(wb, per_layer, contrib=_contrib)
    _rds_input_template(wb, grid, s3123_grid, params, summary_map)   # paste-ready regulator block
    if s3123_grid is not None and len(s3123_grid):
        from .s3123_sheet import write_s3123_sheet
        write_s3123_sheet(wb, s3123_grid, recon=s3123_recon,
                          notes=s3123_notes, as_at=str(params.as_at),
                          qoq=s3123_qoq)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    order = ["Cover", "Executive Summary", "Summary", "RDS Input Template",
             "S3123 & Equity (IG)", "S3123 RDS", "Charts",
             "Changes", "Change Narrative", "Book Movement",
             "WF · Exposure Bridge", "WF · Loss Movement",
             "Portfolio", "Netting Waterfalls", "Space Weather", "Max Risk",
             "Per Layer", "Python Adjustments", "Control",
             "Methodology", "Parameters", "Chart Data", "Audit"]
    wb._sheets.sort(key=lambda s: order.index(s.title)
                    if s.title in order else 99)
    # print hygiene + autofit on every sheet
    _no_autofit = {"Charts", "Cover", "Executive Summary", "Changes",
                   "WF · Exposure Bridge", "WF · Loss Movement"}
    for ws in wb.worksheets:
        try:
            _print_setup(ws, params, ws.title)
            if ws.title not in _no_autofit:
                _autofit(ws)
        except Exception:
            pass
    # Replace computed values on Per Layer and Netting Waterfalls with live
    # formulas (treaty chain + scenario rates from Parameters; netting mirrors
    # Summary). Runs after every source tab exists.
    try:
        from . import linkify
        linkify.linkify(wb)
    except Exception as _e:  # noqa: BLE001
        # Don't fail silently: a linkify error leaves the numbers HARD-CODED
        # (values still correct, but not live formulas). Surface it so a layout
        # drift can't quietly un-link the workbook again.
        import traceback
        print(f"      WARNING: linkify skipped ({_e}) — tabs will show baked "
              f"values, not live formulas. Check tab layouts vs linkify locators.")
        traceback.print_exc()
    # Embed the tab-dependency map as its own worksheet (after Cover).
    try:
        from . import tabmap
        tabmap.add_tab_map(wb)
    except Exception:
        pass
    # Workbook-level organisation: section tab colours + grouped, hyperlinked
    # Contents + frozen data-tab headers. Runs last so it sees the final layout.
    try:
        from . import organise
        organise.organise(wb)
    except Exception:
        pass
    wb.save(path)
    return path
