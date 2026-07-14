"""Ingest raw layer data from the workbook (validation mode), CSV, or SQL."""
from __future__ import annotations
import datetime as dt
import re
import pandas as pd
from openpyxl import load_workbook

# Raw columns as they appear in Input Data B:AA (the SQL paste area)
RAW_COLS = {
    2: "program_id", 3: "layer_id", 5: "coverage", 6: "months_on_risk",
    7: "spacecraft_id", 8: "launch_date", 9: "inception", 10: "expiry",
    11: "orbit", 12: "on_risk_date", 13: "off_risk_date",
    14: "bus_manufacturer", 15: "prime_manufacturer",
    19: "program_type", 20: "underwriting_status", 21: "program_name",
    22: "entity", 23: "mapping_code", 24: "spacecraft_name",
    25: "bus_family", 26: "vehicle_family", 27: "layer_signed_exposure",
    32: "is_consortium",            # AF — consortium participation flag
}

# Workbook computed columns, captured only for reconciliation
WB_COMPUTED = {
    16: "wb_rpf", 17: "wb_on_risk_flag", 29: "wb_months_left",
    30: "wb_leo_debris_rpf", 31: "wb_debris_dr",
    33: "wb_equity_pct", 34: "wb_equity_usd", 35: "wb_layer_occ",
    36: "wb_exposure_usd", 37: "wb_per_sc",
    51: "wb_ext_qs", 54: "wb_net_ext_qs", 56: "wb_s3123_eligible",
    57: "wb_s3123_qs", 58: "wb_net_ext", 59: "wb_igr_qs_rate",
    60: "wb_igr_qs_ceded", 61: "wb_net_of_qs", 62: "wb_xol_ceded",
    64: "wb_net_of_xol",
    77: "wb_pf_fihl", 78: "wb_pf_ful", 79: "wb_pf_fiid", 81: "wb_pf_fibl_ceded",
    87: "wb_sw_fihl", 88: "wb_sw_ful", 89: "wb_sw_fiid", 91: "wb_sw_fibl_ceded",
    96: "wb_gd_fihl", 97: "wb_gd_ful", 98: "wb_gd_fiid", 100: "wb_gd_fibl_ceded",
    103: "wb_sd_fihl", 104: "wb_sd_ful", 105: "wb_sd_fiid", 107: "wb_sd_fibl_ceded",
    72: "wb_maxrisk_fihl", 73: "wb_maxrisk_net",
}


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s or s.upper() in ("NULL", "N/A", "NA", "-", "TBC"):
        return None
    try:
        return dt.date.fromisoformat(s[:10])
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d %b %Y", "%d-%b-%Y",
                "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None   # unparseable date string -> blank (surfaces in data quality)


def load_from_workbook(path: str, sheet: str = "Input Data",
                       first_row: int = 20, with_wb_values: bool = True) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    cols = dict(RAW_COLS)
    if with_wb_values:
        cols.update(WB_COMPUTED)
    max_col = max(cols.keys())

    records = []
    for row in ws.iter_rows(min_row=first_row, max_col=max_col, values_only=True):
        pid = row[1]  # col B (0-indexed 1)
        if pid is None:
            continue
        rec = {name: row[c - 1] for c, name in cols.items()}
        records.append(rec)
    wb.close()

    df = pd.DataFrame(records)
    for col in ["launch_date", "inception", "expiry", "on_risk_date", "off_risk_date"]:
        df[col] = df[col].map(_to_date)
    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


def load_from_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["launch_date", "inception", "expiry",
                                        "on_risk_date", "off_risk_date"])
    for col in ["launch_date", "inception", "expiry", "on_risk_date", "off_risk_date"]:
        df[col] = df[col].dt.date
    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


# ---------------- SQL ingest ----------------

_SQL_TARGETS = {
    "programid": "program_id", "layerid": "layer_id", "coverage": "coverage",
    "monthsonrisk": "months_on_risk", "spacecraftid": "spacecraft_id",
    "launchdate": "launch_date", "inception": "inception", "expiry": "expiry",
    "orbitcategory": "orbit", "orbit": "orbit",
    "onriskdate": "on_risk_date", "offriskdate": "off_risk_date",
    "busmanufacturer": "bus_manufacturer", "primemanufacturer": "prime_manufacturer",
    "programtype": "program_type", "underwritingstatus": "underwriting_status",
    "programname": "program_name", "entity": "entity", "mappingcode": "mapping_code",
    "spacecraftname": "spacecraft_name", "busfamily": "bus_family",
    "vehiclefamily": "vehicle_family",
    "layersignedexposure": "layer_signed_exposure",
    "layersignedexposureusd": "layer_signed_exposure",
    "isconsortium": "is_consortium",
    "action": "override_action",   # 'Add Layer' / 'Remove Layer' from manual table
    "placingbasis": "placing_basis",
    # Renewal pointers from the PBI layer snapshot (J.Pbi.Layers_t) — feed the
    # renewal-policy check. Absent in workbook/CSV runs; optional everywhere.
    "renewedtoprogramid": "renewed_to_program_id",
    "renewedfromprogramid": "renewed_from_program_id",
    "renewedtouwstatus": "renewed_to_status",
    "renewedtoinception": "renewed_to_inception",
    # View's own computed columns — captured as sql_* for engine-vs-view QA
    "qsigr": "sql_igr_qs_rate",
    "monthsleftonrisk": "sql_months_left",
    "externalqs": "sql_ext_qs",
    "netofexternalqs": "sql_net_ext_qs",
    "qsfiblceded": "sql_igr_qs_ceded",
    "netofqsigr": "sql_net_of_qs",
    "xolfiblceded": "sql_xol_ceded",
    "totalfiblceded": "sql_total_fibl_ceded",
    "netofxoligr": "sql_net_of_xol",
}
_SQL_REQUIRED = {"program_id", "layer_id", "spacecraft_id", "inception",
                 "off_risk_date", "orbit", "bus_manufacturer", "entity",
                 "mapping_code", "spacecraft_name", "layer_signed_exposure"}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _pick_driver():
    import pyodbc
    for want in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server",
                 "SQL Server Native Client 11.0", "SQL Server"):
        if want in pyodbc.drivers():
            return want
    raise RuntimeError("No SQL Server ODBC driver installed")


def load_from_sql(sql_cfg: dict, non_consortium: set | None = None) -> pd.DataFrame:
    """Load from the SQL view with automatic column-name mapping.

    Config block (config/<quarter>.yaml):
        ingest:
          source: sql
          non_consortium_spacecraft: [17163, 17950]
          sql:
            server: LON-SQLP-V005
            database: SpaceTrax_Data
            view: rds.vw_SpaceRDS_OnRisk
    """
    import pyodbc
    driver = sql_cfg.get("driver") or _pick_driver()
    conn = (f"DRIVER={{{driver}}};SERVER={sql_cfg['server']};"
            f"DATABASE={sql_cfg['database']};Trusted_Connection=yes;")
    if "18" in driver:
        conn += "TrustServerCertificate=yes;"
    cn = pyodbc.connect(conn, timeout=30)
    # Custom query takes precedence (supports {as_at} placeholder for
    # historical/as-at runs); query_file loads it from a .sql file.
    query = sql_cfg.get("query")
    if not query and sql_cfg.get("query_file"):
        from pathlib import Path as _P
        query = _P(sql_cfg["query_file"]).read_text()
    if query:
        query = query.format(as_at=sql_cfg.get("as_at", ""))
    else:
        query = f"SELECT * FROM {sql_cfg['view']}"
    raw = pd.read_sql(query, cn)
    cn.close()

    rename = {}
    for col in raw.columns:
        target = _SQL_TARGETS.get(_norm(col))
        if target and target not in rename.values():
            rename[col] = target
    df = raw.rename(columns=rename)

    missing = _SQL_REQUIRED - set(df.columns)
    if missing:
        raise ValueError(
            f"SQL view missing required fields: {sorted(missing)}. "
            f"View columns were: {list(raw.columns)}")

    for col in ["launch_date", "inception", "expiry", "on_risk_date",
                "off_risk_date", "renewed_to_inception"]:
        if col in df.columns:
            df[col] = df[col].map(_to_date)

    # Renewal pointers: sentinel 0 → NA (SQL NULLIFs, but be defensive for other
    # sources), ids kept nullable-Int so a missing pointer isn't a spurious 0.
    for col in ["renewed_to_program_id", "renewed_from_program_id"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").replace(0, pd.NA)
            df[col] = df[col].astype("Int64")

    # Consortium flag: from the view if present, else config exceptions list
    if "is_consortium" in df.columns:
        df["is_consortium"] = df["is_consortium"].map(
            lambda v: str(v).strip().lower() in ("1", "true", "yes", "y"))
    else:
        nc = non_consortium or set()
        df["is_consortium"] = ~df["spacecraft_id"].isin(nc)

    # Optional descriptive fields the view doesn't carry
    for col in ["program_type", "underwriting_status"]:
        if col not in df.columns:
            df[col] = None

    df["layer_signed_exposure"] = pd.to_numeric(df["layer_signed_exposure"],
                                                errors="coerce")
    df = df[df["layer_signed_exposure"].notna()].reset_index(drop=True)

    df["layer_key"] = df["program_id"].astype(str) + "_" + df["layer_id"].astype(str)
    return df


def _apply_corrections(df, params):
    """Apply per-layer data corrections from config (e.g. wrong off-risk dates,
    restated exposures). Each correction:
        {program_id, layer_id, field, value, reason}
    Logged to load.last_corrections for the Data Corrections sheet. Date fields
    are parsed; per_sc/exposure corrections also sync layer_signed_exposure so
    the value flows through every downstream calc."""
    corrections = params.raw.get("data_corrections") or []
    applied = []
    df = df.copy()
    df["_pid"] = df["program_id"].astype(str)
    df["_lid"] = df["layer_id"].astype(str)
    DATE_FIELDS = ("off_risk_date", "on_risk_date", "inception", "expiry",
                   "launch_date")
    for c in corrections:
        mask = (df["_pid"] == str(c["program_id"])) & \
               (df["_lid"] == str(c["layer_id"]))
        n = int(mask.sum())
        if not n:
            applied.append({**c, "status": "NOT FOUND", "old": None})
            continue
        field = c["field"]
        old = df.loc[mask, field].iloc[0] if field in df.columns else None
        val = c["value"]
        if field in DATE_FIELDS:
            val = _to_date(val)
        df.loc[mask, field] = val
        # keep the raw signed-exposure column in sync with a per_sc correction
        if field in ("per_sc", "layer_signed_exposure"):
            for sync in ("per_sc", "layer_signed_exposure"):
                if sync in df.columns:
                    df.loc[mask, sync] = val
        applied.append({**c, "status": "APPLIED", "old": old, "rows": n})
    df = df.drop(columns=["_pid", "_lid"])
    load.last_corrections = applied
    return df


def _norm_id(s: pd.Series) -> pd.Series:
    """'344664' from int 344664, float 344664.0, or str '344664.0' alike."""
    return pd.to_numeric(s, errors="coerce").astype("Int64").astype(str)


def _join_lloyds(df: pd.DataFrame, view: pd.DataFrame) -> pd.DataFrame:
    """Join Lloyd's bus type (and, when present, the syndicate's actual signed
    line) onto the layer frame. Pure function — testable without SQL.

    FIX(recon 2026Q1, SW n/a): the old join failed silently two ways —
    (a) it SELECTed a hardcoded column name (any other spelling raised, was
    caught, and fell through to None), and (b) it stringified ids without
    normalising, so a float ProgramId ('344664.0') never matched the layer_key
    ('344664_1') and mapped nothing without an error. Columns are now
    discovered by normalised name, ids are normalised on both sides, the join
    is keyed spacecraft-first (bus type is a spacecraft attribute) with a
    layer fallback, and coverage is reported loudly via last_stats.
    """
    cols = {_norm(c): c for c in view.columns}

    def pick(*names):
        for n in names:
            if n in cols:
                return cols[n]
        return None

    c_bt = pick("lloydsbustype", "lloydbustype", "lloydsbus", "bustype")
    c_pid = pick("programid")
    c_lid = pick("layerid")
    c_scid = pick("spacecraftid")
    c_share = pick("layersignedexposure", "layersignedexposureusd",
                   "signedexposureusd", "signedexposure", "s3123share")
    out = df.copy()
    stats = {"rows_in_view": len(view), "bt_col": c_bt, "share_col": c_share,
             "bt_matched": 0, "share_matched": 0}

    if c_bt:
        if c_scid and "spacecraft_id" in out.columns:
            lut = dict(zip(_norm_id(view[c_scid]), view[c_bt]))
            out["lloyds_bus_type"] = _norm_id(out["spacecraft_id"]).map(lut)
        else:
            out["lloyds_bus_type"] = pd.NA
        if c_pid and c_lid:   # fallback: layer-key join for any gaps
            vkey = _norm_id(view[c_pid]) + "_" + _norm_id(view[c_lid])
            lut2 = dict(zip(vkey, view[c_bt]))
            dkey = (_norm_id(out["program_id"]) + "_"
                    + _norm_id(out["layer_id"]))
            miss = out["lloyds_bus_type"].isna()
            out.loc[miss, "lloyds_bus_type"] = dkey[miss].map(lut2)
        stats["bt_matched"] = int(out["lloyds_bus_type"].notna().sum())
    elif "lloyds_bus_type" not in out.columns:
        out["lloyds_bus_type"] = None

    # O1: capture the syndicate's ACTUAL signed line per layer so s3123.py can
    # use it (share: lloyds_view) instead of the schedule-derived share, once
    # the share-source decision is agreed with JJ.
    if c_share and c_pid and c_lid:
        vkey = _norm_id(view[c_pid]) + "_" + _norm_id(view[c_lid])
        share = pd.to_numeric(view[c_share], errors="coerce")
        lut3 = (pd.DataFrame({"k": vkey, "v": share})
                .groupby("k")["v"].sum().to_dict())
        dkey = _norm_id(out["program_id"]) + "_" + _norm_id(out["layer_id"])
        out["lloyds_signed_share"] = dkey.map(lut3)
        stats["share_matched"] = int(out["lloyds_signed_share"].notna().sum())
    _join_lloyds.last_stats = stats
    return out


def attach_lloyds_bus_type(df: pd.DataFrame, params) -> pd.DataFrame:
    """Add the Lloyd's standardised bus type (e.g. 'Airbus Eurostar 3000 and
    Eurostar NEO') used by the S3123 / Lloyd's RDS scenarios, plus the
    syndicate's per-line signed share when the view carries it. Source order:

      1. already present on the ingested frame (the Lloyd's SQL view carries it);
      2. a join from the Lloyd's bus-type view (rds.vw_SpaceRDS_All_Lloyds_RDS);
      3. an optional fallback mapping CSV (layer_key,lloyds_bus_type).

    A failed or empty join now WARNS LOUDLY instead of silently leaving NA —
    a missing bus type renders Space Weather as n/a on the S3123 tab.
    """
    cfg = (params.raw.get("s3123_rds") or {}).get("lloyds_source", {})
    if "lloyds_bus_type" in df.columns and df["lloyds_bus_type"].notna().any():
        return df
    df = df.copy()

    # PREFERRED: a spacecraft-keyed attributes query (Lloyd's bus type from the
    # (Bus Type, Bus Manufacturer) map + latest altitude), which avoids the
    # broken vw_SpaceRDS_*_Lloyds views entirely. Returns SpacecraftId |
    # LloydsBusType | AltitudeKm; keyed by spacecraft_id onto per-layer.
    qf = cfg.get("spacecraft_attrs_query_file")
    if qf and params.ingest.get("source") == "sql":
        try:
            from pathlib import Path
            import pyodbc
            import warnings as _w
            sc = dict(params.ingest["sql"])
            drv = sc.get("driver") or _pick_driver()
            conn = (f"DRIVER={{{drv}}};SERVER={sc['server']};"
                    f"DATABASE={sc['database']};Trusted_Connection=yes;")
            if "18" in str(drv):
                conn += "TrustServerCertificate=yes;"
            cn = pyodbc.connect(conn, timeout=30)
            with _w.catch_warnings():
                _w.simplefilter("ignore")
                attrs = pd.read_sql(Path(qf).read_text(), cn)
            am = {_norm(c): c for c in attrs.columns}
            c_sid = am.get("spacecraftid") or am.get("seradataspacecraftid")
            c_bt = am.get("lloydsbustype") or am.get("lloydssatellitebustypelist")
            c_alt = am.get("altitudekm") or am.get("altitudelatestkm")
            if c_sid and "spacecraft_id" in df.columns:
                key = _norm_id(df["spacecraft_id"])
                asid = _norm_id(attrs[c_sid])
                if c_bt:
                    df["lloyds_bus_type"] = key.map(dict(zip(asid, attrs[c_bt])))
                if c_alt:
                    df["altitude_km"] = key.map(dict(zip(
                        asid, pd.to_numeric(attrs[c_alt], errors="coerce"))))
                nbt = int(df.get("lloyds_bus_type", pd.Series(dtype=object)).notna().sum())
                nal = int(df.get("altitude_km", pd.Series(dtype=float)).notna().sum())
                print(f"      Lloyd's attrs: bus type on {nbt}/{len(df)} layers, "
                      f"altitude on {nal}/{len(df)} layers")
                if nbt:
                    return df
        except Exception as e:  # noqa: BLE001
            print(f"      Lloyd's attrs query failed ({e}); trying the view")

    view = cfg.get("bus_type_view")
    if view and params.ingest.get("source") == "sql":
        try:
            sql_cfg = dict(params.ingest["sql"])
            driver = sql_cfg.get("driver") or _pick_driver()
            import pyodbc
            conn = (f"DRIVER={{{driver}}};SERVER={sql_cfg['server']};"
                    f"DATABASE={sql_cfg['database']};Trusted_Connection=yes;")
            if "18" in driver:
                conn += "TrustServerCertificate=yes;"
            cn = pyodbc.connect(conn, timeout=30)
            import warnings as _w
            with _w.catch_warnings():
                _w.simplefilter("ignore")   # pandas' non-SQLAlchemy warning
                # FIX: SELECT * forces evaluation of EVERY view column, and
                # this view contains one with a broken varchar->int CAST
                # (dies on spacecraft 'BJ-3C 01'). Discover the schema with
                # TOP 0 (no row evaluation), then select ONLY the columns we
                # need; degrade column-by-column if one is still poisoned.
                hdr = pd.read_sql(f"SELECT TOP 0 * FROM {view}", cn)
                colmap = {_norm(c): c for c in hdr.columns}
                def _pick(*names):
                    for n in names:
                        if n in colmap:
                            return colmap[n]
                    return None
                want = [_pick("programid"), _pick("layerid"),
                        _pick("spacecraftid"),
                        _pick("lloydsbustype", "lloydbustype",
                              "lloydsbus", "bustype"),
                        _pick("layersignedexposure",
                              "layersignedexposureusd",
                              "signedexposureusd", "signedexposure",
                              "s3123share")]
                want = [c for c in want if c]
                if not want:
                    raise ValueError(
                        f"no usable columns found; view has {list(hdr.columns)}")
                def _sel(cols):
                    q = ", ".join(f"[{c}]" for c in cols)
                    return pd.read_sql(f"SELECT {q} FROM {view}", cn)
                try:
                    m = _sel(want)
                except Exception:
                    # one of the wanted columns is the poisoned one — find it
                    ok = []
                    for c in want:
                        try:
                            _sel([c] if not ok else ok + [c])
                            ok.append(c)
                        except Exception:
                            print(f"      lloyds view: column [{c}] is "
                                  f"broken in SQL (conversion error) — "
                                  f"skipped")
                    if not ok:
                        raise
                    m = _sel(ok)
            cn.close()
            df = _join_lloyds(df, m)
            st = _join_lloyds.last_stats
            print(f"      lloyds view: {st['rows_in_view']} rows · bus type "
                  f"col '{st['bt_col']}' matched {st['bt_matched']}/{len(df)} "
                  f"layers · signed share matched {st['share_matched']}")
            if st["bt_matched"] == 0:
                print("      WARNING: bus-type join matched NOTHING — "
                      "Space Weather will render n/a. Check the view's "
                      "id columns vs the layer population.")
            return df
        except Exception as e:  # noqa: BLE001 - view optional; fall through
            print(f"      WARNING: lloyds bus-type view unavailable ({e}) — "
                  f"Space Weather will render n/a unless the fallback CSV "
                  f"is configured")
    csv_path = cfg.get("fallback_mapping_csv")
    if csv_path:
        m = pd.read_csv(csv_path, dtype={"layer_key": str})
        lut = dict(zip(m["layer_key"], m["lloyds_bus_type"]))
        df["lloyds_bus_type"] = df["layer_key"].map(lut)
    elif "lloyds_bus_type" not in df.columns:
        df["lloyds_bus_type"] = None
    return df


def check_consortium_split_coverage(params):
    """Pre-flight data check: the extract's consortium population INNER-JOINs
    rds.param_consortium_splits on the MGU cession covering each layer's
    inception, so a consortium layer incepting AFTER the latest MGU split's End
    Date is silently dropped (this is what hid the Turksat 6A renewal). Warn if
    the MGU split table doesn't reach the as-at date. Returns a warning string or
    None; SQL-only and never raises."""
    ing = getattr(params, "ingest", {}) or {}
    if ing.get("source") != "sql":
        return None
    sql_cfg = dict(ing.get("sql") or {})
    try:
        import pyodbc
        driver = sql_cfg.get("driver") or _pick_driver()
        conn = (f"DRIVER={{{driver}}};SERVER={sql_cfg['server']};"
                f"DATABASE={sql_cfg['database']};Trusted_Connection=yes;")
        if "18" in driver:
            conn += "TrustServerCertificate=yes;"
        cn = pyodbc.connect(conn, timeout=30)
        row = pd.read_sql(
            "SELECT MAX([End Date]) AS mx FROM [SpaceTrax_Data].[rds]."
            "param_consortium_splits WHERE [Controlling Body]='MGU'", cn)
        cn.close()
    except Exception:  # noqa: BLE001 — optional check, never break the run
        return None
    if not len(row) or pd.isna(row["mx"].iloc[0]):
        return None
    mx = pd.to_datetime(row["mx"].iloc[0]).date()
    as_at = getattr(params, "as_at", None)
    if as_at is not None and mx < as_at:
        return (f"MGU consortium-split table ends {mx} (before as-at {as_at}) — "
                f"any consortium layer incepting after {mx} is EXCLUDED from the "
                f"extract (missing split row). Extend rds.param_consortium_splits "
                f"or those layers will be silently dropped.")
    return None


def _rollforward_dicts(prior, rules):
    """Turn `renewal_rollforward` rules into manual_include dicts by cloning the
    prior book's layers for each `from_program` into `to_program` at the renewal
    dates, carrying the per-bird spacecraft / manufacturer / orbit / exposure so
    the fleet's scenario granularity is preserved. Pure; returns a list."""
    out = []
    for rule in rules or []:
        fp = str(rule["from_program"])
        tp = rule["to_program"]
        sub = prior[prior["program_id"].astype(str) == fp]
        for _, r in sub.iterrows():
            scid = r.get("spacecraft_id")
            scid = int(scid) if pd.notna(scid) else None
            out.append({
                "program_id": tp,
                "layer_id": int(r["layer_id"]),
                "spacecraft_id": scid,
                "spacecraft_name": r.get("spacecraft_name"),
                "entity": r.get("entity"),
                "mapping_code": r.get("mapping_code") or rule.get("mapping_code", "ASO"),
                "orbit": r.get("orbit"),
                "bus_manufacturer": r.get("bus_manufacturer"),
                "is_consortium": bool(r.get("is_consortium", False)),
                "inception": rule["inception"],
                "on_risk_date": rule.get("on_risk_date", rule["inception"]),
                "off_risk_date": rule["off_risk_date"],
                "layer_signed_exposure": float(r["per_sc"]),
                "renewed_from_program_id": rule["from_program"],
                "renewal_uw_status": rule.get("renewal_uw_status", ""),
                "renewal_snapshot": rule.get("renewal_snapshot", ""),
                "reason": rule.get("reason",
                                   f"Renewal roll-forward {fp} → {tp} (prior line)"),
            })
    return out


def _rollforward_includes(params):
    """Build manual_include dicts for any `renewal_rollforward` rules by reading
    the frozen prior workbook. Returns [] (never raises) if there are no rules or
    the prior book is unavailable — the run then just carries the hand entries."""
    rules = params.raw.get("renewal_rollforward") or []
    if not rules:
        return []
    prior_path = params.raw.get("prior_workbook")
    if not prior_path:
        print("      renewal roll-forward: skipped (no prior_workbook configured)")
        return []
    try:
        from . import prior_seed
        prior = prior_seed.load_prior_workbook(prior_path)
    except Exception as e:  # noqa: BLE001
        print(f"      renewal roll-forward: skipped (prior load failed: {e})")
        return []
    gen = _rollforward_dicts(prior, rules)
    if gen:
        gx = sum(g["layer_signed_exposure"] for g in gen)
        print(f"      renewal roll-forward: {len(gen)} layer(s) (${gx:,.0f}) "
              f"generated from prior book")
    return gen


def _apply_manual_includes(df, params, extra=None):
    """Inject UW-confirmed manual layers (e.g. renewals not yet bound in the
    source as-at the run date) so they flow through the ENTIRE engine — rpf,
    cessions, scenarios, netting, summary, Max Risk. Each config entry supplies
    the raw layer inputs; the engine recomputes everything derived. Rows are
    tagged manual_include=True and logged to load.last_manual_includes for the
    Python Adjustments tab. `extra` carries roll-forward-generated dicts. Each entry:
        {program_id, layer_id, spacecraft_id, spacecraft_name, entity,
         mapping_code, orbit, bus_manufacturer, is_consortium, inception,
         on_risk_date, off_risk_date, layer_signed_exposure, reason}
    """
    incs = (params.raw.get("manual_include") or []) + list(extra or [])
    load.last_manual_includes = []
    df = df.copy()
    if "manual_include" not in df.columns:
        df["manual_include"] = False
    if not incs:
        return df
    DATE_FIELDS = ("off_risk_date", "on_risk_date", "inception", "expiry",
                   "launch_date")
    existing_keys = set(df["layer_key"].astype(str)) if "layer_key" in df.columns else set()
    added_keys = set()   # intra-list dedup (a hand entry + a roll-forward for the same layer)
    rows, logged = [], []
    for it in incs:
        # SAFEGUARD against double-counting: once the source binds the renewal,
        # skip the manual inclusion. Trip if the same layer_key is already in the
        # feed, OR the spacecraft already has a layer incepting on/after this
        # inclusion's inception (i.e. the renewal is now live in source).
        lk = f"{it.get('program_id')}_{it.get('layer_id')}"
        scid = str(it.get("spacecraft_id"))
        inc_date = _to_date(it.get("inception"))
        skip = None
        if lk in added_keys:
            skip = "duplicate manual entry (same layer_key already injected)"
        elif lk in existing_keys:
            skip = "already in source feed (same layer_key)"
        elif (scid and inc_date is not None and "spacecraft_id" in df.columns
              and "inception" in df.columns):
            same = df[df["spacecraft_id"].astype(str) == scid]
            if len(same):
                inc2 = pd.to_datetime(same["inception"], errors="coerce").dt.date
                if (inc2.dropna() >= inc_date).any():
                    skip = "renewal already bound in source (spacecraft on-risk from inception)"
        if skip:
            logged.append({**it, "status": f"SKIPPED — {skip}"})
            continue

        row = {c: None for c in df.columns}
        for k, v in it.items():
            if k == "reason":
                continue
            row[k] = _to_date(v) if k in DATE_FIELDS else v
        # derived / synced fields the engine relies on
        if it.get("program_id") is not None and it.get("layer_id") is not None:
            row["layer_key"] = f"{it['program_id']}_{it['layer_id']}"
        exp = it.get("layer_signed_exposure", it.get("per_sc"))
        if exp is not None:
            row["layer_signed_exposure"] = float(exp)
            row["per_sc"] = float(exp)
        if row.get("on_risk_date") is None and row.get("inception") is not None:
            row["on_risk_date"] = row["inception"]
        row["manual_include"] = True
        rows.append(row)
        added_keys.add(lk)
        logged.append({**it, "status": "INCLUDED"})
    add = pd.DataFrame(rows).reindex(columns=df.columns)
    out = pd.concat([df, add], ignore_index=True)
    load.last_manual_includes = logged
    return out


def _reconcile_rollforward(source_df, params, extra):
    """Reconcile CARRIED renewals (renewal_rollforward + manual_include) against
    what the SOURCE extract already bound, and WARN on any overlap the per-bird
    safeguard in `_apply_manual_includes` cannot catch.

    The per-bird safeguard dedups on layer_key / spacecraft_id, so it is blind to
    a source row that carries the SAME programme but with a NULL spacecraft_id
    (e.g. an ARABSAT / Eutelsat 'Fleet IO — Top Layer' aggregate). When that
    happens the roll-forward birds and the source aggregate can BOTH survive →
    double-count, or the source aggregate lands NULL-orbit and misses the GEO
    scenarios. Non-destructive: this only reports (retiring a carry rule that
    still supplies per-bird orbit detail could silently drop exposure — the
    analyst decides). Stored on load.last_rollforward_reconcile.
    """
    load.last_rollforward_reconcile = []
    if source_df is None or "program_id" not in source_df.columns:
        return
    src_pids = set(source_df["program_id"].dropna().astype(str))
    carried = {}
    for r in (params.raw.get("renewal_rollforward") or []):
        if r.get("to_program") is not None:
            carried.setdefault(str(r["to_program"]), "renewal_rollforward")
    for g in (extra or []):
        if g.get("program_id") is not None:
            carried.setdefault(str(g["program_id"]), "renewal_rollforward")
    for it in (params.raw.get("manual_include") or []):
        if it.get("program_id") is not None:
            carried.setdefault(str(it["program_id"]), "manual_include")
    hits = []
    for pid, kind in carried.items():
        if pid not in src_pids:
            continue                       # carried and NOT in source → correct
        sub = source_df[source_df["program_id"].astype(str) == pid]
        named = int(sub["spacecraft_id"].notna().sum()) if "spacecraft_id" in sub else 0
        hits.append({"program_id": pid, "carried_by": kind,
                     "source_layers": int(len(sub)), "source_per_bird": named,
                     "source_null_spacecraft": int(len(sub) - named)})
    load.last_rollforward_reconcile = hits
    if hits:
        print(f"      ⚠  renewal reconcile: {len(hits)} carried programme(s) also "
              f"present in the source feed — verify no double-count:")
        for h in hits:
            print(f"         · programme {h['program_id']} ({h['carried_by']}): "
                  f"source has {h['source_layers']} layer(s) "
                  f"[{h['source_per_bird']} per-bird, "
                  f"{h['source_null_spacecraft']} NULL-spacecraft]. "
                  "Per-bird rows are deduped by the safeguard; NULL-spacecraft "
                  "aggregates are NOT — retire the carry rule or exclude the "
                  "source aggregate.")


def load(params) -> pd.DataFrame:
    ing = params.ingest
    load.last_corrections = []
    load.last_manual_includes = []
    load.last_rollforward_reconcile = []
    if ing["source"] == "workbook":
        df = load_from_workbook(ing["workbook_path"], ing.get("sheet", "Input Data"),
                                ing.get("first_data_row", 20))
    elif ing["source"] == "csv":
        df = load_from_csv(ing["csv_path"])
    elif ing["source"] == "sql":
        nc = set(ing.get("non_consortium_spacecraft", []) or [])
        sql_cfg = dict(ing["sql"])
        sql_cfg.setdefault("as_at", str(getattr(params, "as_at", "")))
        df = load_from_sql(sql_cfg, non_consortium=nc)
    else:
        raise ValueError(f"Unknown source {ing['source']}")
    df = _apply_corrections(df, params)
    extra = _rollforward_includes(params)
    _reconcile_rollforward(df, params, extra)   # WARN on carried-vs-source overlap
    df = _apply_manual_includes(df, params, extra=extra)
    if (params.raw.get("s3123_rds") or {}).get("enabled"):
        df = attach_lloyds_bus_type(df, params)
    return df


def load_lloyds_rds_summary(params):
    """Read JJ's two Lloyd's RDS views verbatim for the 'Lloyds RDS Summary' tab:

      * rds.vw_SpaceRDS_All_Lloyds_RDS      -> the four headline RDS
      * rds.vw_SpaceRDS_SpaceWeather_Lloyds -> the Space-Weather bus-type block

    The RDS methodology lives in these views (same database as the IG book), so
    the tab ties to JJ's filed 'Lloyds RDS Summary-final' to the dollar without
    re-deriving anything. Returns {'rds': df|None, 'space_weather': df|None} or
    None when not configured. NEVER raises — a missing/unreadable view degrades
    to None and the sheet renders a visible banner instead.
    """
    cfg = ((params.raw.get("s3123_rds") or {}).get("lloyds_summary") or {})
    if not cfg.get("enabled"):
        return None
    if params.ingest.get("source") != "sql":
        print("      Lloyd's RDS summary: skipped (ingest source is not SQL)")
        return None
    try:
        import pyodbc
        sql_cfg = dict(params.ingest["sql"])
        driver = sql_cfg.get("driver") or _pick_driver()
        conn = (f"DRIVER={{{driver}}};SERVER={sql_cfg['server']};"
                f"DATABASE={sql_cfg['database']};Trusted_Connection=yes;")
        if "18" in str(driver):
            conn += "TrustServerCertificate=yes;"
        cn = pyodbc.connect(conn, timeout=30)
    except Exception as e:  # noqa: BLE001
        print(f"      Lloyd's RDS summary: skipped (SQL connect failed: {e})")
        return None

    import warnings as _w

    def _read(view):
        if not view:
            return None
        try:
            with _w.catch_warnings():
                _w.simplefilter("ignore")   # pandas' non-SQLAlchemy warning
                return pd.read_sql(f"SELECT * FROM {view}", cn)
        except Exception as e:  # noqa: BLE001
            print(f"      Lloyd's RDS summary: view {view} unreadable ({e})")
            return None

    out = {"rds": _read(cfg.get("rds_view")),
           "space_weather": _read(cfg.get("space_weather_view"))}
    n1 = 0 if out["rds"] is None else len(out["rds"])
    n2 = 0 if out["space_weather"] is None else len(out["space_weather"])
    print(f"      Lloyd's RDS summary: {n1} RDS row(s), {n2} bus-type row(s) read")
    if not n1 and not n2:
        return None
    return out
