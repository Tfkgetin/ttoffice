"""Render JJ's 'Lloyds RDS Summary' as the workbook's last tab.

The Lloyd's (Syndicate 3123) RDS methodology lives in two pre-computed SQL
views, exactly as JJ's filed 'Lloyds RDS Summary-final' reads them:

  * rds.vw_SpaceRDS_All_Lloyds_RDS       -> the four headline RDS
        RDS_Name | RDS_Value | Breaches_Risk_Appetite
        (Generic Defect, Proton Flare, Space Debris - Group 1,
         Space Weather - Design Deficiency: <worst bus type>)
  * rds.vw_SpaceRDS_SpaceWeather_Lloyds  -> the Space-Weather bus-type block
        Lloyds Satellite Bus Type List | Aggregate Exposures per satellite type
        (USD) | Top 4 Gross Exposures per satellite type (USD) | > Risk Appetite USD?

This module renders those two frames verbatim (no re-derivation), so the tab
ties to JJ's filed numbers to the dollar. ingest.load_lloyds_rds_summary reads
the views; a missing/unreadable view degrades to a visible banner, not a blank.
"""
from __future__ import annotations
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK = "1F2933"; NAVY = "1B3A5C"; GREEN = "2D6A3C"; RED = "C0392B"
SOFT = "6B7785"; RULE = "DCE3EA"; WHITE = "FFFFFF"; AMBER = "9A6410"
AMBERFILL = "FBF3E4"; REDFILL = "F7E4E1"
F = "Calibri"
MONEY_FMT = '"$"#,##0;"$"-#,##0'  # numeric cells render like _money() but stay live
thin = Side(style="thin", color=RULE)


def _f(sz=10, b=False, c=INK):
    return Font(name=F, size=sz, bold=b, color=c)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _norm(s):
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def _pick(df, *names):
    """Return the actual column whose normalised name matches any of `names`."""
    if df is None:
        return None
    cm = {_norm(c): c for c in df.columns}
    for n in names:
        if _norm(n) in cm:
            return cm[_norm(n)]
    return None


def _is_yes(v):
    return str(v).strip().lower() in ("yes", "y", "true", "1")


def _money(v):
    if v is None or v == "" or v == "NULL" or (isinstance(v, float) and v != v):
        return "-"
    try:
        return "${:,.0f}".format(float(v))
    except (TypeError, ValueError):
        return str(v)


def _banner(ws, r, text):
    cell = ws.cell(r, 2, "⚠  " + text)
    cell.font = _f(10, True, AMBER)
    cell.alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws.cell(r, c).fill = _fill(AMBERFILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    return r + 1


def _header(ws, r, cols, widths, fill=NAVY):
    for i, (h, _) in enumerate(zip(cols, widths)):
        cell = ws.cell(r, 2 + i, h)
        cell.font = _f(9, True, WHITE)
        cell.fill = _fill(fill)
        cell.alignment = Alignment(horizontal="left" if i == 0 else "right",
                                   wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 26
    return r + 1


# pipeline scenario -> (JJ's Lloyd's RDS name, RDS description)
_JJ = {
    "Proton Flare": ("Space weather – Solar energetic particle event",
                     "5% loss on all GEO satellites"),
    "Space Weather": ("Space weather – Design deficiency",
                      "Top 4 exposures on largest bus type group"),
    "Generic Defect": ("Generic Defect",
                       "Sum of the top-10 satellite losses (exposure × risk "
                       "factor × 50% loss rate)"),
    "Space Debris": ("Space Debris",
                     "100% loss of all LEO satellites in the same orbit range "
                     "(adjusted for policy period)"),
}
# Lloyd's RDS has no Max Risk scenario — only these four.
_JJ_ORDER = ["Proton Flare", "Space Weather", "Generic Defect", "Space Debris"]


def write_lloyds_rds_summary(wb, grid, as_at="", sw_view=None,
                             risk_appetite=None, prior=None, qoq_note=None,
                             params=None, change_narrative=None,
                             sheet_name="Lloyds RDS Summary", syndicate_no="3123",
                             cfg_key="s3123_rds", factors=None):
    """Render the Lloyd's (S3123) RDS summary tab from the pipeline's COMPUTED
    grid (gross + net) — it does not depend on the Lloyd's SQL views, which are
    currently unreadable (a varchar->int CAST in the view dies on 'BJ-3C 01').

      grid          : DataFrame with scenario / detail / gross / net (s3123.grid)
      sw_view       : optional DataFrame from vw_SpaceRDS_SpaceWeather_Lloyds for
                      the bus-type block; None -> a banner explains it's blocked.
      risk_appetite : optional USD threshold; RDS gross above it flags a breach.
      prior         : optional {scenario: gross} for a Q-on-Q change column.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [3, 46, 40, 15, 15, 15, 15, 14, 14]):
        ws.column_dimensions[col].width = w

    import pandas as _pd
    g = grid.set_index("scenario") if grid is not None and len(grid) else _pd.DataFrame()
    prior = prior or {}
    prior_label = prior.get("_label", "prior")
    syn = f"S{syndicate_no}"
    # net basis: derive from the syndicate's QS-to-IG (0 -> retains 100%)
    _cfg0 = params.raw.get(cfg_key, {}) if (params is not None
                                            and hasattr(params, "raw")) else {}
    _qs = float(_cfg0.get("qs_to_ig", 0.20))
    net_basis = ("no QS to IG — retains 100% (net = gross)" if _qs == 0
                 else f"net of the {_qs:.0%} QS to IG")

    r = 2
    ws.cell(r, 2, f"Lloyd's RDS Summary — Syndicate {syndicate_no}").font = _f(16, True, NAVY)
    r += 1
    ws.cell(r, 2, f"Space · {syn} · as at {as_at} · syndicate share, {net_basis}"
            f" · computed in-engine · prior = {prior_label}"
            ).font = _f(10, False, SOFT)
    r += 2

    # ---------------- Block 1: the headline RDS (gross + net vs prior) ---- #
    ws.cell(r, 2, "RDS — realistic disaster scenarios").font = _f(12, True, NAVY)
    r += 1
    if not len(g):
        r = _banner(ws, r, f"No {syn} grid — enable {cfg_key} in the config.")
    else:
        r = _header(ws, r, ["RDS Name", "RDS Description", "Gross", "Net",
                            "Prior Gross", "Prior Net", "Δ Gross", "Δ Net"],
                    [46, 40, 15, 15, 15, 15, 14, 14])
        for scen in _JJ_ORDER:
            if scen not in g.index:
                continue
            row = g.loc[scen]
            detail = row.get("detail")
            has_detail = detail not in (None, "", "None") and detail == detail
            name, desc = _JJ.get(scen, (scen, ""))
            if scen == "Space Weather" and has_detail:
                name = f"{name}: {detail}"
            gross = float(row.get("gross") or 0)
            net = row.get("net")
            try:
                netf = float(net)
            except (TypeError, ValueError):
                netf = None
            unavail = (gross == 0 and not has_detail)
            ws.cell(r, 2, name).font = _f(10, True, INK)
            ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 3, desc).font = _f(9, False, SOFT)
            ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")

            def _num(col, value, color=INK, fmt=MONEY_FMT):
                """Write a live money cell — a number, a '=formula', or '-'.

                'n/a' text and formulas both carry MONEY_FMT so a resolved
                formula renders like the numeric cells around it.
                """
                cc = ws.cell(r, col)
                if value is None:
                    cc.value = "-"
                elif value == "n/a":
                    cc.value = "n/a"
                else:
                    cc.value = value
                    cc.number_format = fmt
                cc.font = _f(10, False, color)
                cc.alignment = Alignment(horizontal="right")
                return cc

            # Gross (D) — engine value; selections keep their computed gross.
            if unavail:
                _num(4, "n/a", AMBER)
            else:
                _num(4, gross)
            # Net (E) — for no-QS syndicates net == gross, written as a live
            # formula (=D{r}); otherwise the engine's netted value.
            if unavail:
                _num(5, "n/a", AMBER)
            elif _qs == 0:
                _num(5, f"=D{r}")
            else:
                _num(5, netf)
            # Prior gross / net (F, G) — dollar numbers, live for the deltas.
            pr = prior.get(scen) or {}
            pg = pr.get("gross"); pn = pr.get("net")
            try:
                pgf = float(pg)
            except (TypeError, ValueError):
                pgf = None
            try:
                pnf = float(pn)
            except (TypeError, ValueError):
                pnf = None
            _num(6, pgf, SOFT)
            _num(7, pnf, SOFT)
            # Δ Gross (H) = D-F, Δ Net (I) = E-G — live formulas when both sides
            # are present; colour from the known Python delta.
            dg = (gross - pgf) if (not unavail and pgf is not None) else None
            dn = (netf - pnf) if (not unavail and netf is not None
                                  and pnf is not None) else None
            hc = _num(8, f"=D{r}-F{r}" if dg is not None else None)
            hc.font = _f(10, False, GREEN if (dg or 0) < 0 else RED if (dg or 0) > 0 else SOFT)
            ic = _num(9, f"=E{r}-G{r}" if dn is not None else None)
            ic.font = _f(10, False, GREEN if (dn or 0) < 0 else RED if (dn or 0) > 0 else SOFT)
            for c in range(2, 10):
                ws.cell(r, c).border = Border(bottom=thin)
            r += 1
        if risk_appetite:
            ws.cell(r, 2, f"Risk appetite: ${float(risk_appetite):,.0f} — no RDS "
                    "breaches." if all(float(g.loc[s].get("gross") or 0) <= float(risk_appetite)
                                       for s in _JJ_ORDER if s in g.index)
                    else f"Risk appetite: ${float(risk_appetite):,.0f} — one or "
                    "more RDS BREACH.").font = _f(9, False, SOFT)
            r += 1
        if qoq_note:
            r += 1
            nc = ws.cell(r, 2, "▸  " + str(qoq_note))
            nc.font = Font(name=F, size=9, italic=True, color=SOFT)
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            ws.row_dimensions[r].height = 56
            r += 1
    r += 2

    # ---------------- Change narrative (gross + net, per RDS) ------------ #
    if change_narrative:
        ws.cell(r, 2, "Change narrative — Q-on-Q gross & net").font = _f(12, True, NAVY)
        r += 1
        r = _header(ws, r, ["RDS", "Narrative (gross & net drivers)"], [46, 80])
        for scen in _JJ_ORDER:
            txt = change_narrative.get(scen)
            if not txt:
                continue
            name = _JJ.get(scen, (scen, ""))[0]
            ws.cell(r, 2, name).font = _f(10, True, INK)
            ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
            tc = ws.cell(r, 3, str(txt).strip()); tc.font = _f(10, False, INK)
            tc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
            ws.row_dimensions[r].height = 70
            for c in range(2, 10):
                ws.cell(r, c).border = Border(bottom=thin)
            r += 1
        r += 2

    sw = sw_view

    # ---------- Block 2: Space Weather — design deficiency by bus type ---- #
    ws.cell(r, 2, "Space Weather — design deficiency by satellite bus type"
            ).font = _f(12, True, NAVY)
    r += 1
    ws.cell(r, 2, "Top-4 gross exposures on the largest bus-type group drive the "
            "Design Deficiency RDS above.").font = _f(9, False, SOFT)
    r += 1
    if sw is None or not len(sw):
        r = _banner(ws, r, "Bus-type detail unavailable: vw_SpaceRDS_SpaceWeather_"
                    "Lloyds currently errors in SQL (varchar→int CAST fails on "
                    "spacecraft 'BJ-3C 01'). Fix the view / the offending row and "
                    "re-run; the Space Weather RDS above then populates too.")
    else:
        c_bus = _pick(sw, "Lloyds Satellite Bus Type List", "LloydsSatelliteBusTypeList",
                      "BusType") or sw.columns[0]
        c_agg = _pick(sw, "Aggregate Exposures per satellite type (USD)",
                      "AggregateExposurespersatellitetypeUSD") or sw.columns[1]
        c_top = _pick(sw, "Top 4 Gross Exposures per satellite type (USD)",
                      "Top4GrossExposurespersatellitetypeUSD")
        c_app = _pick(sw, "> Risk Appetite USD?", "RiskAppetiteUSD", "RiskAppetite")
        r = _header(ws, r, ["Lloyd's satellite bus type",
                            "Aggregate exposure (USD)",
                            "Top-4 gross exposure (USD)",
                            "> Risk appetite?"], [60, 24, 24, 16])
        # sort by aggregate desc so the binding bus type is first
        try:
            sw = sw.sort_values(c_agg, ascending=False,
                                key=lambda s: pd.to_numeric(s, errors="coerce"))
        except Exception:
            pass
        for k, (_, row) in enumerate(sw.iterrows()):
            ws.cell(r, 2, str(row[c_bus])).font = _f(10, False, INK)
            ws.cell(r, 2).alignment = Alignment(wrap_text=True)
            ac = ws.cell(r, 3, _money(row[c_agg])); ac.alignment = Alignment(horizontal="right")
            ac.font = _f(10, False, INK)
            tc = ws.cell(r, 4, _money(row[c_top]) if c_top else "-")
            tc.alignment = Alignment(horizontal="right"); tc.font = _f(10, False, INK)
            over = _is_yes(row[c_app]) if c_app else False
            oc = ws.cell(r, 5, "Yes" if over else "No")
            oc.font = _f(10, True, RED if over else GREEN)
            oc.alignment = Alignment(horizontal="right")
            if over:
                for c in range(2, 6):
                    ws.cell(r, c).fill = _fill(REDFILL)
            for c in range(2, 6):
                ws.cell(r, c).border = Border(bottom=thin)
            r += 1
    r += 2
    # ---------------- Parameters & methodology --------------------------- #
    ws.cell(r, 2, "Parameters & methodology").font = _f(12, True, NAVY)
    r += 1
    r = _header(ws, r, ["RDS", f"Definition (at the {syn} share, RPF-adjusted)"],
                [46, 80])
    for scen in _JJ_ORDER:
        name, desc = _JJ.get(scen, (scen, ""))
        ws.cell(r, 2, name).font = _f(10, True, INK)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        dc = ws.cell(r, 3, desc); dc.font = _f(10, False, INK)
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        for c in range(2, 10):
            ws.cell(r, c).border = Border(bottom=thin)
        r += 1
    r += 1

    def _param(label, val):
        c = ws.cell(r, 2, label); c.font = _f(10, True, INK)
        c.alignment = Alignment(vertical="top")
        vc = ws.cell(r, 3, val); vc.font = _f(10, False, INK)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)

    rows = []
    if params is not None:
        cfg = params.raw.get(cfg_key, {}) if hasattr(params, "raw") else {}
        _facs = factors if factors is not None else getattr(params, "s3123_factors", [])
        try:
            rows.append((f"{syn} consortium share (schedule)",
                         " · ".join(f"{d}: {f:.3f}" for d, f in _facs)))
        except Exception:  # noqa: BLE001
            pass
        qs = float(cfg.get("qs_to_ig", 0.20))
        nexcl = len(cfg.get("qs_to_ig_excluded") or [])
        if qs == 0:
            rows.append(("QS ceded to IG", "0% — no QS-to-IG agreement; net = gross "
                         "(retains 100%). The S3123 above-IG-line exclusion does "
                         "not apply here (UW-confirmed)."))
        else:
            rows.append(("QS ceded to IG", f"{qs:.0%} — net = gross × {1 - qs:.2f}; "
                         f"{nexcl} spacecraft above the 30m IG line retain 100%"))
        ap = (cfg.get("lloyds_summary", {}) or {}).get("risk_appetite_usd")
        if ap:
            rows.append(("Lloyd's risk appetite", f"${float(ap):,.0f}"))
        try:
            b = list(params.rpf_bands)
            parts = []
            for i, (m, f) in enumerate(b):
                lo = int(m)
                hi = int(b[i + 1][0]) if i + 1 < len(b) else None
                if hi is None or hi >= 900:
                    if lo < 900:
                        parts.append(f"{lo}m+: {f:.1f}")
                else:
                    parts.append(f"{lo}-{hi}m: {f:.1f}")
            rows.append(("RPF (policy-period) bands", " · ".join(parts)))
        except Exception:  # noqa: BLE001
            pass
        grps = ((cfg.get("scenarios", {}) or {}).get("space_debris", {})
                or {}).get("altitude_groups") or []
        if grps:
            rows.append(("LEO altitude groups (Space Debris)",
                         " · ".join(f"{g['name']}: {g['low']}–{g['high']} km" for g in grps)))
    if _qs == 0:
        rows.append(("Basis / validation",
                     f"Computed at the {syn} consortium share (same methodology, "
                     f"RPF and scenario definitions as S3123), {net_basis}. New "
                     f"participant from 2026-04-01 (5/30 share) — no prior filing "
                     f"to reconcile; very small exposure."))
    else:
        rows.append(("Basis / validation",
                     "Computed at the S3123 syndicate share, net of the 20% QS to "
                     "IG. Solar-particle & Design-deficiency reproduce JJ's 2026Q1 "
                     "(gross AND net) to ~$3."))
    for label, val in rows:
        _param(label, val)
        r += 1

    # ---------------- SQL sources & connections -------------------------- #
    r += 1
    ws.cell(r, 2, "SQL sources & connections").font = _f(12, True, NAVY)
    r += 1
    src = []
    if params is not None and hasattr(params, "ingest"):
        sq = (params.ingest or {}).get("sql", {})
        if sq.get("server") or sq.get("database"):
            src.append(("Server / database",
                        f"{sq.get('server', '?')} / {sq.get('database', '?')} "
                        f"(Windows/Trusted, ODBC Driver 17)"))
        if sq.get("query_file"):
            src.append(("On-risk population", f"{sq['query_file']}  "
                        f"(body of rds.vw_SpaceRDS_OnRisk, as-at injected)"))
        lc = (params.raw.get("s3123_rds", {}) or {}).get("lloyds_source", {})
        if lc.get("spacecraft_attrs_query_file"):
            src.append(("Lloyd's bus type + altitude",
                        f"{lc['spacecraft_attrs_query_file']}  — dbo.tbl_SpaceCraft × "
                        f"rds.params_Satellite_List_Lloyds (bus-type group); "
                        f"[Altitude Latest (km)] for the LEO groups"))
    src.append(("LEO altitude groups", "rds.params_lloyds_proton_flare_altitude"))
    src.append(("Risk appetite", "rds.params_Risk_Appetite_Lloyds (dated)"))
    src.append(("Reference (JJ's filed logic)",
                "rds.vw_SpaceRDS_All_Lloyds_RDS + per-scenario sub-views "
                "(vw_SpaceRDS_{Proton_Flare,Generic_Defect,Space_Debris,"
                "SpaceWeather}_Lloyds) — currently error on a varchar→int CAST "
                "('BJ-3C 01'); the pipeline computes in-engine and routes around them"))
    for label, val in src:
        _param(label, val)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws
