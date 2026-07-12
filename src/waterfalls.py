"""Movement waterfalls — quarter-on-quarter exposure & loss bridges (live-linked).

House translation of the Aviation War "VF waterfalls" into the Space RDS idiom.
AW bridges a single headline PML between two snapshots and slices the movement by
region / country / sector with an entity selector. The Space analogues are
scenario / entity / orbit / manufacturer, so this builds:

  WF · Exposure Bridge
    • portfolio in-force bridge  Opening → +New → −Run-off → ±Reval → Closing
      (the 1:1 AW analogue — every step links to the Portfolio bridge cells)
    • in-force composition by entity / orbit / manufacturer / scenario
      (each value is a live SUMIFS against Per Layer — nothing hard-coded)
    • per-slice opening→closing bridges when a prior per-layer snapshot exists.
      Closing is a live SUMIFS on Per Layer; the three prior-derived movements
      (+New / −Run-off / ±Reval — the prior snapshot is the one input not on a
      current tab) are consolidated in a labelled anchor block, and every bridge
      cell references it (Opening is derived), so no bridge cell is baked.

  WF · Loss Movement
    • all-scenario levels (current vs prior net) with the worst-case scenario flagged
    • all-scenario movement (Δ net by scenario, diverging)
    • worst-case deep-dive: gross AND net bridge, isolating the
      reinsurance / cession effect (ΔGross − ΔNet) — FIBL is the IGR receiver, so
      part of any move is structural cession, not new risk
    • per entity × scenario net table
    every loss figure links to the Changes "By scenario (net/gross)" blocks.

Charts use the stacked-bar floating-waterfall technique used on the Netting tab
(Base invisible + Down + Up + Total); openpyxl cannot emit AW's native chartEx
waterfall and this renders identically everywhere. Scenarios are single-event RDS
views and are NEVER summed.

Linking is resolved at build time by scanning the Portfolio / Changes tabs for
their labels, and by the fixed Per Layer column layout (entity=C, orbit=G,
manufacturer=H, per_sc=L). Portfolio and Changes are written before this runs;
Per Layer is written after, but the formulas resolve at recalc when every tab
exists.

Wiring (excel_report.write_results, next to _changes):

    from . import waterfalls
    waterfalls.build_movement_waterfalls(wb, per_layer, grid, changes, params,
                                         prior_layers=(changes or {}).get("prior_layers"))
"""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter as _L
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint

from .excel_report import (
    F_HDR, F_CELL, F_KEY, F_TOTAL, F_OK, F_SUB, F_SECT,
    FILL_HDR, FILL_ALT, MONEY, PCT, NAVY, GREEN, SOFT,
    _title, _section, _hdr, _cell,
)

RED = "C0392B"
NAVY_LT = "5B86B0"
AMBER = "9A6410"
MONEY_M = '#,##0,,"m";(#,##0,,"m");"–"'

SCEN_ORDER = ["Proton Flare", "Space Weather", "Generic Defect",
              "Space Debris", "Max Risk"]
ENT_ORDER = ["FIHL", "FUL", "FIBL", "FIID"]
GROUP = "FIHL"

# Per Layer fixed column layout (written by excel_report._per_layer).
PL = "'Per Layer'"
PL_ENT, PL_ORB, PL_MFR, PL_PSC = "C", "G", "H", "L"


# --------------------------------------------------------------------------- #
#  Source-cell locator (link, don't hard-code)                                  #
# --------------------------------------------------------------------------- #
def _find_row(ws, text, col=2, contains=False):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and ((text in v) if contains else (v == text)):
            return r
    return None


def _locate(wb, n_layers):
    """Resolve the source cells the bridges link to. Returns None on the first
    run (no Changes tab), in which case callers fall back to Python values."""
    ref = {"pl_first": 2, "pl_last": 1 + int(n_layers)}
    ref["pl_total"] = f"SUM({PL}!${PL_PSC}${ref['pl_first']}:${PL_PSC}${ref['pl_last']})"

    if "Portfolio" in wb.sheetnames:
        p = wb["Portfolio"]
        # Renewal-aware Portfolio bridge: Opening / +New / ↻Renewals / −Lapsed /
        # ±Revaluation / Closing (labels are lower-case 'run-off', so match on
        # 'Lapsed'; add the Renewals step so the waterfall ties).
        ref["p_open"] = _find_row(p, "Opening", contains=True)
        ref["p_new"] = _find_row(p, "New business", contains=True)
        ref["p_ren"] = _find_row(p, "Renewals", contains=True)
        ref["p_run"] = _find_row(p, "Lapsed", contains=True)
        ref["p_reval"] = _find_row(p, "Revaluation", contains=True)
        ref["p_close"] = _find_row(p, "Closing", contains=True)

    if "Changes" not in wb.sheetnames:
        return ref
    c = wb["Changes"]
    ref["c_new"] = _find_row(c, "New exposure")
    ref["c_drop"] = _find_row(c, "Dropped exposure")
    ref["c_move"] = _find_row(c, "Net move on continuing layers")
    for key, header in (("net_base", "By scenario (net)"),
                        ("gross_base", "By scenario (gross)")):
        h = _find_row(c, header)
        base = None
        if h:
            for r in range(h, h + 8):
                if c.cell(row=r, column=2).value == "FIHL":
                    base = r
                    break
        ref[key] = {ent: (base + 5 * i) if base else None
                    for i, ent in enumerate(ENT_ORDER)}
    return ref


def _fcell(ws, r, c, value, alt=False, fmt=MONEY, bold=False):
    """Write a formula (or value) keeping house styling + number format."""
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = F_TOTAL if bold else F_CELL
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    if alt:
        cell.fill = FILL_ALT
    return cell


def _scen_ref(ref, block, entity, scenario, col):
    """A1 ref into a Changes scenario block. col in {'cur','prior','delta'}."""
    base = ref.get(block, {}).get(entity)
    if base is None:
        return None
    col_letter = {"cur": "D", "prior": "E", "delta": "F"}[col]
    return f"Changes!{col_letter}{base + SCEN_ORDER.index(scenario)}"


# --------------------------------------------------------------------------- #
#  Core floating-waterfall block (formula scaffold + stacked chart)            #
# --------------------------------------------------------------------------- #
def _waterfall_block(ws, r0, c0, heading, steps, reported_close=None,
                     chart_h=7.8, chart_w=15.5, tie=True, chart=True):
    """steps: (label, detail, source, kind). source may be a number OR an
    '=...' formula string; both render and tie out the same way."""
    if heading:
        ws.cell(row=r0, column=c0, value=heading).font = F_SECT
    heads = ["Step", "Detail", "Source ($)", "Cumul", "Base", "Down", "Up", "Total"]
    hr = r0 + 1
    for j, h in enumerate(heads):
        c = ws.cell(row=hr, column=c0 + j, value=h)
        c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center" if j > 1 else "left")
    ws.column_dimensions[_L(c0)].width = 20
    ws.column_dimensions[_L(c0 + 1)].width = 15
    ws.column_dimensions[_L(c0 + 2)].width = 19
    ws.column_dimensions[_L(c0 + 3)].width = 17
    for off in (4, 5, 6, 7):
        ws.column_dimensions[_L(c0 + off)].width = 16
    SRC, CUM = _L(c0 + 2), _L(c0 + 3)
    first = hr + 1; r = first
    for i, (label, detail, value, kind) in enumerate(steps):
        is_total = (kind == "total"); is_last = (i == len(steps) - 1)
        ws.cell(row=r, column=c0, value=label).font = (F_TOTAL if is_total else F_CELL)
        if detail is not None:
            dc = ws.cell(row=r, column=c0 + 1, value=detail); dc.font = F_SUB
            dc.alignment = Alignment(horizontal="right")
        if is_total and is_last:
            sc = ws.cell(row=r, column=c0 + 2, value=f"={CUM}{r - 1}")
        else:
            sc = ws.cell(row=r, column=c0 + 2, value=value)
        sc.number_format = MONEY; sc.font = (F_TOTAL if is_total else F_CELL)
        cum = (f"={SRC}{r}" if is_total else f"={CUM}{r - 1}+{SRC}{r}")
        ws.cell(row=r, column=c0 + 3, value=cum).number_format = MONEY
        if is_total:
            ws.cell(row=r, column=c0 + 4, value=0)
            ws.cell(row=r, column=c0 + 5, value=0)
            ws.cell(row=r, column=c0 + 6, value=0)
            ws.cell(row=r, column=c0 + 7, value=f"={CUM}{r}")
        else:
            ws.cell(row=r, column=c0 + 4, value=f"=IF({SRC}{r}>0,{CUM}{r - 1},{CUM}{r})")
            ws.cell(row=r, column=c0 + 5, value=f"=IF({SRC}{r}<0,-{SRC}{r},0)")
            ws.cell(row=r, column=c0 + 6, value=f"=IF({SRC}{r}>0,{SRC}{r},0)")
            ws.cell(row=r, column=c0 + 7, value=0)
        for off in (3, 4, 5, 6, 7):
            cc = ws.cell(row=r, column=c0 + off)
            cc.number_format = MONEY; cc.font = F_CELL
        r += 1
    last = r - 1
    if chart:
        ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
        ch.gapWidth = 45; ch.height = chart_h; ch.width = chart_w
        ch.y_axis.numFmt = MONEY_M; ch.y_axis.title = "$m"
        ch.x_axis.delete = False; ch.y_axis.delete = False
        cats = Reference(ws, min_col=c0, min_row=first, max_row=last)
        for off in (4, 5, 6, 7):
            ch.add_data(Reference(ws, min_col=c0 + off, min_row=hr, max_row=last),
                        titles_from_data=True)
        ch.set_categories(cats)
        sb, sd, su, st = ch.series
        sb.graphicalProperties = GraphicalProperties(); sb.graphicalProperties.noFill = True
        sb.graphicalProperties.line.noFill = True
        sd.graphicalProperties = GraphicalProperties(solidFill=RED)
        su.graphicalProperties = GraphicalProperties(solidFill=GREEN)
        st.graphicalProperties = GraphicalProperties(solidFill=NAVY)
        ch.legend = None
        ws.add_chart(ch, f"{_L(c0)}{last + 2}")
    if tie and reported_close is not None:
        tr = last + 2
        ws.column_dimensions[_L(c0 + 9)].width = 16
        ws.column_dimensions[_L(c0 + 10)].width = 2
        ws.column_dimensions[_L(c0 + 11)].width = 16
        ws.cell(row=tr, column=c0 + 9, value="Reported closing").font = F_SUB
        rc = ws.cell(row=tr, column=c0 + 11, value=reported_close)
        rc.number_format = MONEY; rc.font = F_CELL
        ws.cell(row=tr + 1, column=c0 + 9, value="Bridge closing").font = F_SUB
        bc = ws.cell(row=tr + 1, column=c0 + 11, value=f"={CUM}{last}")
        bc.number_format = MONEY; bc.font = F_CELL
        ws.cell(row=tr + 2, column=c0 + 9, value="Tie-out").font = F_KEY
        cv = ws.cell(row=tr + 2, column=c0 + 11,
                     value=f'=IF(ROUND({CUM}{last}-{_L(c0 + 11)}{tr},0)=0,"OK","CHECK")')
        cv.font = F_OK; cv.alignment = Alignment(horizontal="center")
    return last + 16


# --------------------------------------------------------------------------- #
#  Composition panel: ranked table (live SUMIFS) + horizontal bar              #
# --------------------------------------------------------------------------- #
def _comp_panel(ws, r0, c0, title, items, total_ref, chart=False):
    """items: list of (member, exposure_formula, sort_value). Ranked desc.
    A ranked table with a Share column reads clearly on its own, so the bar
    chart is OFF by default (it just duplicated the table) — pass chart=True to
    restore it for a panel that genuinely benefits from the visual."""
    items = [(k, f, float(v)) for k, f, v in items if v]
    items.sort(key=lambda x: x[2], reverse=True)
    ws.cell(row=r0, column=c0, value=title).font = F_SECT
    hr = r0 + 1
    for j, h in enumerate(["", "Exposure ($)", "Share"]):
        c = ws.cell(row=hr, column=c0 + j, value=h); c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="left" if j == 0 else "center")

    def _setw(col, w):
        d = ws.column_dimensions[_L(col)]; d.width = max(d.width or 0, w)
    _setw(c0, 30); _setw(c0 + 1, 16); _setw(c0 + 2, 12)
    first = hr + 1
    for i, (k, formula, _v) in enumerate(items):
        alt = i % 2 == 0
        _cell(ws, first + i, c0, k, alt=alt)
        _fcell(ws, first + i, c0 + 1, formula, alt=alt, fmt=MONEY)
        exp = f"{_L(c0 + 1)}{first + i}"
        _fcell(ws, first + i, c0 + 2, f"=IFERROR({exp}/({total_ref}),0)", alt=alt, fmt=PCT)
    last = first + len(items) - 1
    # in-cell data bars on the Share column give the visual ranking without a
    # separate chart taking half the page.
    from openpyxl.formatting.rule import DataBarRule
    ws.conditional_formatting.add(
        f"{_L(c0 + 2)}{first}:{_L(c0 + 2)}{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="8FB0CF"))
    if not chart:
        return r0 + len(items) + 3
    ch = BarChart(); ch.type = "bar"; ch.height = max(4.2, 0.6 * len(items) + 1.2)
    ch.width = 11; ch.legend = None
    ch.x_axis.delete = False; ch.y_axis.delete = False; ch.y_axis.numFmt = MONEY_M
    ch.x_axis.scaling.min = 0
    data = Reference(ws, min_col=c0 + 1, min_row=first, max_row=last)
    cats = Reference(ws, min_col=c0, min_row=first, max_row=last)
    ch.add_data(data, titles_from_data=False); ch.set_categories(cats)
    ch.series[0].graphicalProperties = GraphicalProperties(solidFill=NAVY)
    ws.add_chart(ch, f"{_L(c0 + 4)}{r0 + 1}")
    chart_rows = int(ch.height / 0.5) + 2
    return r0 + max(len(items) + 2, chart_rows) + 1


# --------------------------------------------------------------------------- #
#  Diverging Δ bars (formula values, colour by known sign)                      #
# --------------------------------------------------------------------------- #
def _delta_bars(ws, r0, c0, title, labels, value_refs, signs,
                chart_h=7.0, chart_w=15.0):
    ws.cell(row=r0, column=c0, value=title).font = F_SECT
    hr = r0 + 1
    for j, h in enumerate(["Scenario", "Δ Net ($)"]):
        c = ws.cell(row=hr, column=c0 + j, value=h); c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="left" if j == 0 else "center")
    ws.column_dimensions[_L(c0)].width = 18
    ws.column_dimensions[_L(c0 + 1)].width = 16
    first = hr + 1
    for i, (lab, vref) in enumerate(zip(labels, value_refs)):
        _cell(ws, first + i, c0, lab, alt=(i % 2 == 0))
        _fcell(ws, first + i, c0 + 1, vref, fmt=MONEY)
    last = first + len(labels) - 1
    ch = BarChart(); ch.type = "col"; ch.height = chart_h; ch.width = chart_w
    ch.legend = None; ch.y_axis.numFmt = MONEY_M; ch.y_axis.title = "$m"
    ch.x_axis.delete = False; ch.y_axis.delete = False
    ch.add_data(Reference(ws, min_col=c0 + 1, min_row=first, max_row=last),
                titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=c0, min_row=first, max_row=last))
    s = ch.series[0]
    s.graphicalProperties = GraphicalProperties(solidFill=SOFT)
    s.invertIfNegative = False          # stop Excel whitening negative bars
    for i, d in enumerate(signs):
        if d is None:
            continue
        dp = DataPoint(idx=i, invertIfNegative=False)
        dp.graphicalProperties = GraphicalProperties(
            solidFill=(GREEN if d >= 0 else RED))
        s.data_points.append(dp)
    ws.add_chart(ch, f"{_L(c0)}{last + 2}")
    return last + 16


def _clustered(ws, r0, c0, title, cats, series_specs, chart_h=7.0, chart_w=15.0):
    """series_specs: list of (name, [value_refs], colour)."""
    if title:
        ws.cell(row=r0, column=c0, value=title).font = F_SECT
    hr = r0 + 1
    ws.cell(row=hr, column=c0, value="").fill = FILL_HDR
    ws.column_dimensions[_L(c0)].width = 18
    for j, (nm, _vals, _col) in enumerate(series_specs):
        c = ws.cell(row=hr, column=c0 + 1 + j, value=nm)
        c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[_L(c0 + 1 + j)].width = 15
    first = hr + 1
    for i, cat in enumerate(cats):
        _cell(ws, first + i, c0, cat, alt=(i % 2 == 0))
        for j, (_nm, vals, _col) in enumerate(series_specs):
            _fcell(ws, first + i, c0 + 1 + j, vals[i], fmt=MONEY)
    last = first + len(cats) - 1
    ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
    ch.height = chart_h; ch.width = chart_w
    ch.y_axis.numFmt = MONEY_M; ch.y_axis.title = "$m"
    ch.x_axis.delete = False; ch.y_axis.delete = False
    for j, (_nm, _vals, col) in enumerate(series_specs):
        ch.add_data(Reference(ws, min_col=c0 + 1 + j, min_row=hr, max_row=last),
                    titles_from_data=True)
        ch.series[j].graphicalProperties = GraphicalProperties(solidFill=col)
    ch.set_categories(Reference(ws, min_col=c0, min_row=first, max_row=last))
    ws.add_chart(ch, f"{_L(c0)}{last + 2}")
    return last + 16


# --------------------------------------------------------------------------- #
#  Exposure tab                                                                 #
# --------------------------------------------------------------------------- #
def _exposure_steps_linked(ref):
    pf = "Portfolio"
    steps = [("Opening in-force", "prior", f"={pf}!D{ref['p_open']}", "total"),
             ("+ New business", "written", f"={pf}!D{ref['p_new']}", "auto")]
    if ref.get("p_ren"):
        steps.append(("↻ Renewals", "renewed", f"={pf}!D{ref['p_ren']}", "auto"))
    steps += [
        ("− Lapsed / run-off", "expired", f"={pf}!D{ref['p_run']}", "auto"),
        ("± Revaluation", "continuing", f"={pf}!D{ref['p_reval']}", "auto"),
        ("Closing in-force", "current", None, "total"),
    ]
    return steps


def _sumifs(crit_col, member, first, last):
    return (f"=SUMIFS({PL}!${PL_PSC}${first}:${PL_PSC}${last},"
            f"{PL}!${crit_col}${first}:${crit_col}${last},\"{member}\")")


def _wf_exposure(wb, per_layer, changes, params, ref, prior_layers=None):
    ws = wb.create_sheet("WF · Exposure Bridge")
    _title(ws, "Space RDS — Exposure Movement Waterfall",
           f"On-risk per-S/C signed exposure  ·  "
           f"{(changes or {}).get('prior_as_at', 'prior')} → {params.as_at}  ·  "
           "all entities  ·  live-linked to Portfolio / Per Layer")
    ws.column_dimensions["A"].width = 2

    if not ref or ref.get("p_open") is None:
        _section(ws, 5, "No prior quarter to bridge")
        ws.cell(row=7, column=2, value="Initial automation deployment — the "
                "movement waterfall appears once a prior run is persisted."
                ).font = F_SUB
        return

    r = _waterfall_block(ws, 5, 2, "Portfolio exposure bridge ($m on-risk)",
                         _exposure_steps_linked(ref),
                         reported_close=f"=Portfolio!D{ref['p_close']}")
    L = (changes or {}).get("layers", {})
    if L:
        from . import renewals as _rnw
        sp = _rnw.split_movement(L.get("new"), L.get("dropped"))

        def _x(d):
            return float(pd.to_numeric(d.get("per_sc"), errors="coerce").fillna(0).sum()) \
                if d is not None and len(d) else 0.0
        new_x, lap_x = _x(sp["new_biz"]), _x(sp["lapsed"])
        renewed = _x(sp["ren_new"]) - _x(sp["ren_old"])
        mv = float(L.get("summary", {}).get("net_move", 0.0))
        net = new_x + renewed - lap_x + mv
        ws.cell(row=r, column=2, value=(
            f"Read: the book {'grew' if net >= 0 else 'shrank'} {abs(net) / 1e6:,.1f}m — "
            f"{new_x / 1e6:,.1f}m genuine new vs {lap_x / 1e6:,.1f}m genuine run-off; "
            "roll-forward renewals net out (not counted as run-off). Every step "
            "links to the Portfolio bridge; green = added, red = removed, navy = level."
            )).font = F_SUB
    r += 3

    first, last = ref["pl_first"], ref["pl_last"]
    total = ref["pl_total"]
    r = _section(ws, r, "Closing in-force composition  (live SUMIFS on Per Layer)")
    r = _comp_panel(ws, r, 2, "By entity", [
        (m, _sumifs(PL_ENT, m, first, last), v)
        for m, v in per_layer.groupby("entity")["per_sc"].sum().items()], total) + 1
    if "orbit" in per_layer:
        r = _comp_panel(ws, r, 2, "By orbit", [
            (m, _sumifs(PL_ORB, m, first, last), v)
            for m, v in per_layer.groupby("orbit")["per_sc"].sum().items()], total) + 1
    if "bus_manufacturer" in per_layer:
        top = per_layer.groupby("bus_manufacturer")["per_sc"].sum().sort_values(
            ascending=False)
        items = [(m, _sumifs(PL_MFR, m, first, last), v) for m, v in top.head(6).items()]
        if len(top) > 6:
            covered = "+".join(f"SUMIFS({PL}!${PL_PSC}${first}:${PL_PSC}${last},"
                               f"{PL}!${PL_MFR}${first}:${PL_MFR}${last},\"{m}\")"
                               for m, _v in top.head(6).items())
            items.append(("Other", f"={total}-({covered})", float(top.iloc[6:].sum())))
        r = _comp_panel(ws, r, 2, "By bus manufacturer (top 6)", items, total) + 1
    # scenario base: Proton Flare sees the GEO fleet only; others the full book
    geo = f"=SUMIFS({PL}!${PL_PSC}${first}:${PL_PSC}${last},{PL}!${PL_ORB}${first}:${PL_ORB}${last},\"*GEO*\")"
    allp = f"={total}"
    base_v = float(per_layer["per_sc"].sum())
    geo_v = float(per_layer.loc[per_layer.get("orbit", pd.Series("", index=per_layer.index))
                                .astype(str).str.contains("GEO", na=False), "per_sc"].sum())
    r = _comp_panel(ws, r, 2, "In-scope exposure base by scenario", [
        ("Proton Flare", geo, geo_v), ("Space Weather", allp, base_v),
        ("Generic Defect", allp, base_v), ("Space Debris", allp, base_v),
        ("Max Risk", allp, base_v)], total) + 1

    if prior_layers is not None and len(prior_layers):
        # Per-slice bridges are retained but HIDDEN (user 2026Q2): the portfolio
        # bridge + composition above cover the movement story; the per-entity /
        # per-orbit detail is collapsed out of the main view. Build without charts
        # (they would float over hidden rows), then hide the whole section.
        ws.cell(row=r, column=2, value=(
            "▸  Per-slice movement bridges (by entity / orbit) are hidden below — "
            "unhide the rows to view. Renewal-aware; data live off 'WF · Movement "
            "Data'.")).font = F_SUB
        r += 2
        slice_start = r
        end = _slice_bridges(ws, r, per_layer, prior_layers, changes, first, last,
                             charts=False)
        for i in range(slice_start, (end or slice_start) + 1):
            ws.row_dimensions[i].hidden = True
    else:
        ws.cell(row=r, column=2, value=(
            "Per-slice opening→closing bridges (by entity / orbit / manufacturer) "
            "populate when the pipeline supplies the prior per-layer snapshot; the "
            "prior set is not a workbook tab, so those alone are not cell-linked.")
            ).font = F_SUB


MD = "'WF · Movement Data'"          # hidden helper tab (see _write_movement_data)
MD_SIDE, MD_ENT, MD_ORB, MD_CLS, MD_PSC = "A", "B", "C", "D", "E"


def _write_movement_data(wb, per_layer, prior_layers, changes):
    """Hidden helper tab: every CURRENT and PRIOR layer tagged with a renewal-
    AWARE movement class, so the slice bridges compute each movement as a live
    SUMIFS — renewal-aware, with NO baked numbers.

      current layer -> new / renewed / continuing   (renewed = a renewal of a
                       prior spacecraft, even onto a new programme id)
      prior   layer -> lapsed / renewed / continuing

    Classes come from renewals.split_movement (spacecraft-level pairing), so a
    fleet that renewed onto a new programme id lands in 'renewed', NOT in both
    run-off and new. Returns (first_row, last_row) or None."""
    from . import renewals as _rnw
    lyr = (changes or {}).get("layers", {}) or {}
    sp = _rnw.split_movement(lyr.get("new"), lyr.get("dropped"))

    def _keys(fr):
        return (set(fr["layer_key"].astype(str))
                if fr is not None and len(fr) and "layer_key" in fr.columns else set())
    new_k, rin_k = _keys(sp["new_biz"]), _keys(sp["ren_new"])
    rout_k, lap_k = _keys(sp["ren_old"]), _keys(sp["lapsed"])

    def _cls(lk, side):
        if side == "cur":
            return "new" if lk in new_k else "renewed" if lk in rin_k else "continuing"
        return "lapsed" if lk in lap_k else "renewed" if lk in rout_k else "continuing"

    rows = []
    for df, side in ((per_layer, "cur"), (prior_layers, "prior")):
        if df is None or not len(df) or "layer_key" not in df.columns:
            continue
        d = df.copy(); d["layer_key"] = d["layer_key"].astype(str)
        for _, x in d.iterrows():
            rows.append((side, str(x.get("entity", "")), str(x.get("orbit", "")),
                         _cls(x["layer_key"], side), float(x.get("per_sc") or 0)))
    if not rows:
        return None
    name = "WF · Movement Data"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(["side", "entity", "orbit", "mv_class", "per_sc"])
    for row in rows:
        ws.append(list(row))
    ws.sheet_state = "hidden"
    return 2, len(rows) + 1


def _md_sumifs(first, last, side=None, cls=None, dimcol=None, member=None):
    """A SUMIFS over the Movement Data tab, filtered by side / class / slice."""
    parts = [f"{MD}!${MD_PSC}${first}:${MD_PSC}${last}"]
    if side is not None:
        parts += [f"{MD}!${MD_SIDE}${first}:${MD_SIDE}${last}", f'"{side}"']
    if cls is not None:
        parts += [f"{MD}!${MD_CLS}${first}:${MD_CLS}${last}", f'"{cls}"']
    if dimcol is not None:
        parts += [f"{MD}!${dimcol}${first}:${dimcol}${last}", f'"{member}"']
    return "SUMIFS(" + ",".join(parts) + ")"


def _slice_bridges(ws, r, per_layer, prior_layers, changes, pl_first, pl_last,
                   dims=("entity", "orbit"), charts=True):
    # Manufacturer slice-bridges dropped (user 2026Q2): too many thin per-maker
    # waterfalls; manufacturer concentration is still on the 'By bus manufacturer'
    # composition panel above. Entity / orbit bridges are retained.
    #
    # Every cell here is a live SUMIFS over the hidden 'WF · Movement Data' tab —
    # NOTHING baked — and the split is RENEWAL-AWARE: renewals (even onto a new
    # programme id) show in a '↻ Renewals (reprice)' step, so run-off is only the
    # genuinely-lapsed exposure, matching the portfolio bridge above.
    md = _write_movement_data(ws.parent, per_layer, prior_layers, changes)
    if md is None:
        return r
    first, last = md
    cur = per_layer.copy()
    slices = []
    for dim in dims:
        if dim not in cur:
            continue
        dimcol = MD_ENT if dim == "entity" else MD_ORB
        plcol = PL_ENT if dim == "entity" else PL_ORB
        for m in (cur.groupby(dim)["per_sc"].sum()
                  .sort_values(ascending=False).head(6).index.tolist()):
            slices.append((dim, str(m), dimcol, plcol))

    for dim, m, dimcol, plcol in slices:
        def s(**kw):
            return _md_sumifs(first, last, dimcol=dimcol, member=m, **kw)
        opening = s(side="prior")
        # reported closing ties to the authoritative Per Layer book (not the
        # helper tab), so any classification gap surfaces as a CHECK.
        closing = _sumifs(plcol, m, pl_first, pl_last)[1:]  # strip leading '='
        new = s(side="cur", cls="new")
        ren_in = s(side="cur", cls="renewed")
        ren_out = s(side="prior", cls="renewed")
        runoff = s(side="prior", cls="lapsed")
        cont_cur = s(side="cur", cls="continuing")
        cont_pri = s(side="prior", cls="continuing")
        steps = [("Opening", "prior", f"={opening}", "total"),
                 ("+ New business", "genuine", f"={new}", "auto"),
                 ("↻ Renewals (reprice)", "renewed", f"={ren_in}-{ren_out}", "auto"),
                 ("− Run-off", "lapsed", f"=-{runoff}", "auto"),
                 ("± Reval", "continuing", f"={cont_cur}-{cont_pri}", "auto"),
                 ("Closing", "current", None, "total")]
        label = dim.replace("bus_manufacturer", "manufacturer")
        r = _waterfall_block(ws, r, 2, f"{label}: {m}", steps,
                             reported_close=f"={closing}",
                             chart_h=6.6, chart_w=13, chart=charts) + 1
    return r


# --------------------------------------------------------------------------- #
#  Loss tab                                                                     #
# --------------------------------------------------------------------------- #
def _binding_scenario(scen_df, entity=GROUP):
    block = scen_df[(scen_df["entity"] == entity) & scen_df["cur_net"].notna()]
    if not len(block):
        return None
    return block.loc[block["cur_net"].idxmax(), "scenario"]


def _wf_loss(wb, summary, changes, params, ref):
    ws = wb.create_sheet("WF · Loss Movement")
    _title(ws, "Space RDS — Loss Movement",
           "QoQ change in scenario loss, gross & net  ·  worst-case deep-dive  ·  "
           "per-entity detail below  ·  live-linked to Changes")
    ws.column_dimensions["A"].width = 2

    scen = None if not changes else changes.get("scenarios")
    if scen is None or not len(scen) or not ref or ref.get("net_base", {}).get("FIHL") is None:
        _section(ws, 5, "No prior quarter to compare")
        ws.cell(row=7, column=2, value="Initial automation deployment — loss "
                "movement appears once a prior run is persisted.").font = F_SUB
        return

    scen = scen.copy()
    bind = _binding_scenario(scen, GROUP)

    cats_flag = [f"{s}  \u25c6" if s == bind else s for s in SCEN_ORDER]
    has_gross = ref.get("gross_base", {}).get(GROUP) is not None

    # ONE movement chart — \u0394 gross AND net by scenario. The tab's whole point
    # is the QoQ change; the levels (prior/current) live in the entity table
    # below, so we don't repeat them as extra level charts.
    ndel = [f"={_scen_ref(ref, 'net_base', GROUP, s, 'delta')}" for s in SCEN_ORDER]
    r = _section(ws, 5, f"{GROUP} loss movement by scenario \u2014 \u0394 vs prior")
    if has_gross:
        gdel = [f"={_scen_ref(ref, 'gross_base', GROUP, s, 'delta')}" for s in SCEN_ORDER]
        r = _clustered(ws, r, 2, "", cats_flag,
                       [("\u0394 Gross", gdel, NAVY_LT), ("\u0394 Net", ndel, NAVY)])
    else:
        signs = []
        for s in SCEN_ORDER:
            row = scen[(scen["entity"] == GROUP) & (scen["scenario"] == s)]
            signs.append(_num(row["d_net"].iloc[0]) if len(row) else None)
        r = _delta_bars(ws, r, 2, "", SCEN_ORDER, ndel, signs)
    ws.cell(row=r, column=2, value=(
        f"\u25c6 worst-case scenario ({bind}). Positive = loss grew vs prior, "
        "negative = shrank. Scenarios are single-event views, never summed \u2014 "
        "the worst one sets capital.")).font = F_SUB
    r += 2

    if bind:
        nb = ref["net_base"][GROUP] + SCEN_ORDER.index(bind)
        gb = ref.get("gross_base", {}).get(GROUP)
        r = _section(ws, r, f"Worst-case deep-dive — {GROUP} {bind}")
        net_steps = [("Prior Q", str(changes.get("prior_as_at", "prior")),
                      f"=Changes!E{nb}", "total"),
                     (f"\u0394 {bind}", None, f"=Changes!F{nb}", "auto"),
                     ("Current Q", str(params.as_at), None, "total")]
        _waterfall_block(ws, r, 2, "Net loss ($m)", net_steps,
                         reported_close=f"=Changes!D{nb}", chart_h=6.6, chart_w=12)
        gr = None
        if gb is not None:
            gr = gb + SCEN_ORDER.index(bind)
            gr_steps = [("Prior Q", str(changes.get("prior_as_at", "prior")),
                         f"=Changes!E{gr}", "total"),
                        (f"\u0394 {bind}", None, f"=Changes!F{gr}", "auto"),
                        ("Current Q", str(params.as_at), None, "total")]
            _waterfall_block(ws, r, 14, "Gross loss ($m)", gr_steps,
                             reported_close=f"=Changes!D{gr}", chart_h=6.6,
                             chart_w=12, tie=False)
        r += 1 + len(net_steps) + 16
        if gr is not None:
            ws.cell(row=r, column=2,
                    value="Reinsurance / cession effect  (\u0394gross \u2212 \u0394net)").font = F_KEY
            _fcell(ws, r, 5, f"=Changes!F{gr}-Changes!F{nb}", fmt=MONEY)
            ws.cell(row=r + 1, column=2, value=(
                "The part of the worst-case gross move absorbed by the "
                "reinsurance structure rather than reaching the net account.")).font = F_SUB
            r += 3

    r = _section(ws, r, "QoQ loss by entity \u00d7 scenario ($) \u2014 gross AND net  "
                        "(links to Changes)")
    r = _hdr(ws, r, 2, ["Entity", "Scenario", "Cur Gross", "Prior Gross",
                        "\u0394 Gross", "Cur Net", "Prior Net", "\u0394 Net",
                        "\u0394 % Net"],
             [10, 18, 14, 14, 14, 14, 14, 14, 9])
    k = 0
    for ent in ENT_ORDER:
        base = ref["net_base"].get(ent)
        gbase = ref.get("gross_base", {}).get(ent)
        if base is None:
            continue
        for j, scn in enumerate(SCEN_ORDER):
            rr = base + j; alt = k % 2 == 0
            _cell(ws, r, 2, ent, alt=alt, bold=(j == 0))
            _cell(ws, r, 3, scn, alt=alt)
            # Gross (links to the Changes 'By scenario (gross)' block when present)
            if gbase is not None:
                gr = gbase + j
                _fcell(ws, r, 4, f"=Changes!D{gr}", alt=alt, fmt=MONEY)
                _fcell(ws, r, 5, f"=Changes!E{gr}", alt=alt, fmt=MONEY)
                _fcell(ws, r, 6, f"=Changes!F{gr}", alt=alt, fmt=MONEY)
            else:
                for cc in (4, 5, 6):
                    _cell(ws, r, cc, None, alt=alt)
            # Net
            _fcell(ws, r, 7, f"=Changes!D{rr}", alt=alt, fmt=MONEY)
            _fcell(ws, r, 8, f"=Changes!E{rr}", alt=alt, fmt=MONEY)
            _fcell(ws, r, 9, f"=Changes!F{rr}", alt=alt, fmt=MONEY)
            _fcell(ws, r, 10, f"=IFERROR(Changes!F{rr}/Changes!E{rr},0)", alt=alt, fmt=PCT)
            r += 1; k += 1


# --------------------------------------------------------------------------- #
#  Entry point                                                                  #
# --------------------------------------------------------------------------- #
def build_movement_waterfalls(wb, per_layer, summary, changes, params,
                              prior_layers=None):
    ref = _locate(wb, len(per_layer))
    _wf_exposure(wb, per_layer, changes, params, ref, prior_layers=prior_layers)
    _wf_loss(wb, summary, changes, params, ref)


# --------------------------------------------------------------------------- #
#  Helpers                                                                       #
# --------------------------------------------------------------------------- #
def _num(v):
    try:
        return float(v) if v is not None and v == v else None
    except (TypeError, ValueError):
        return None
