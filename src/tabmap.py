"""Tab-dependency map → embedded as the 'Map' worksheet.

The diagram is defined here in code (NODES / EDGES), rendered to SVG, and
rasterised to PNG for embedding. Rendering prefers cairosvg (pure-Python); if
that is not installed it falls back to the PNG bundled in assets/ so the
pipeline always produces the tab. To refresh the bundled asset after a
structure change, render build_svg() to assets/Space_RDS_tabmap.png.
"""
import html
import os
import tempfile
import warnings

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.worksheet.properties import PageSetupProperties

INK = "#1F2933"; SOFT = "#6B7785"; RULE = "#DCE3EA"; CREAM = "#EEF3F8"
WHITE = "#FFFFFF"; GREEN = "#2D6A3C"
SEC = {"A": "#1B3A5C", "B": "#3E6C8C", "C": "#5A6B7D", "D": "#97A1AC"}
SECNAME = {"A": "Headline", "B": "Movement",
           "C": "Exposure & structure", "D": "Basis & assurance"}
TXT = {"A": WHITE, "B": WHITE, "C": WHITE, "D": INK}

W, H = 1480, 980
NW, NH = 158, 48

# id -> (label, section, x, y)
NODES = {
    "Parameters": ("Parameters", "D", 60, 430),
    "Per Layer": ("Per Layer", "C", 300, 430),
    "Summary": ("Summary", "A", 565, 120),
    "Portfolio": ("Portfolio", "C", 565, 330),
    "Space Weather": ("Space Weather", "C", 565, 470),
    "Changes": ("Changes", "B", 565, 610),
    "Control": ("Control", "D", 565, 760),
    "Executive Summary": ("Executive Summary", "A", 880, 95),
    "Netting Waterfalls": ("Netting Waterfalls", "C", 880, 235),
    "Chart Data": ("Chart Data", "D", 880, 470),
    "WF · Exposure Bridge": ("WF · Exposure Bridge", "B", 880, 590),
    "WF · Loss Movement": ("WF · Loss Movement", "B", 880, 720),
    "Charts": ("Charts", "A", 1180, 470),
}
# (from, to, kind)  kind: 'live' | 'engine'
EDGES = [
    ("Parameters", "Per Layer", "live"),
    ("Per Layer", "Summary", "engine"),
    ("Per Layer", "Portfolio", "live"),
    ("Per Layer", "Space Weather", "live"),
    ("Per Layer", "Control", "live"),
    ("Per Layer", "Executive Summary", "live"),
    ("Summary", "Executive Summary", "live"),
    ("Summary", "Netting Waterfalls", "live"),
    ("Summary", "Chart Data", "live"),
    ("Changes", "Portfolio", "live"),
    ("Changes", "WF · Loss Movement", "live"),
    ("Portfolio", "WF · Exposure Bridge", "live"),
    ("Space Weather", "Chart Data", "live"),
    ("Chart Data", "Charts", "live"),
]


def _cy(n):
    return NODES[n][3] + NH / 2


def _node(nid):
    lab, sec, x, y = NODES[nid]
    col = SEC[sec]; tc = TXT[sec]; t = html.escape(lab)
    return (f'<rect x="{x}" y="{y}" width="{NW}" height="{NH}" rx="7" fill="{col}"/>'
            f'<text x="{x+16}" y="{y+NH/2+4.5}" font-size="13.5" font-weight="600" '
            f'fill="{tc}" font-family="Arial">{t}</text>'
            f'<text x="{x+NW-13}" y="{y+15}" text-anchor="middle" font-size="9.5" '
            f'font-weight="700" fill="{tc}" opacity="0.85" font-family="Arial">{sec}</text>')


def _edge(a, b, kind):
    x1, y1 = NODES[a][2] + NW, _cy(a)
    x2, y2 = NODES[b][2], _cy(b)
    dx = (x2 - x1) * 0.45
    if kind == "engine":
        stroke, dash, mk = GREEN, 'stroke-dasharray="6 5"', "url(#arrG)"
    else:
        stroke, dash, mk = SOFT, "", "url(#arr)"
    return (f'<path d="M {x1} {y1} C {x1+dx} {y1}, {x2-dx} {y2}, {x2-2} {y2}" '
            f'fill="none" stroke="{stroke}" stroke-width="1.7" {dash} marker-end="{mk}"/>')


def build_svg():
    edges = "\n".join(_edge(*e) for e in EDGES)
    nodes = "\n".join(_node(n) for n in NODES)
    heads = [("INPUT", 150), ("BACKBONE", 379), ("CORE GRIDS &amp; STRUCTURE", 644),
             ("DERIVED OUTPUTS", 959), ("VISUALS", 1259)]
    heads_svg = "\n".join(
        f'<text x="{x}" y="62" text-anchor="middle" font-size="11" font-weight="700" '
        f'fill="{SOFT}" letter-spacing="1.5" font-family="Arial">{t}</text>'
        for t, x in heads)
    cover = (
        f'<rect x="60" y="78" width="1360" height="26" rx="6" fill="{CREAM}" '
        f'stroke="{RULE}" stroke-width="1" stroke-dasharray="4 4"/>'
        f'<rect x="70" y="84" width="22" height="14" rx="3" fill="{SEC["A"]}"/>'
        f'<text x="81" y="94.5" text-anchor="middle" font-size="9" font-weight="700" '
        f'fill="{WHITE}" font-family="Arial">A</text>'
        f'<text x="102" y="95" font-size="12" font-weight="600" fill="{INK}" '
        f'font-family="Arial">Cover</text>'
        f'<text x="148" y="95" font-size="11.5" fill="{SOFT}" font-family="Arial">'
        f'— Contents hyperlinks jump to every tab (navigation only, not a data feed)</text>')
    standalone = (f'<text x="60" y="838" font-size="11" font-weight="700" fill="{SOFT}" '
                  f'letter-spacing="1" font-family="Arial">STANDALONE · provenance '
                  f'&amp; assurance (no live feed in/out)</text>')
    sx = 60
    for lab, sec in [("Max Risk", "C"), ("Python Adjustments", "D"),
                     ("Methodology", "D"), ("Audit", "D")]:
        col = SEC[sec]; tc = TXT[sec]; w = 160
        standalone += (
            f'<rect x="{sx}" y="852" width="{w}" height="38" rx="7" fill="{col}"/>'
            f'<text x="{sx+16}" y="875" font-size="12.5" font-weight="600" fill="{tc}" '
            f'font-family="Arial">{html.escape(lab)}</text>'
            f'<text x="{sx+w-13}" y="867" text-anchor="middle" font-size="9" '
            f'font-weight="700" fill="{tc}" opacity="0.85" font-family="Arial">{sec}</text>')
        sx += w + 14
    lx, ly = 60, 928
    seclegend = (f'<text x="60" y="{ly-15}" font-size="11" font-weight="700" fill="{SOFT}" '
                 f'letter-spacing="1" font-family="Arial">CATEGORY  ·  box colour = '
                 f'workbook section</text>')
    for s in ["A", "B", "C", "D"]:
        seclegend += (
            f'<rect x="{lx}" y="{ly-9}" width="18" height="18" rx="4" fill="{SEC[s]}"/>'
            f'<text x="{lx+25}" y="{ly+5}" font-size="11.5" fill="{INK}" '
            f'font-family="Arial">{s} · {html.escape(SECNAME[s])}</text>')
        lx += 215
    arrowlegend = (
        f'<line x1="1118" y1="916" x2="1156" y2="916" stroke="{SOFT}" stroke-width="1.7" '
        f'marker-end="url(#arr)"/>'
        f'<text x="1163" y="920" font-size="11" fill="{INK}" font-family="Arial">'
        f'live Excel formula / chart link</text>'
        f'<line x1="1118" y1="936" x2="1156" y2="936" stroke="{GREEN}" stroke-width="1.7" '
        f'stroke-dasharray="6 5" marker-end="url(#arrG)"/>'
        f'<text x="1163" y="940" font-size="11" fill="{INK}" font-family="Arial">'
        f'aggregated by the Python engine</text>')
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="{W}" height="{H}" viewBox="0 0 {W} {H}" font-family="Arial">
<defs>
  <marker id="arr" markerWidth="9" markerHeight="9" refX="7" refY="3" orient="auto"><path d="M0,0 L7,3 L0,6 Z" fill="{SOFT}"/></marker>
  <marker id="arrG" markerWidth="9" markerHeight="9" refX="7" refY="3" orient="auto"><path d="M0,0 L7,3 L0,6 Z" fill="{GREEN}"/></marker>
</defs>
<rect width="{W}" height="{H}" fill="{WHITE}"/>
<text x="60" y="38" font-size="22" font-weight="700" fill="{SEC['A']}">Space RDS — how the tabs feed each other</text>
<line x1="60" y1="48" x2="1420" y2="48" stroke="{RULE}" stroke-width="1"/>
{heads_svg}
{cover}
{edges}
{nodes}
{standalone}
<line x1="60" y1="903" x2="1420" y2="903" stroke="{RULE}" stroke-width="1"/>
{seclegend}
{arrowlegend}
</svg>'''


_ASSET = os.path.join(os.path.dirname(__file__), "assets", "Space_RDS_tabmap.png")


def _render_png(out_path):
    """Render the map SVG to PNG.

    Tries cairosvg first; if that is unavailable or fails (common on Windows
    where the native cairo library is hard to install), falls back to
    playwright (headless Chromium). Returns (ok, reason).
    """
    svg_bytes = build_svg().encode("utf-8")

    # --- attempt 1: cairosvg ---
    try:
        import cairosvg
        cairosvg.svg2png(bytestring=svg_bytes,
                         write_to=out_path, output_width=W * 2, output_height=H * 2)
        return True, None
    except ImportError:
        cairo_err = "cairosvg not installed"
    except Exception as e:
        cairo_err = f"cairosvg failed ({e!r})"

    # --- attempt 2: playwright (headless Chromium) ---
    try:
        from playwright.sync_api import sync_playwright
        import base64, glob as _glob
        svg_b64 = base64.b64encode(svg_bytes).decode()
        html = (f'<html><body style="margin:0;padding:0;">'
                f'<img src="data:image/svg+xml;base64,{svg_b64}" '
                f'width="{W}" height="{H}"/></body></html>')
        # Auto-detect a pre-installed chromium (e.g. cloud / CI environments).
        # Falls back to letting playwright find its own managed browser.
        _exe_candidates = sorted(_glob.glob(
            os.path.join(os.environ.get("PLAYWRIGHT_BROWSERS_PATH", ""),
                         "chromium*/chrome-linux/chrome")))
        _launch_kw = {"executable_path": _exe_candidates[-1]} if _exe_candidates else {}
        with sync_playwright() as p:
            browser = p.chromium.launch(**_launch_kw)
            page = browser.new_page(viewport={"width": W * 2, "height": H * 2})
            page.set_content(html)
            page.wait_for_load_state("networkidle")
            page.screenshot(path=out_path, full_page=False)
            browser.close()
        return True, None
    except ImportError:
        pw_err = "playwright not installed"
    except Exception as e:
        pw_err = f"playwright failed ({e!r})"

    return False, f"no renderer succeeded — cairo: {cairo_err}; playwright: {pw_err}"


def add_tab_map(wb, sheet="Map", after="Cover"):
    """Add/refresh the 'Map' worksheet with the tab-dependency diagram."""
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws["B2"] = "Tab map — how the tabs feed each other"
    ws["B2"].font = Font(name="Calibri", size=13, bold=True, color="1B3A5C")

    png = None
    tmp = os.path.join(tempfile.gettempdir(), "spacerds_tabmap.png")
    ok, reason = _render_png(tmp)
    if ok:
        png = tmp
    elif os.path.exists(_ASSET):
        png = _ASSET                          # fall back to the bundled image
        reason = None
    if png:
        img = XLImage(png)
        # Fit within a landscape A4/letter page (≈1050 px wide at 96 DPI).
        # Source PNG is 2× the SVG dimensions for crispness; display at 1× SVG size
        # scaled to page width so no horizontal scrolling is needed.
        scale = 1050 / W          # ~0.71 — keeps aspect ratio
        img.width  = int(W * scale)
        img.height = int(H * scale)
        # Widen column B so the image is not clipped by a narrow default column.
        ws.column_dimensions["B"].width = 130
        ws.row_dimensions[4].height = img.height * 0.75  # pts ≈ px × 0.75
        ws.add_image(img, "B4")
    else:
        # Never ship a silently-blank tab: explain why on the sheet itself and
        # raise a warning so the build pipeline/log shows the real cause.
        why = reason or "no renderer available and no bundled fallback image"
        ws["B4"] = (f"⚠ Map diagram could not be embedded — {why}. "
                    "Install cairosvg (pip install cairosvg) or add "
                    "assets/Space_RDS_tabmap.png next to this module.")
        ws["B4"].font = Font(name="Calibri", size=11, italic=True, color="B00020")
        warnings.warn(f"add_tab_map: map image not embedded — {why}",
                      RuntimeWarning, stacklevel=2)

    # position right after the Cover tab
    if after in wb.sheetnames:
        sheets = wb._sheets
        s = sheets.pop(sheets.index(ws))
        sheets.insert(wb.sheetnames.index(after) + 1, s)
    return ws


def fix_picture_ids(xlsx_path):
    """Repair image drawing ids after a LibreOffice round-trip.

    LibreOffice's recalc/save can renumber an embedded picture's cNvPr id to 0,
    which Excel treats as corrupt and silently drops (the image shows blank).
    Call this on any workbook that has been saved by LibreOffice. Idempotent.
    """
    import zipfile
    import shutil
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}
    changed = False
    next_id = 900  # unique replacement id, bumped per picture (Excel rejects dup ids)
    for n in names:
        if n.startswith("xl/drawings/drawing") and n.endswith(".xml"):
            txt = data[n].decode("utf-8")
            if ("<xdr:pic>" in txt or "<pic>" in txt) and 'cNvPr id="0"' in txt:
                while 'cNvPr id="0"' in txt:
                    txt = txt.replace('cNvPr id="0"', f'cNvPr id="{next_id}"', 1)
                    next_id += 1
                data[n] = txt.encode("utf-8")
                changed = True
    if changed:
        tmp = xlsx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                zout.writestr(n, data[n])
        shutil.move(tmp, xlsx_path)
    return changed
