"""Workbook-level organisation pass for the Space RDS pack.

Turns a flat 18-tab workbook into a navigable, sectioned senior deliverable:

  • tabs are coloured by section in a dark→light gradient (headline → assurance),
    so the structure of the pack is legible from the tab strip alone;
  • the Cover's Contents is rebuilt grouped into the same four sections, with a
    matching colour key, one-line purposes, and a clickable hyperlink per tab;
  • a short reading note frames how the pack should be interpreted.

Wiring (excel_report.write_results, after the sheet order is sorted):

    from . import organise
    organise.organise(wb)

`organise` is idempotent and order-independent — it reads the section map below,
so re-running simply refreshes colours and the Contents.
"""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import WorksheetProperties

_FB = "Calibri"
INK = "1F2933"; SOFT = "6B7785"; RULE = "DCE3EA"; NAVY = "1B3A5C"; WHITE = "FFFFFF"

# Section map — order mirrors the workbook tab order. (letter, name, colour, tabs)
SECTIONS = [
    ("A", "Headline", "1B3A5C", [
        ("Cover", None),
        ("Map", "How the tabs feed each other — data-flow map"),
        ("Executive Summary",
         "Worst-case by entity, ranked scenarios, watch items"),
        ("Summary", "Full netting cascade per scenario & entity (gross \u2192 net)"),
        ("Charts", "Net risk shape, reinsurance benefit, concentration, RI support"),
    ]),
    ("B", "Quarter-on-quarter movement", "3E6C8C", [
        ("Changes", "Layer adds / drops and scenario moves vs the prior run"),
        ("WF \u00b7 Exposure Bridge",
         "Opening \u2192 closing exposure bridge and composition"),
        ("WF \u00b7 Loss Movement",
         "Net loss movement and worst-case deep-dive"),
    ]),
    ("C", "Exposure & structure", "5A6B7D", [
        ("Portfolio", "Exposure concentration by entity, orbit, manufacturer"),
        ("Netting Waterfalls", "Gross-to-net cascade per entity"),
        ("Space Weather", "Worst-case scenario detail by bus manufacturer"),
        ("Max Risk", "Largest-layer scenario detail"),
        ("Per Layer", "Full layer-level data backbone"),
    ]),
    ("D", "Basis & assurance", "97A1AC", [
        ("Python Adjustments", "Automated add / remove with documented reasons"),
        ("Control", "Run controls and reconciliation"),
        ("Methodology", "Assumptions and scenario definitions"),
        ("Parameters", "Treaty terms and scenario factors"),
        ("Chart Data", "Formula-linked backing tables for Charts"),
        ("Audit", "Provenance and run trail"),
    ]),
]

READING_NOTE = ("How to read: scenarios are single-event RDS views and are never "
                "summed \u2014 the worst-case scenario sets capital. Figures are USD, "
                "per-spacecraft signed exposure. Tab colours match the sections above.")


def apply_tab_colours(wb):
    for _letter, _name, colour, tabs in SECTIONS:
        for tab, _desc in tabs:
            if tab in wb.sheetnames:
                wb[tab].sheet_properties.tabColor = colour


def _find_contents_row(ws):
    for r in range(1, (ws.max_row or 1) + 1):
        if ws.cell(row=r, column=2).value == "Contents":
            return r
    return None


def build_contents(wb, cover="Cover"):
    if cover not in wb.sheetnames:
        return
    ws = wb[cover]
    hdr = _find_contents_row(ws)
    if hdr is None:
        hdr = (ws.max_row or 1) + 2
        c = ws.cell(row=hdr, column=2, value="Contents")
        c.font = Font(name=_FB, size=12, bold=True, color=NAVY)

    # clear any previous contents body (cols A..E below the header)
    lo, hi = hdr + 1, (ws.max_row or hdr) + 40
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= lo:
            ws.unmerge_cells(str(rng))
    for r in range(lo, hi):
        for cc in range(1, 6):
            cell = ws.cell(row=r, column=cc)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

    ws.column_dimensions["A"].width = 3.2
    r = hdr + 2
    for letter, name, colour, tabs in SECTIONS:
        # left colour chip with the section letter
        chip = ws.cell(row=r, column=1, value=letter)
        chip.fill = PatternFill("solid", start_color=colour)
        chip.font = Font(name=_FB, size=9, bold=True, color=WHITE)
        chip.alignment = Alignment(horizontal="center", vertical="center")
        sh = ws.cell(row=r, column=2, value=name)
        sh.font = Font(name=_FB, size=10.5, bold=True, color=colour)
        for cc in range(2, 6):
            ws.cell(row=r, column=cc).border = Border(
                bottom=Side(style="thin", color=RULE))
        r += 1
        for tab, desc in tabs:
            if tab == cover or tab not in wb.sheetnames:
                continue
            link = ws.cell(row=r, column=2,
                           value=f'=HYPERLINK("#\'{tab}\'!A1","{tab}")')
            link.font = Font(name=_FB, size=10, bold=True, color=NAVY, underline="single")
            link.alignment = Alignment(indent=1)
            if desc:
                dc = ws.cell(row=r, column=3, value=desc)
                dc.font = Font(name=_FB, size=9, color=INK)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            r += 1
        r += 1

    note = ws.cell(row=r, column=2, value=READING_NOTE)
    note.font = Font(name=_FB, size=8.5, italic=True, color=SOFT)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=5)
    ws.row_dimensions[r].height = 14
    ws.row_dimensions[r + 1].height = 14


def freeze_data_tabs(wb):
    """Freeze the header on the long, scrolling data tab(s)."""
    if "Per Layer" in wb.sheetnames and not wb["Per Layer"].freeze_panes:
        wb["Per Layer"].freeze_panes = "A2"


def organise(wb):
    apply_tab_colours(wb)
    build_contents(wb)
    freeze_data_tabs(wb)
